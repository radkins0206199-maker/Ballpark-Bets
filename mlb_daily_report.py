import math
import json
import os
import pickle
import datetime
import time
import io
import requests
import pandas as pd
import numpy as np
import pytz
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.isotonic import IsotonicRegression
from xgboost import XGBClassifier
import smtplib
from email.message import EmailMessage

MLB_BASE = "https://statsapi.mlb.com/api/v1"

# ── Always use Eastern Time for dates ─────────────────────────
# Render runs in UTC. Games run until ~1 AM ET on the West Coast.
# Using UTC causes tomorrow's date to appear after ~8 PM ET.
ET = pytz.timezone("America/New_York")

def today_et():
    """Return today's date in Eastern Time as YYYY-MM-DD string."""
    return datetime.datetime.now(ET).strftime("%Y-%m-%d")

def yesterday_et():
    """Return yesterday's date in Eastern Time as YYYY-MM-DD string."""
    yd = datetime.datetime.now(ET) - datetime.timedelta(days=1)
    return yd.strftime("%Y-%m-%d")

CURRENT_SEASON = datetime.datetime.now(ET).year


# ── Phase 4.1: Elo Rating System ──────────────────────────────
# Provides a continuously updated team quality estimate.
# More responsive than win% — updates after every game.
# K-factor adjusted for margin of victory (blowouts move Elo more).
# Pre-season prior: all teams start at 1500, regressed from last season.

_elo_cache = {}   # team_name → current elo rating
_elo_loaded = False

ELO_BASE        = 1500
ELO_K           = 20     # base K-factor
ELO_HOME_ADV    = 30     # home field advantage in Elo points
ELO_REGRESSION  = 0.33   # off-season regression toward mean

def _elo_expected(rating_a: float, rating_b: float) -> float:
    """Expected win probability for team A vs team B."""
    return 1.0 / (1.0 + 10 ** ((rating_b - rating_a) / 400.0))

def _elo_k_factor(run_diff: int) -> float:
    """K-factor multiplier based on margin of victory."""
    abs_diff = abs(run_diff)
    if abs_diff <= 1: return 1.0
    if abs_diff <= 3: return 1.2
    if abs_diff <= 5: return 1.4
    return 1.6

def load_elo_ratings():
    """Load current Elo ratings from Supabase elo_ratings table."""
    global _elo_cache, _elo_loaded
    if _elo_loaded:
        return _elo_cache

    try:
        import psycopg2
        db_url = os.environ.get("DATABASE_URL", "")
        if not db_url:
            return {}
        conn = psycopg2.connect(db_url, sslmode='require', connect_timeout=10)
        cur  = conn.cursor()
        cur.execute("SELECT team_name, elo_rating FROM elo_ratings")
        rows = cur.fetchall()
        cur.close(); conn.close()
        _elo_cache  = {row[0]: float(row[1]) for row in rows}
        _elo_loaded = True
        print(f"  [Elo] Loaded {len(_elo_cache)} team ratings")
    except Exception as e:
        print(f"  [Elo] Load failed: {e} — using base ratings")
        _elo_cache  = {}
        _elo_loaded = True
    return _elo_cache

def get_elo(team_name: str) -> float:
    """Get current Elo rating for a team. Returns base if not found."""
    ratings = load_elo_ratings()
    return ratings.get(team_name, ELO_BASE)

def update_elo_after_resolve(home_team: str, away_team: str,
                              home_score: int, away_score: int):
    """
    Update Elo ratings after a game resolves.
    Called from auto_fill_results after actual winner is determined.
    """
    global _elo_cache

    ratings = load_elo_ratings()
    home_elo = ratings.get(home_team, ELO_BASE)
    away_elo = ratings.get(away_team, ELO_BASE)

    # Expected outcome (home has advantage)
    home_expected = _elo_expected(home_elo + ELO_HOME_ADV, away_elo)
    away_expected = 1.0 - home_expected

    # Actual outcome
    home_actual = 1.0 if home_score > away_score else 0.0
    away_actual = 1.0 - home_actual

    # K-factor adjusted for margin
    run_diff = home_score - away_score
    k = ELO_K * _elo_k_factor(run_diff)

    # Update ratings
    new_home = round(home_elo + k * (home_actual - home_expected), 1)
    new_away = round(away_elo + k * (away_actual - away_expected), 1)

    # Save to cache and DB
    _elo_cache[home_team] = new_home
    _elo_cache[away_team] = new_away

    try:
        import psycopg2
        db_url = os.environ.get("DATABASE_URL", "")
        if db_url:
            conn = psycopg2.connect(db_url, sslmode='require', connect_timeout=10)
            cur  = conn.cursor()
            for team, rating in [(home_team, new_home), (away_team, new_away)]:
                cur.execute("""
                    INSERT INTO elo_ratings (team_name, elo_rating, last_updated)
                    VALUES (%s, %s, NOW())
                    ON CONFLICT (team_name) DO UPDATE SET
                        elo_rating   = EXCLUDED.elo_rating,
                        last_updated = NOW()
                """, (team, rating))
            conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"  [Elo] Update failed: {e}")


def elo_win_probability(home_team: str, away_team: str) -> float:
    """
    Convert Elo ratings to win probability for the home team.
    Includes home field advantage.
    """
    home_elo = get_elo(home_team)
    away_elo = get_elo(away_team)
    prob = _elo_expected(home_elo + ELO_HOME_ADV, away_elo)
    return round(prob, 4)


class CalibratedModel:
    """Wraps a fitted classifier with an isotonic regression calibrator.
    Trained on a held-out calibration set so predict_proba() outputs
    well-calibrated probabilities without expensive cross-validation."""
    def __init__(self, base_model, calibrator):
        self.base_model = base_model
        self.calibrator = calibrator

    def predict_proba(self, X):
        raw = self.base_model.predict_proba(X)[:, 1]
        cal = np.clip(self.calibrator.predict(raw), 0.0, 1.0)
        return np.column_stack([1.0 - cal, cal])

    def predict(self, X):
        return (self.predict_proba(X)[:, 1] >= 0.5).astype(int)


def _train_calibrated(estimator, X_tr, y_tr, X_cal, y_cal):
    """Fit estimator on training split, then fit isotonic calibrator on cal split."""
    estimator.fit(X_tr, y_tr)
    raw_cal = estimator.predict_proba(X_cal)[:, 1]
    iso = IsotonicRegression(out_of_bounds="clip")
    iso.fit(raw_cal, y_cal)
    return CalibratedModel(estimator, iso)


def _train_calibrated_crossfit(estimator_factory, X, y, n_folds=5):
    """
    Phase 4.5 Step 7 — cross-fitted calibration.

    The Aug 11 diagnostic showed the calibration curve was badly broken:
    predictions of 0.845 won only 0.545 of the time, and the model emitted
    probabilities as extreme as 0.020 and 0.953. Root cause was fitting an
    isotonic calibrator on a ~60-row holdout. Isotonic is non-parametric and
    on 60 points will map cleanly-separating regions to 0.0/1.0.

    Cross-fitting fixes this by giving the calibrator EVERY row's
    out-of-fold prediction (~1000 instead of ~60) with no leakage:
      - split into k time-ordered folds
      - for each fold: fit on the other k-1, predict this fold
      - concatenate all out-of-fold predictions, fit ONE calibrator on them
      - refit the base estimator on all data, pair with that calibrator

    Isotonic is retained (rather than Platt) because the observed curve is
    non-monotone and flat in the low range — a shape a two-parameter sigmoid
    cannot represent. At n~1000 with cross-fitting, isotonic is appropriate.
    """
    import numpy as np
    n = len(X)
    if n < 100:
        # Too small to cross-fit safely — fall back to a simple split.
        split = int(n * 0.75)
        est = estimator_factory()
        return _train_calibrated(est, X.iloc[:split], y.iloc[:split],
                                 X.iloc[split:], y.iloc[split:])

    fold_edges = [int(n * i / n_folds) for i in range(n_folds + 1)]
    oof_pred = np.zeros(n, dtype=float)

    for k in range(n_folds):
        lo, hi = fold_edges[k], fold_edges[k + 1]
        test_idx  = list(range(lo, hi))
        train_idx = [i for i in range(n) if i < lo or i >= hi]
        if len(train_idx) < 30 or not test_idx:
            continue
        est_k = estimator_factory()
        est_k.fit(X.iloc[train_idx], y.iloc[train_idx])
        oof_pred[lo:hi] = est_k.predict_proba(X.iloc[test_idx])[:, 1]

    iso = IsotonicRegression(out_of_bounds="clip", y_min=0.02, y_max=0.98)
    iso.fit(oof_pred, np.asarray(y))

    final_est = estimator_factory()
    final_est.fit(X, y)
    return CalibratedModel(final_est, iso), oof_pred

# In-process cache so we don't hit the API twice for the same team/pitcher
_team_stats_cache       = {}
_pitcher_stats_cache    = {}
_standings_cache        = {}
_savant_cache           = {}
_vegas_cache            = {}
_kalshi_cache           = {}
_lineup_cache           = {}
_bullpen_fatigue_cache  = {}
_weather_cache          = {}

SCRIPT_DIR        = os.path.dirname(os.path.abspath(__file__))
TRACKER_CSV       = os.path.join(SCRIPT_DIR, "results_tracker.csv")
RETRAIN_FLAG      = os.path.join(SCRIPT_DIR, ".last_retrain_date")
MODEL_CACHE       = os.path.join(SCRIPT_DIR, "ml_model_cache.pkl")
PREDICTIONS_JSON  = os.path.join(SCRIPT_DIR, "daily_predictions.json")
TRACKER_JSON      = os.path.join(SCRIPT_DIR, "results_tracker.json")

# Core display columns + ML feature columns stored for auto-recalibration
TRACKER_COLS = [
    "Date", "Game Time", "Home Team", "Away Team", "Predicted Winner",
    "Home Win %", "Model Edge", "Confidence", "Actual Winner",
    # ML training features
    "team_wpct", "opp_wpct", "team_rpg", "opp_rpg",
    "pitcher_era", "opp_pitcher_era", "bullpen_era", "opp_bullpen_era",
    "park_factor", "temp", "wind_speed", "wind_dir_out", "home",
    "pitcher_recent_delta", "opp_pitcher_recent_delta",
]

# ============================================================
# A. LIVE DATA FROM MLB STATS API
# ============================================================

def _get(url, params=None):
    try:
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        print(f"  [API] WARNING: {e}")
        return None


def fetch_standings():
    """Return dict of team_id -> (wpct, rpg) from live standings."""
    global _standings_cache
    if _standings_cache:
        return _standings_cache

    data = _get(f"{MLB_BASE}/standings",
                params={"leagueId": "103,104", "season": CURRENT_SEASON, "standingsTypes": "regularSeason"})
    if not data:
        return {}

    result = {}
    for division in data.get("records", []):
        for rec in division.get("teamRecords", []):
            team_id = rec["team"]["id"]
            wpct = float(rec.get("winningPercentage", "0.500"))
            games = rec.get("gamesPlayed", 1) or 1
            runs = rec.get("runsScored", 0) or 0
            rpg = round(runs / games, 2)
            result[team_id] = {"wpct": wpct, "rpg": rpg}

    _standings_cache = result
    return result


# ============================================================
# B. BASEBALL SAVANT / STATCAST DATA
# ============================================================

def fetch_savant_team_batting():
    """
    Aggregate Baseball Savant batting Statcast (hard hit %, barrel rate, xwOBA)
    to the team level by joining player IDs with MLB API roster data.
    Results are cached for the run.
    """
    global _savant_cache
    if _savant_cache:
        return _savant_cache

    # 1. player_id → team_id from MLB API active roster
    players_data = _get(f"{MLB_BASE}/sports/1/players",
                        {"season": CURRENT_SEASON, "gameType": "R"})
    if not players_data:
        return {}
    player_to_team = {}
    for p in players_data.get("people", []):
        pid = p.get("id")
        tid = p.get("currentTeam", {}).get("id")
        if pid and tid:
            player_to_team[pid] = tid

    # 2. team_id → full team name
    teams_data = _get(f"{MLB_BASE}/teams", {"sportId": 1, "season": CURRENT_SEASON})
    team_id_to_name = {}
    if teams_data:
        for t in teams_data.get("teams", []):
            team_id_to_name[t["id"]] = t["name"]

    # 3. Baseball Savant batting leaderboard (player-level)
    try:
        r = requests.get(
            "https://baseballsavant.mlb.com/leaderboard/custom",
            params={
                "year": CURRENT_SEASON, "type": "batter", "filter": "",
                "sort": "4", "sortDir": "desc", "min": "25",
                "selections": "hard_hit_percent,barrel_batted_rate,xwoba",
                "csv": "true",
            },
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=15,
        )
        df = pd.read_csv(io.StringIO(r.text))
    except Exception as e:
        print(f"  [Savant] WARNING: Batting leaderboard failed: {e}")
        return {}

    # 4. Join and aggregate by team
    df["team_id"]   = df["player_id"].map(player_to_team)
    df["team_name"] = df["team_id"].map(team_id_to_name)
    df = df.dropna(subset=["team_name"])

    result = {}
    for team_name, grp in df.groupby("team_name"):
        result[team_name] = {
            "hard_hit":    round(float(grp["hard_hit_percent"].mean()) / 100, 4),
            "barrel_rate": round(float(grp["barrel_batted_rate"].mean()) / 100, 4),
            "xwoba":       round(float(grp["xwoba"].mean()), 4),
        }

    _savant_cache = result
    print(f"  [Savant] Team batting stats loaded for {len(result)} teams")
    return result


def get_savant_pitcher_stats(pitcher_id):
    """
    Fetch Baseball Savant Statcast data for a specific pitcher:
    xwOBA allowed, hard hit % allowed, barrel rate against.
    Returns None if unavailable.
    """
    if not pitcher_id:
        return None
    try:
        r = requests.get(
            "https://baseballsavant.mlb.com/statcast_search/csv",
            params={
                "all": "true", "hfGT": "R|", "hfSea": f"{CURRENT_SEASON}|",
                "player_type": "pitcher",
                "pitchers_lookup[]": pitcher_id,
                "group_by": "name",
                "sort_col": "pitches", "sort_order": "desc", "min_pas": "0",
            },
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=15,
        )
        df = pd.read_csv(io.StringIO(r.text))
        if df.empty:
            return None
        row = df.iloc[0]
        return {
            "xwoba":       float(row.get("xwoba",                  0.320)),
            "hard_hit_pct": float(row.get("hardhit_percent",       35.0)),
            "barrel_rate":  float(row.get("barrels_per_bbe_percent", 7.0)),
        }
    except Exception:
        return None


def xwoba_to_era_adj(xwoba, base_era):
    """
    Blend a pitcher's ERA with a Statcast xwOBA-implied ERA for a more
    predictive quality metric.
      - League avg: xwOBA ≈ 0.315, ERA ≈ 4.20
      - Each 0.020 xwOBA ≈ 0.50 ERA difference
      - Final = 60% ERA + 40% xwOBA-derived ERA
    """
    LEAGUE_XWOBA = 0.315
    LEAGUE_ERA   = 4.20
    xwoba_era = LEAGUE_ERA + (xwoba - LEAGUE_XWOBA) * 25
    return round(base_era * 0.60 + xwoba_era * 0.40, 2)


_FORMULA_CHARS = ("=", "+", "-", "@", "\t", "\r")

# Only these CSV columns contain free-form text from untrusted third-party
# feeds (team names, pitcher names, weather strings).  Numeric columns such
# as "Model Edge" or "Home Win %" are intentionally excluded so that negative
# numeric strings stored as text (e.g. "-3.2") are not prefixed and can still
# be parsed correctly by downstream consumers.
_CSV_TEXT_COLS = frozenset({
    "Date", "Game Time", "Home Team", "Away Team",
    "Predicted Winner", "Confidence", "Actual Winner",
})


def sanitize_cell(val):
    """Neutralize spreadsheet formula injection in untrusted string values.

    Spreadsheet applications (Excel, Google Sheets, LibreOffice) treat cell
    values that begin with '=', '+', '-', '@', TAB, or CR as formulas and
    execute them on open.  This function prefixes any such string with a
    leading single-quote so the application renders it as plain text instead.
    Strings that start with '+' or '-' but parse as valid numbers are left
    untouched because they represent legitimate numeric text, not formulas.
    Non-string values (numbers, None, booleans) are returned unchanged.
    """
    if not isinstance(val, str):
        return val
    if val.startswith(("=", "@", "\t", "\r")):
        return "'" + val
    if val.startswith(("+", "-")):
        try:
            float(val)
            return val  # Legitimate signed-number string; not a formula
        except (ValueError, TypeError):
            return "'" + val
    return val


def _sanitize_dataframe(df):
    """Apply sanitize_cell() to free-form text columns of a DataFrame.

    Only columns listed in _CSV_TEXT_COLS are sanitized.  Numeric columns
    (e.g. 'Model Edge', 'Home Win %') are left untouched so that values such
    as '-3.2' remain parseable as numbers by downstream consumers.
    """
    for col in _CSV_TEXT_COLS:
        if col in df.columns:
            df[col] = df[col].map(lambda v: sanitize_cell(v) if isinstance(v, str) else v)
    return df


def update_results_tracker(games_list):
    """
    Append today's predictions to results_tracker.csv.
    Removes any existing rows for today first (safe to re-run).
    Uses Eastern Time date — not UTC — to prevent midnight date rollover.
    """
    today    = today_et()  # ET date, not UTC
    new_rows = []
    for g in games_list:
        new_rows.append({
            "Date":             today,
            "Home Team":        g.get("Home Team", ""),
            "Away Team":        g.get("Away Team", ""),
            "Predicted Winner": g.get("Predicted Winner", ""),
            "Home Win %":       g.get("Home Win Probability", ""),
            "Confidence":       g.get("Confidence", ""),
            "Actual Winner":    "",
            # ML training features — used for monthly auto-recalibration
            "team_wpct":        g.get("team_wpct", ""),
            "opp_wpct":         g.get("opp_wpct", ""),
            "team_rpg":         g.get("team_rpg", ""),
            "opp_rpg":          g.get("opp_rpg", ""),
            "pitcher_era":      g.get("pitcher_era", ""),
            "opp_pitcher_era":  g.get("opp_pitcher_era", ""),
            "bullpen_era":      g.get("bullpen_era", ""),
            "opp_bullpen_era":  g.get("opp_bullpen_era", ""),
            "park_factor":      g.get("park_factor", ""),
            "temp":             g.get("temp", ""),
            "wind_speed":       g.get("wind_speed", ""),
            "wind_dir_out":     g.get("wind_dir_out", ""),
            "home":             1,
            "pitcher_recent_delta":     g.get("pitcher_recent_delta", 0.0),
            "opp_pitcher_recent_delta": g.get("opp_pitcher_recent_delta", 0.0),
        })
    new_df = pd.DataFrame(new_rows).reindex(columns=TRACKER_COLS, fill_value="")

    if os.path.exists(TRACKER_CSV):
        existing = _safe_read_tracker_csv()
        existing = existing[existing["Date"] != today]
        combined = pd.concat([existing, new_df], ignore_index=True)
    else:
        combined = new_df

    combined = _sanitize_dataframe(combined)
    combined.to_csv(TRACKER_CSV, index=False)
    return combined


# ============================================================
# C. FEATURES 1–6: AUTO-RESULTS, VEGAS, LINEUP, BULLPEN, FORM, RETRAIN
# ============================================================

def auto_fill_results():
    """
    Feature 1 — Auto-fill Results Tracker.
    Fetches final scores from MLB API and writes the actual winner
    into any unresolved rows. Backfills up to 7 past days to recover
    from Render restarts. Uses Eastern Time throughout.
    Also resolves any rows in Supabase that are missing actual_winner.
    """
    today = today_et()

    # ── Supabase-first backfill ────────────────────────────────
    # Even if local CSV is missing, resolve unresolved Supabase rows
    # This ensures missed days are always captured
    try:
        import urllib.request as _req
        sb_url = os.environ.get("SUPABASE_URL", "https://wkxpdmfabiepkfdxbdie.supabase.co")
        sb_key = os.environ.get("SUPABASE_KEY", "")

        # Get unresolved rows from Supabase (actual_winner is null)
        url = f"{sb_url}/rest/v1/results?actual_winner=is.null&select=date,home_team,away_team,predicted_winner&order=date.desc&limit=100"
        req = _req.Request(url)
        req.add_header("apikey", sb_key)
        req.add_header("Authorization", f"Bearer {sb_key}")
        with _req.urlopen(req, timeout=10) as resp:
            unresolved_sb = json.loads(resp.read())

        # Group by date
        dates_to_resolve = {}
        for row in unresolved_sb:
            d = row.get("date", "")
            if d and d < today:
                if d not in dates_to_resolve:
                    dates_to_resolve[d] = []
                dates_to_resolve[d].append(row)

        if dates_to_resolve:
            print(f"  [Tracker] Resolving {len(dates_to_resolve)} missing date(s) from Supabase: {sorted(dates_to_resolve.keys())}")
            for date_str, rows in dates_to_resolve.items():
                data = _get(f"{MLB_BASE}/schedule",
                            {"sportId": 1, "date": date_str, "hydrate": "linescore,team"})
                if not data:
                    continue
                results = {}
                for date_entry in data.get("dates", []):
                    for game in date_entry.get("games", []):
                        state = game.get("status", {}).get("abstractGameState", "")
                        if state != "Final":
                            continue
                        teams = game.get("teams", {})
                        home = teams.get("home", {})
                        away = teams.get("away", {})
                        home_name = home.get("team", {}).get("name", "")
                        away_name = away.get("team", {}).get("name", "")
                        home_score = home.get("score", 0) or 0
                        away_score = away.get("score", 0) or 0
                        if home_score == 0 and away_score == 0:
                            continue
                        winner = home_name if home_score > away_score else away_name
                        results[(home_name, away_name)] = winner

                # Resolve via psycopg2 (parameterized — handles spaces/special chars safely)
                # The REST API path failed on any team name containing a space because
                # the URL wasn't encoded. psycopg2 parameterized queries avoid this entirely.
                if results:
                    try:
                        import psycopg2
                        db_url = os.environ.get("DATABASE_URL", "")
                        if db_url:
                            pconn = psycopg2.connect(db_url, sslmode='require', connect_timeout=10)
                            pcur  = pconn.cursor()
                            for row in rows:
                                key = (row.get("home_team", ""), row.get("away_team", ""))
                                if key in results:
                                    actual = results[key]
                                    predicted = row.get("predicted_winner", "")
                                    correct = "✓ Correct" if actual == predicted else "✗ Wrong"
                                    # Find scores for Elo update
                                    hs = as_ = None
                                    try:
                                        pcur.execute("""
                                            UPDATE results
                                            SET actual_winner=%s, correct=%s, updated_at=NOW()
                                            WHERE date=%s AND home_team=%s AND away_team=%s
                                        """, (actual, correct, date_str, key[0], key[1]))
                                        if pcur.rowcount > 0:
                                            print(f"    [Supabase] Resolved {key[0]} vs {key[1]} on {date_str}: {actual} ({correct})")
                                    except Exception as pe:
                                        print(f"    [Supabase] Patch failed {key}: {pe}")
                            pconn.commit()
                            pcur.close()
                            pconn.close()
                    except Exception as e:
                        print(f"  [Supabase] Backfill resolve failed: {e}")
        else:
            print(f"  [Tracker] All Supabase rows resolved")

        # ── Insert MISSING days from predictions table ─────────
        # Checks last 7 days — if predictions exist but no results row, inserts it
        import datetime
        for days_back in range(1, 8):
            check_date = (datetime.date.today() - datetime.timedelta(days=days_back)).isoformat()
            if check_date >= today:
                continue
            try:
                # Check if results rows exist for this date
                check_url = f"{sb_url}/rest/v1/results?date=eq.{check_date}&select=date&limit=1"
                check_req = _req.Request(check_url)
                check_req.add_header("apikey", sb_key)
                check_req.add_header("Authorization", f"Bearer {sb_key}")
                with _req.urlopen(check_req, timeout=8) as resp:
                    existing = json.loads(resp.read())

                if existing:
                    continue  # results exist for this date, skip

                # No results — check predictions table for this date
                pred_url = f"{sb_url}/rest/v1/predictions?date=eq.{check_date}&select=games,best_bets&limit=1"
                pred_req = _req.Request(pred_url)
                pred_req.add_header("apikey", sb_key)
                pred_req.add_header("Authorization", f"Bearer {sb_key}")
                with _req.urlopen(pred_req, timeout=8) as resp:
                    pred_rows = json.loads(resp.read())

                if not pred_rows:
                    print(f"  [Tracker] No predictions for {check_date} — cannot backfill")
                    continue

                # Get actual scores from MLB API
                score_data = _get(f"{MLB_BASE}/schedule",
                                  {"sportId": 1, "date": check_date, "hydrate": "linescore,team"})
                if not score_data:
                    continue

                # Parse predictions
                raw_games = pred_rows[0].get("games", [])
                if isinstance(raw_games, str):
                    try: raw_games = json.loads(raw_games)
                    except: raw_games = []

                pred_by_key = {(g.get("Home Team",""), g.get("Away Team","")): g for g in raw_games}

                inserted_count = 0
                for date_entry in score_data.get("dates", []):
                    for game in date_entry.get("games", []):
                        if game.get("status", {}).get("abstractGameState") != "Final":
                            continue
                        teams = game.get("teams", {})
                        home = teams.get("home", {})
                        away = teams.get("away", {})
                        home_name = home.get("team", {}).get("name", "")
                        away_name = away.get("team", {}).get("name", "")
                        home_score = home.get("score", 0) or 0
                        away_score = away.get("score", 0) or 0
                        if not home_name or not away_name or (home_score == 0 and away_score == 0):
                            continue

                        actual_winner = home_name if home_score > away_score else away_name
                        pred = pred_by_key.get((home_name, away_name), {})
                        predicted_winner = pred.get("Predicted Winner", "")
                        correct = ("✓ Correct" if actual_winner == predicted_winner else "✗ Wrong") if predicted_winner else None

                        # Phase 4.1: Update Elo ratings after each resolved game
                        try:
                            update_elo_after_resolve(home_name, away_name, home_score, away_score)
                        except Exception as elo_err:
                            print(f"  [Elo] Update skipped: {elo_err}")

                        row_data = {
                            "date":             check_date,
                            "home_team":        home_name,
                            "away_team":        away_name,
                            "predicted_winner": predicted_winner,
                            "actual_winner":    actual_winner,
                            "correct":          correct,
                            "home_score":       home_score,
                            "away_score":       away_score,
                            "confidence":       pred.get("Confidence", ""),
                            "model_edge":       pred.get("Model Edge"),
                            "home_win_prob":    pred.get("Home Win Probability"),
                            "game_time":        pred.get("Game Time", ""),
                        }
                        insert_url = f"{sb_url}/rest/v1/results"
                        insert_body = json.dumps(row_data).encode()
                        insert_req = _req.Request(insert_url, data=insert_body, method="POST")
                        insert_req.add_header("apikey", sb_key)
                        insert_req.add_header("Authorization", f"Bearer {sb_key}")
                        insert_req.add_header("Content-Type", "application/json")
                        insert_req.add_header("Prefer", "resolution=ignore-duplicates,return=minimal")
                        try:
                            with _req.urlopen(insert_req, timeout=8):
                                inserted_count += 1
                        except Exception as ie:
                            print(f"    [Tracker] Insert failed {home_name}: {ie}")

                if inserted_count:
                    print(f"  [Tracker] ✅ Backfilled {inserted_count} games for {check_date}")

            except Exception as day_err:
                print(f"  [Tracker] Backfill error for {check_date}: {day_err}")

    except Exception as e:
        print(f"  [Tracker] Supabase backfill error: {e}")

    # ── Local CSV backfill (if file exists) ───────────────────
    if not os.path.exists(TRACKER_CSV):
        return

    df = _safe_read_tracker_csv()
    today = today_et()

    # Remove future game rows (dates after today ET)
    df = df[df["Date"] <= today]

    # Find all past dates with unresolved rows
    unresolved = df[
        df["Actual Winner"].isna() |
        (df["Actual Winner"].str.strip() == "") |
        (df["Actual Winner"].str.lower() == "nan")
    ]
    unresolved_dates = [d for d in unresolved["Date"].dropna().unique() if d < today]

    if not unresolved_dates:
        print(f"  [Tracker] All past rows resolved (today ET={today})")
        df.to_csv(TRACKER_CSV, index=False)
        return

    print(f"  [Tracker] Backfilling: {sorted(unresolved_dates)}")
    total_updated = 0

    for date_str in sorted(unresolved_dates):
        data = _get(f"{MLB_BASE}/schedule",
                    {"sportId": 1, "date": date_str, "hydrate": "linescore,team"})
        if not data:
            continue

        results = {}
        for date_entry in data.get("dates", []):
            for game in date_entry.get("games", []):
                state  = game.get("status", {}).get("abstractGameState", "")
                detail = game.get("status", {}).get("detailedState", "").lower()
                if state != "Final" or "postponed" in detail:
                    continue
                teams      = game.get("teams", {})
                home       = teams.get("home", {})
                away       = teams.get("away", {})
                home_name  = home.get("team", {}).get("name", "")
                away_name  = away.get("team", {}).get("name", "")
                home_score = home.get("score", 0) or 0
                away_score = away.get("score", 0) or 0
                if home_score == 0 and away_score == 0:
                    continue
                winner = home_name if home_score > away_score else away_name
                results[(home_name, away_name)] = winner
                # Phase 4.1: Update Elo after resolve
                try:
                    update_elo_after_resolve(home_name, away_name, home_score, away_score)
                except Exception:
                    pass

        updated = 0
        for idx, row in df.iterrows():
            if str(row.get("Date", "")).strip() != date_str:
                continue
            if str(row.get("Actual Winner", "")).strip() not in ("", "nan", "None"):
                continue
            key = (str(row.get("Home Team", "")).strip(),
                   str(row.get("Away Team", "")).strip())
            if key in results:
                actual    = results[key]
                predicted = str(row.get("Predicted Winner", "")).strip()
                df.at[idx, "Actual Winner"] = actual
                df.at[idx, "Correct?"]      = "✓ Correct" if actual == predicted else "✗ Wrong"
                updated += 1
                total_updated += 1

        print(f"    {date_str}: filled {updated} result(s) from {len(results)} final games")

    df.to_csv(TRACKER_CSV, index=False)
    print(f"  [Tracker] Done. Total updated: {total_updated}")

    # Sync resolved results to Supabase
    if total_updated:
        _sync_results_supabase(df)


def _sync_results_supabase(df):
    """Upsert resolved results from CSV into Supabase — including feature columns for retraining."""
    import urllib.request as _req
    sb_url = os.environ.get("SUPABASE_URL", "https://wkxpdmfabiepkfdxbdie.supabase.co")
    sb_key = os.environ.get("SUPABASE_KEY", "")

    def safe_float(v):
        try: return float(v) if v and str(v).strip() not in ('', 'nan', 'None') else None
        except: return None

    def safe_bool(v):
        if str(v).strip().lower() in ('true','1','yes'): return True
        if str(v).strip().lower() in ('false','0','no'): return False
        return None

    resolved = df[df['Actual Winner'].notna() & (df['Actual Winner'].str.strip() != '')]
    rows = []
    for _, r in resolved.iterrows():
        rows.append({
            'date':                     str(r.get('Date','')).strip(),
            'home_team':                str(r.get('Home Team','')).strip(),
            'away_team':                str(r.get('Away Team','')).strip(),
            'predicted_winner':         str(r.get('Predicted Winner','')).strip() or None,
            'actual_winner':            str(r.get('Actual Winner','')).strip() or None,
            'correct':                  str(r.get('Correct?','')).strip() or None,
            'confidence':               str(r.get('Confidence','')).strip() or None,
            'model_edge':               safe_float(r.get('Model Edge')),
            'home_win_prob':            safe_float(r.get('Home Win %')),
            'game_time':                str(r.get('Game Time','')).strip() or None,
            # Feature columns for model retraining
            'home_sp_era':              safe_float(r.get('pitcher_era')),
            'away_sp_era':              safe_float(r.get('opp_pitcher_era')),
            'bullpen_era':              safe_float(r.get('bullpen_era')),
            'opp_bullpen_era':          safe_float(r.get('opp_bullpen_era')),
            'park_factor':              safe_float(r.get('park_factor')),
            'temp':                     safe_float(r.get('temp')),
            'wind_speed':               safe_float(r.get('wind_speed')),
            'wind_dir_out':             safe_bool(r.get('wind_dir_out')),
            'home':                     safe_bool(r.get('home')),
            'pitcher_recent_delta':     safe_float(r.get('pitcher_recent_delta')),
            'opp_pitcher_recent_delta': safe_float(r.get('opp_pitcher_recent_delta')),
        })

    if not rows:
        return

    try:
        # Use psycopg2 direct connection — bypasses REST API host restrictions
        import psycopg2, psycopg2.extras
        db_url = os.environ.get("DATABASE_URL", "")
        if db_url:
            conn = psycopg2.connect(db_url, sslmode='require')
            cur  = conn.cursor()
            for row in rows:
                cols = list(row.keys())
                vals = [row[c] for c in cols]
                placeholders = ','.join(['%s'] * len(cols))
                col_str = ','.join(cols)
                update_str = ','.join([f"{c}=EXCLUDED.{c}" for c in cols if c not in ('date','home_team','away_team')])
                cur.execute(f"""
                    INSERT INTO results ({col_str}) VALUES ({placeholders})
                    ON CONFLICT (date, home_team, away_team) DO UPDATE SET {update_str}
                """, vals)
            conn.commit()
            cur.close(); conn.close()
            print(f"  [Supabase] {len(rows)} results synced with features (psycopg2)")
            return
    except Exception as e:
        print(f"  [Supabase] psycopg2 sync failed: {e} — trying REST API")

    # Fallback to REST API
    try:
        body = json.dumps(rows, default=str).encode()
        url = f"{sb_url}/rest/v1/results"
        request = _req.Request(url, data=body, method='POST')
        request.add_header('apikey', sb_key)
        request.add_header('Authorization', f'Bearer {sb_key}')
        request.add_header('Content-Type', 'application/json')
        request.add_header('Prefer', 'resolution=merge-duplicates,return=minimal')
        with _req.urlopen(request, timeout=15) as resp:
            print(f"  [Supabase] {len(rows)} results synced via REST (HTTP {resp.status})")
    except Exception as e:
        print(f"  [Supabase] Results sync failed: {e}")


def _vig_free_prob(home_ml, away_ml):
    """
    Convert American moneylines to a vig-removed (true) home win probability.
    Removes the bookmaker's juice so both sides sum to 100%.
    """
    def to_raw(ml):
        return (100 / (ml + 100)) if ml > 0 else (abs(ml) / (abs(ml) + 100))
    h = to_raw(home_ml)
    a = to_raw(away_ml)
    total = h + a
    return h / total if total > 0 else 0.500


def fetch_vegas_odds():
    """
    Phase 3 — Enhanced Vegas Odds with bookmaker tier weighting.
    Sharp books (Pinnacle, Circa) move first on real information.
    Square books (FanDuel, DraftKings) follow the market.
    Weighting sharp books more heavily gives a better consensus line.
    Also stores opening lines to DB for CLV tracking later.
    """
    global _vegas_cache
    if _vegas_cache:
        return _vegas_cache

    api_key = os.environ.get("ODDS_API_KEY", "").strip()
    if not api_key:
        print("  [Vegas] ODDS_API_KEY not set — win% fallback active")
        return {}

    # ── Phase 3: Bookmaker tier weights ──────────────────────────
    # Sharp books set the line — weight them 3x
    # Medium books follow quickly — weight 2x
    # Square books are last to move — weight 1x
    BOOK_WEIGHTS = {
        # Sharp (tier 1) — 3x weight
        "pinnacle":    3.0,
        "circa":       3.0,
        "bookmaker":   3.0,
        "lowvig":      2.5,
        # Medium (tier 2) — 2x weight
        "betrivers":   2.0,
        "pointsbet":   2.0,
        "williamhill": 2.0,
        "superbook":   2.0,
        # Square (tier 3) — 1x weight
        "fanduel":     1.0,
        "draftkings":  1.0,
        "betmgm":      1.0,
        "caesars":     1.0,
        "barstool":    1.0,
        "unibet":      1.0,
        "betway":      1.0,
    }
    DEFAULT_WEIGHT = 1.5  # unknown books get medium-low weight

    try:
        r = requests.get(
            "https://api.the-odds-api.com/v4/sports/baseball_mlb/odds/",
            params={
                "apiKey":      api_key,
                "regions":     "us",
                "markets":     "h2h,spreads",
                "oddsFormat":  "american",
                "dateFormat":  "iso",
            },
            timeout=10,
        )
        if r.status_code != 200:
            print(f"  [Vegas] Odds API returned {r.status_code} — win% fallback active")
            return {}

        result = {}
        for game in r.json():
            home_team  = game.get("home_team", "")
            away_team  = game.get("away_team", "")
            bookmakers = game.get("bookmakers", [])
            if not bookmakers:
                continue

            # ── Weighted moneyline consensus ──────────────────────
            weighted_probs   = []
            weighted_home_ml = []
            total_weight_ml  = 0.0

            # ── Run line (spreads) ────────────────────────────────
            weighted_home_rl = []
            weighted_away_rl = []
            home_rl_points   = []
            total_weight_rl  = 0.0

            # ── Per-book data for tier analysis ───────────────────
            sharp_probs  = []   # Pinnacle/Circa only
            square_probs = []   # FanDuel/DraftKings only

            for bm in bookmakers:
                bm_key    = bm.get("key", "").lower()
                bm_weight = BOOK_WEIGHTS.get(bm_key, DEFAULT_WEIGHT)
                is_sharp  = BOOK_WEIGHTS.get(bm_key, 0) >= 2.5
                is_square = BOOK_WEIGHTS.get(bm_key, 99) <= 1.0

                for mkt in bm.get("markets", []):
                    if mkt["key"] == "h2h":
                        outcomes = {o["name"]: o["price"] for o in mkt.get("outcomes", [])}
                        hml = outcomes.get(home_team)
                        aml = outcomes.get(away_team)
                        if hml is not None and aml is not None:
                            prob = _vig_free_prob(hml, aml)
                            weighted_probs.append(prob * bm_weight)
                            weighted_home_ml.append(hml * bm_weight)
                            total_weight_ml += bm_weight
                            if is_sharp:  sharp_probs.append(prob)
                            if is_square: square_probs.append(prob)

                    elif mkt["key"] == "spreads":
                        for o in mkt.get("outcomes", []):
                            if o["name"] == home_team:
                                weighted_home_rl.append(o["price"] * bm_weight)
                                home_rl_points.append(o.get("point", -1.5))
                                total_weight_rl += bm_weight
                            elif o["name"] == away_team:
                                weighted_away_rl.append(o["price"] * bm_weight)

            if not weighted_probs or total_weight_ml == 0:
                continue

            # Weighted consensus probability
            consensus_prob    = sum(weighted_probs) / total_weight_ml
            consensus_home_ml = round(sum(weighted_home_ml) / total_weight_ml)

            # Sharp vs square divergence
            sharp_consensus  = round(sum(sharp_probs) / len(sharp_probs), 4) if sharp_probs else None
            square_consensus = round(sum(square_probs) / len(square_probs), 4) if square_probs else None
            book_divergence  = None
            if sharp_consensus and square_consensus:
                book_divergence = round(sharp_consensus - square_consensus, 4)
                if abs(book_divergence) > 0.03:
                    direction = "home" if book_divergence > 0 else "away"
                    print(f"  [Vegas] 📊 {home_team} sharp/square divergence: {book_divergence:+.3f} → sharps favor {direction}")

            # Run line consensus
            rl_data = {}
            if weighted_home_rl and weighted_away_rl and total_weight_rl > 0:
                avg_home_rl = round(sum(weighted_home_rl) / total_weight_rl)
                avg_away_rl = round(sum(weighted_away_rl) / total_weight_rl)
                avg_rl_pts  = round(sum(home_rl_points) / len(home_rl_points), 1)
                home_cover_prob = _vig_free_prob(avg_home_rl, avg_away_rl)
                rl_data = {
                    "home_rl_line":      avg_rl_pts,
                    "away_rl_line":      -avg_rl_pts,
                    "home_rl_odds":      avg_home_rl,
                    "away_rl_odds":      avg_away_rl,
                    "home_cover_prob":   round(home_cover_prob, 4),
                    "away_cover_prob":   round(1 - home_cover_prob, 4),
                    "rl_is_live":        True,
                }
                print(f"  [Vegas] {home_team} RL: {avg_rl_pts:+.1f} ({avg_home_rl}) | {away_team} RL: {-avg_rl_pts:+.1f} ({avg_away_rl})")

            result[(home_team, away_team)] = {
                "home_moneyline":    consensus_home_ml,
                "away_moneyline":    -consensus_home_ml,
                "home_prob_novig":   round(consensus_prob, 4),
                "is_live":           True,
                "sharp_consensus":   sharp_consensus,
                "square_consensus":  square_consensus,
                "book_divergence":   book_divergence,
                "books_used":        len(bookmakers),
                **rl_data,
            }

        _vegas_cache = result
        rl_count = sum(1 for v in result.values() if v.get("rl_is_live"))
        sharp_count = sum(1 for v in result.values() if v.get("sharp_consensus"))
        print(f"  [Vegas] Live consensus odds loaded for {len(result)} games · {rl_count} with run lines · {sharp_count} with sharp data")
        return result

    except Exception as e:
        print(f"  [Vegas] WARNING: {e} — win% fallback active")
        return {}




# ══════════════════════════════════════════════════════════════
# SPORTSBOOK INTELLIGENCE — RLM, STEAM, SHARP, PUBLIC FADE
# ══════════════════════════════════════════════════════════════

# Opening lines cache — stored at 9 AM, compared to current at prediction time
_opening_lines_cache = {}

def capture_closing_lines():
    """
    Closing Line Value capture.

    CLV is the fastest honest read on whether a model has real edge:
    it is detectable at ~100 bets where win rate needs 500+, because it
    measures the model against the market's final price rather than against
    the enormous variance of individual game outcomes. Beating the closing
    line IS the definition of edge; win rate is edge plus noise.

    Must run CLOSE to first pitch. A line captured at 9 AM is an opening
    line, not a closing line, and would make CLV meaningless. This is
    designed to be called on a schedule through the day: it only writes
    for games starting within CLOSING_WINDOW_MIN, and never overwrites a
    closing price once stored.

        CLV = model_prob - closing_implied_prob   (both vig-free, home side)

    Positive mean CLV on the model's picks = the model is finding value the
    market later agrees with. That is real edge. Negative = it is not.
    """
    CLOSING_WINDOW_MIN = 90     # capture within 90 min of first pitch
    db_url = os.environ.get("DATABASE_URL", "")
    if not db_url:
        print("  [CLV] No DATABASE_URL — skipping")
        return {"captured": 0}

    # Force a fresh odds pull; the cached copy is from the 9 AM run.
    global _vegas_cache
    _vegas_cache = {}
    all_odds = fetch_vegas_odds()
    if not all_odds:
        print("  [CLV] No live odds available — skipping")
        return {"captured": 0}

    today = today_et()
    now   = datetime.datetime.now(ET)
    captured = 0
    skipped_early = 0

    try:
        import psycopg2
        conn = psycopg2.connect(db_url, sslmode="require", connect_timeout=10)
        cur  = conn.cursor()

        # Only today's games that still need a closing price.
        cur.execute("""
            SELECT home_team, away_team, game_time, home_win_prob,
                   opening_home_implied, predicted_winner
            FROM results
            WHERE date = %s AND closing_home_implied IS NULL
        """, (today,))
        rows = cur.fetchall()

        for home, away, game_time, home_prob, opening_prob, pick in rows:
            odds = all_odds.get((home, away))
            if not odds:
                continue
            closing_prob = odds.get("home_prob_novig")
            if closing_prob is None:
                continue

            # Only capture near first pitch — otherwise it isn't a closing line.
            if game_time:
                mins_to_start = _minutes_until_game(game_time, now)
                if mins_to_start is not None and mins_to_start > CLOSING_WINDOW_MIN:
                    skipped_early += 1
                    continue

            # CLV is a MARKET-to-MARKET comparison: did the line move toward
            # our pick between when we made it and when it closed?
            #
            #     CLV = closing_implied - opening_implied   (on the picked side)
            #
            # It deliberately does NOT reference model probability. Comparing
            # model_prob to closing_prob just restates the model's own edge,
            # and for an overconfident model that is positive almost by
            # construction — the first implementation did exactly that and
            # produced a meaningless +18.7% with 100% of picks "positive".
            #
            # Positive CLV here means the market agreed with us after the fact,
            # which is real evidence independent of what the model claims.
            clv = None
            if opening_prob is not None and pick:
                op, cp = float(opening_prob), float(closing_prob)
                if pick == home:
                    clv = round(cp - op, 4)
                else:
                    clv = round((1.0 - cp) - (1.0 - op), 4)

            cur.execute("""
                UPDATE results
                SET closing_home_implied = %s,
                    model_clv            = %s
                WHERE date = %s AND home_team = %s AND away_team = %s
                  AND closing_home_implied IS NULL
            """, (closing_prob, clv, today, home, away))
            if cur.rowcount > 0:
                captured += 1

        conn.commit()
        cur.close()
        conn.close()

        if captured:
            print(f"  [CLV] ✅ Closing lines captured for {captured} game(s)")
        if skipped_early:
            print(f"  [CLV] {skipped_early} game(s) >{CLOSING_WINDOW_MIN}min out — will capture closer to first pitch")
        if not captured and not skipped_early:
            print(f"  [CLV] Nothing to capture")

    except Exception as e:
        print(f"  [CLV] Capture failed: {e}")
        return {"captured": 0, "error": str(e)}

    return {"captured": captured, "pending": skipped_early}


def _minutes_until_game(game_time_str, now_et):
    """
    Minutes from now until first pitch. Returns None if unparseable so the
    caller degrades to capturing rather than silently skipping forever.
    game_time is stored like '7:07 PM ET'.
    """
    try:
        s = str(game_time_str).replace("ET", "").strip()
        if not s or s.lower() == "nan":
            return None
        t = datetime.datetime.strptime(s, "%I:%M %p").time()
        start = now_et.replace(hour=t.hour, minute=t.minute, second=0, microsecond=0)
        return (start - now_et).total_seconds() / 60.0
    except Exception:
        return None


def clv_summary(days: int = 30) -> dict:
    """
    Aggregate CLV over recent games. This is the number that answers
    "does the model beat the market?" faster than win rate can.

    Reported on the model's PICKED side, which is what actually matters —
    CLV on a side you didn't bet is not evidence of anything.
    """
    db_url = os.environ.get("DATABASE_URL", "")
    if not db_url:
        return {}
    try:
        import psycopg2, numpy as np
        conn = psycopg2.connect(db_url, sslmode="require", connect_timeout=10)
        cur  = conn.cursor()
        cur.execute("""
            SELECT home_team, predicted_winner,
                   opening_home_implied, closing_home_implied, model_cohort
            FROM results
            WHERE closing_home_implied IS NOT NULL
              AND opening_home_implied IS NOT NULL
              AND predicted_winner IS NOT NULL
              AND date >= %s
        """, ((datetime.datetime.now(ET) - datetime.timedelta(days=days)).strftime("%Y-%m-%d"),))
        rows = cur.fetchall()
        cur.close(); conn.close()

        picked_clv, cohort_b_clv = [], []
        for home, pick, open_p, close_p, cohort in rows:
            op, cp = float(open_p), float(close_p)
            # Market-to-market on the picked side. Did the line move our way?
            if pick == home:
                v = cp - op
            else:
                v = (1.0 - cp) - (1.0 - op)
            picked_clv.append(v)
            if cohort == "B":
                cohort_b_clv.append(v)

        def _agg(vals):
            if not vals:
                return None
            a = np.array(vals)
            return {
                "n": len(a),
                "mean_clv": round(float(a.mean()), 4),
                "pct_positive": round(float((a > 0).mean()), 4),
            }

        return {"all": _agg(picked_clv), "cohort_b": _agg(cohort_b_clv)}
    except Exception as e:
        print(f"  [CLV] Summary failed: {e}")
        return {}


def store_opening_lines(all_odds: dict):
    """
    Phase 3 — Store opening lines at 9 AM for CLV tracking.
    Saves to in-memory cache AND to Supabase results table.
    This is the foundation for closing line value computation later.
    """
    global _opening_lines_cache
    _opening_lines_cache = {
        k: {"home_prob": v["home_prob_novig"], "stored_at": today_et()}
        for k, v in all_odds.items()
    }

    if not all_odds:
        return

    # Write opening lines to results table for CLV tracking
    try:
        import psycopg2
        db_url = os.environ.get("DATABASE_URL", "")
        if not db_url:
            return
        today = today_et()
        conn = psycopg2.connect(db_url, sslmode='require', connect_timeout=10)
        cur  = conn.cursor()
        updated = 0
        for (home, away), odds in all_odds.items():
            opening_prob = odds.get("home_prob_novig")
            if opening_prob is None:
                continue
            cur.execute("""
                UPDATE results
                SET opening_home_implied = %s
                WHERE date = %s AND home_team = %s AND away_team = %s
                AND opening_home_implied IS NULL
            """, (opening_prob, today, home, away))
            if cur.rowcount > 0:
                updated += 1
        conn.commit()
        cur.close()
        conn.close()
        if updated:
            print(f"  [Vegas] ✅ Opening lines stored for {updated} games")
        elif all_odds:
            # Silence here previously hid an ordering bug for days: the
            # UPDATE matched no rows and simply printed nothing. Any game
            # without an opening price can never have CLV computed, so this
            # must be visible.
            print(f"  [Vegas] ⚠️  Opening lines stored for 0 of {len(all_odds)} games "
                  f"— results rows may not exist yet. CLV will be uncomputable for today.")
    except Exception as e:
        print(f"  [Vegas] Opening line storage skipped: {e}")

def detect_reverse_line_movement(home_team: str, away_team: str,
                                  opening_prob: float, current_prob: float,
                                  public_pct_home: float = None) -> dict:
    """
    RLM = Public heavily on one side, but line moves toward the other side.
    This is the most reliable sharp money signal available.
    """
    movement = current_prob - opening_prob  # positive = line moved toward home

    if public_pct_home is None:
        # Without public data, just track line movement direction
        if abs(movement) >= 0.04:
            return {
                "detected":        True,
                "sharp_side":      home_team if movement < 0 else away_team,
                "movement":        round(movement * 100, 1),
                "signal_strength": min(10, int(abs(movement) * 150)),
                "label":           f"⚡ Sharp line movement {movement*100:+.1f}%",
                "type":            "line_movement",
            }
        return {"detected": False}

    # Full RLM with public data
    public_on_home = public_pct_home > 60
    line_moved_away_from_home = movement < -0.03

    public_on_away = public_pct_home < 40
    line_moved_away_from_away = movement > 0.03

    if public_on_home and line_moved_away_from_home:
        strength = min(10, int((public_pct_home - 60) * 0.4 + abs(movement) * 120))
        return {
            "detected":        True,
            "sharp_side":      away_team,
            "public_side":     home_team,
            "movement":        round(movement * 100, 1),
            "public_pct":      public_pct_home,
            "signal_strength": strength,
            "label":           f"⚡ Sharp money on {away_team.split()[-1]} — public fading wrong side",
            "type":            "rlm",
        }
    elif public_on_away and line_moved_away_from_away:
        strength = min(10, int((40 - public_pct_home) * 0.4 + abs(movement) * 120))
        return {
            "detected":        True,
            "sharp_side":      home_team,
            "public_side":     away_team,
            "movement":        round(movement * 100, 1),
            "public_pct":      public_pct_home,
            "signal_strength": strength,
            "label":           f"⚡ Sharp money on {home_team.split()[-1]} — public fading wrong side",
            "type":            "rlm",
        }

    return {"detected": False}


def detect_steam_move(opening_prob: float, current_prob: float,
                       hours_since_open: float = 24.0) -> dict:
    """
    Phase 3 — Steam detection with line velocity.
    Velocity matters: a 5% move in 20 minutes is steam.
    A 5% move over 12 hours is normal market efficiency.
    """
    movement = abs(current_prob - opening_prob)
    velocity = movement / max(hours_since_open, 0.1)  # % per hour
    time_factor = max(0.5, 1.0 - (hours_since_open / 168.0))
    weighted = movement * time_factor

    # Steam: high velocity OR large weighted movement
    is_steam = (velocity >= 0.004 and movement >= 0.03) or weighted >= 0.05

    if is_steam:
        return {
            "detected":        True,
            "magnitude":       round(movement * 100, 1),
            "velocity":        round(velocity * 100, 2),
            "weighted":        round(weighted * 100, 1),
            "signal_strength": min(10, int(weighted * 140)),
            "label":           f"🔥 Steam move — {movement*100:.1f}% ({velocity*100:.2f}%/hr)",
        }
    return {"detected": False}


def detect_trap_line(home_team: str, current_implied: float,
                      historical_avg_implied: float = None) -> dict:
    """
    Trap line = book offering attractive price on popular team to bait public.
    Usually set up when a popular team just had a big win.
    """
    POPULAR_TEAMS = {
        "New York Yankees", "Los Angeles Dodgers", "Boston Red Sox",
        "Chicago Cubs", "Atlanta Braves", "Houston Astros",
        "New York Mets", "San Francisco Giants",
    }
    if home_team not in POPULAR_TEAMS:
        return {"detected": False}

    if historical_avg_implied and current_implied < historical_avg_implied - 0.05:
        return {
            "detected": True,
            "label":    f"⚠️ Potential trap — {home_team.split()[-1]} unusually cheap",
            "note":     "Book may be baiting public. Verify before tailing.",
        }
    return {"detected": False}


def get_public_betting_data(home_team: str, away_team: str) -> dict:
    """
    Fetch public betting percentages from aggregators.
    Tries multiple free sources. Falls back to neutral if unavailable.
    """
    # In production: integrate with Action Network API or SportsBookReview
    # For now returns neutral — will be populated when API key added
    # Add ACTION_NETWORK_KEY to Render env when available
    return {
        "home_pct":    50,
        "away_pct":    50,
        "total_pct":   50,
        "source":      "neutral_fallback",
        "is_live":     False,
    }


# ══════════════════════════════════════════════════════════════
# STATCAST ENHANCED PITCHER FEATURES — Stuff+, xFIP, SIERA
# ══════════════════════════════════════════════════════════════

def get_statcast_pitcher_quality(pitcher_id: int, pitcher_name: str = "") -> dict:
    """
    Fetch Stuff+ and location quality from Baseball Savant.
    Stuff+ > 100 = above average pitch quality.
    xFIP removes home run luck from ERA.
    """
    try:
        import pybaseball
        # Statcast pitcher data from pybaseball
        data = pybaseball.statcast_pitcher(
            start_dt=f"{today_et()[:4]}-03-01",
            end_dt=today_et(),
            player_id=pitcher_id
        )
        if data is None or data.empty:
            return _statcast_defaults()

        # Compute key metrics
        total_pitches = len(data)
        if total_pitches < 50:
            return _statcast_defaults()

        # Estimated Stuff+ from pitch characteristics
        avg_velo = data["release_speed"].mean() if "release_speed" in data else 92.0
        avg_spin = data["release_spin_rate"].mean() if "release_spin_rate" in data else 2200.0
        avg_break = data["pfx_z"].abs().mean() if "pfx_z" in data else 8.0

        # Simplified Stuff+ proxy (proper version uses MLB's model)
        stuff_proxy = round(
            100 +
            (avg_velo - 92.0) * 2.5 +
            (avg_spin - 2200) * 0.008 +
            (avg_break - 8.0) * 1.5
        , 1)

        # xFIP proxy from WHIP and K/BB
        whiff_rate = (data["description"].str.contains("swinging_strike", na=False)).mean()

        return {
            "stuff_plus":     max(60, min(160, stuff_proxy)),
            "avg_velo":       round(avg_velo, 1),
            "avg_spin":       round(avg_spin),
            "whiff_rate":     round(whiff_rate, 3),
            "pitch_count_ytd": total_pitches,
            "statcast_live":  True,
        }
    except Exception as e:
        print(f"  [Statcast] {pitcher_name}: {e}")
        return _statcast_defaults()


def _statcast_defaults() -> dict:
    return {
        "stuff_plus":     100,
        "avg_velo":       92.0,
        "avg_spin":       2200,
        "whiff_rate":     0.24,
        "pitch_count_ytd": 0,
        "statcast_live":  False,
    }


# ══════════════════════════════════════════════════════════════
# LIGHTGBM ENSEMBLE — adds to existing XGBoost model
# ══════════════════════════════════════════════════════════════

def get_lightgbm_prediction(features: list, model_path: str = "lgbm_model.pkl") -> float:
    """
    LightGBM prediction to blend with XGBoost.
    Falls back to None if model not yet trained (first week).
    """
    try:
        import lightgbm as lgb
        import pickle, os
        if not os.path.exists(model_path):
            return None
        with open(model_path, "rb") as f:
            lgbm = pickle.load(f)
        prob = lgbm.predict_proba([features])[0][1]
        return round(float(prob), 4)
    except Exception:
        return None


def retrain_lightgbm(X_train, y_train, model_path: str = "lgbm_model.pkl"):
    """Train or retrain LightGBM model alongside XGBoost."""
    try:
        import lightgbm as lgb
        import pickle
        lgbm = lgb.LGBMClassifier(
            n_estimators=200,
            learning_rate=0.05,
            num_leaves=31,
            min_child_samples=5,
            random_state=42,
            verbose=-1,
        )
        lgbm.fit(X_train, y_train)
        with open(model_path, "wb") as f:
            pickle.dump(lgbm, f)
        print(f"  [LightGBM] Retrained on {len(y_train)} rows")
        return lgbm
    except Exception as e:
        print(f"  [LightGBM] Training failed: {e}")
        return None


# ══════════════════════════════════════════════════════════════
# PROBABILITY CALIBRATION — Brier score, log loss, buckets
# ══════════════════════════════════════════════════════════════

def compute_calibration_metrics(conn) -> dict:
    """
    Compute Brier score, log loss, and per-bucket calibration.
    Run after each retrain to verify model quality.
    """
    try:
        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT home_win_prob, correct
            FROM results
            WHERE actual_winner IS NOT NULL
              AND home_win_prob IS NOT NULL
              AND correct IS NOT NULL
            ORDER BY date DESC
            LIMIT 300
        """)
        rows = cur.fetchall()
        cur.close()

        if len(rows) < 20:
            return {"error": "insufficient data", "sample_size": len(rows)}

        import numpy as np
        probs    = np.array([r["home_win_prob"] / 100.0 for r in rows])
        outcomes = np.array([1 if r["correct"] and "✓" in (r["correct"] or "") else 0 for r in rows])

        # Brier score (lower = better, 0.25 = random coin flip)
        brier = float(np.mean((probs - outcomes) ** 2))

        # Log loss
        eps = 1e-7
        probs_clipped = np.clip(probs, eps, 1 - eps)
        ll = float(-np.mean(outcomes * np.log(probs_clipped) + (1 - outcomes) * np.log(1 - probs_clipped)))

        # Bucket calibration
        buckets = []
        for lower in [0.50, 0.55, 0.60, 0.65, 0.70, 0.75, 0.80]:
            upper = lower + 0.05
            mask  = (probs >= lower) & (probs < upper)
            if mask.sum() < 5:
                continue
            buckets.append({
                "bucket":          f"{int(lower*100)}-{int(upper*100)}%",
                "count":           int(mask.sum()),
                "predicted_avg":   round(float(probs[mask].mean() * 100), 1),
                "actual_win_pct":  round(float(outcomes[mask].mean() * 100), 1),
                "error":           round(float(abs(probs[mask].mean() - outcomes[mask].mean()) * 100), 1),
            })

        result = {
            "brier_score":       round(brier, 4),
            "log_loss":          round(ll, 4),
            "sample_size":       len(rows),
            "overall_accuracy":  round(float(outcomes.mean() * 100), 1),
            "bucket_analysis":   buckets,
            "computed_at":       today_et(),
            "grade": "A" if brier < 0.19 else "B" if brier < 0.22 else "C" if brier < 0.25 else "D",
        }
        print(f"  [Calibration] Brier: {brier:.4f} | LogLoss: {ll:.4f} | Grade: {result['grade']}")
        return result

    except Exception as e:
        print(f"  [Calibration] Failed: {e}")
        return {"error": str(e)}


def save_calibration_metrics(conn, metrics: dict):
    """Store calibration history in Supabase."""
    try:
        import json, psycopg2
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO calibration_history
                (brier_score, log_loss, sample_size, overall_accuracy,
                 bucket_analysis, grade, computed_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            metrics.get("brier_score"),
            metrics.get("log_loss"),
            metrics.get("sample_size"),
            metrics.get("overall_accuracy"),
            json.dumps(metrics.get("bucket_analysis", [])),
            metrics.get("grade"),
            metrics.get("computed_at"),
        ))
        conn.commit()
        cur.close()
        print(f"  [Calibration] Saved to Supabase")
    except Exception as e:
        print(f"  [Calibration] Save failed: {e}")


# ══════════════════════════════════════════════════════════════
# PREDICTION FAILURE ANALYSIS
# ══════════════════════════════════════════════════════════════

FAILURE_CLASSES = {
    "BULLPEN_VARIANCE":      "Bullpen allowed significantly more runs than expected",
    "SP_UNDERPERFORMANCE":   "Starting pitcher underperformed their weighted ERA",
    "SP_OVERPERFORMANCE":    "Opposing starter overperformed — model underweighted upside",
    "WEATHER_ANOMALY":       "Weather significantly worse than forecast",
    "MARKET_MISREAD":        "Sharp money was on losing side — market was right",
    "DEFENSIVE_VARIANCE":    "Unusual defensive errors changed game outcome",
    "OFFENSIVE_OUTLIER":     "Team scored far above or below expected runs",
    "UMPIRE_VARIANCE":       "Strike zone deviated significantly from umpire norm",
    "LINEUP_CHANGE":         "Significant unannounced lineup change before game",
    "RANDOM_VARIANCE":       "No identifiable cause — within expected variance",
}

def classify_prediction_failure(prediction: dict, actual: dict) -> tuple[str, float]:
    """
    Rule-based failure classifier. Returns (class, confidence).
    Fast, no AI required — AI explanation called separately.
    """
    # Bullpen variance: bullpen allowed 3+ more runs than ERA suggests
    pred_bull_era = prediction.get("bullpen_era", 4.0)
    expected_bull_runs = pred_bull_era / 9.0 * 3.5
    actual_bull_runs = actual.get("bullpen_runs_allowed", expected_bull_runs)
    if actual_bull_runs > expected_bull_runs + 2.5:
        return "BULLPEN_VARIANCE", 0.82

    # SP underperformance: pitched much worse than ERA suggests
    pred_sp_era = prediction.get("home_sp_era", 4.0)
    actual_game_era = actual.get("sp_era_game", pred_sp_era)
    if actual_game_era > pred_sp_era * 1.6 and actual_game_era > 6.0:
        return "SP_UNDERPERFORMANCE", 0.78

    # Market misread: sharp money signal was present but model went other way
    if prediction.get("sharp_money_signal") and not prediction.get("model_agrees_with_sharp"):
        return "MARKET_MISREAD", 0.71

    # Weather anomaly
    pred_wind = prediction.get("wind_speed", 0)
    actual_wind = actual.get("actual_wind_speed", pred_wind)
    if abs(actual_wind - pred_wind) > 12:
        return "WEATHER_ANOMALY", 0.65

    # Offensive outlier: team scored 3+ more or fewer than expected
    pred_runs = prediction.get("predicted_home_runs", 4.5)
    actual_runs = actual.get("actual_home_runs", pred_runs)
    if abs(actual_runs - pred_runs) > 3.5:
        return "OFFENSIVE_OUTLIER", 0.60

    return "RANDOM_VARIANCE", 0.55


def compute_sample_weight(failure_class: str, edge_score: int) -> float:
    """
    Weight training samples by failure type.
    Informative failures get more weight in retraining.
    Random noise gets downweighted.
    """
    base_weights = {
        "BULLPEN_VARIANCE":    0.85,
        "SP_UNDERPERFORMANCE": 1.20,
        "SP_OVERPERFORMANCE":  1.20,
        "MARKET_MISREAD":      1.50,  # sharp money right, model wrong = very informative
        "WEATHER_ANOMALY":     0.55,  # not predictable pre-game
        "DEFENSIVE_VARIANCE":  0.70,
        "OFFENSIVE_OUTLIER":   0.80,
        "UMPIRE_VARIANCE":     0.65,
        "LINEUP_CHANGE":       0.60,  # info unavailable at prediction time
        "RANDOM_VARIANCE":     0.45,  # pure noise — downweight heavily
    }
    weight = base_weights.get(failure_class, 1.0)
    # High-confidence picks that were wrong are very informative
    if edge_score >= 8:
        weight *= 1.25
    return round(min(weight, 2.0), 3)


# ══════════════════════════════════════════════════════════════
# LOCK OF THE DAY + DANGEROUS UNDERDOG
# ══════════════════════════════════════════════════════════════

def select_lock_of_the_day(games: list) -> dict | None:
    """
    Lock of Day = highest conviction pick.
    Criteria: edge_score >= 8 AND win_prob >= 65% AND vegas_is_live
    Returns the single best pick, or None if no game qualifies.
    """
    candidates = []
    for g in games:
        edge  = g.get("Model Edge", 0) or 0
        prob  = g.get("Home Win Probability", 50) or 50
        score = g.get("Edge Score", 0) or 0
        live  = g.get("Vegas Is Live", False)
        sharp = g.get("Sharp Money Signal", False)

        if score >= 8 and prob >= 62 and live:
            # Composite lock score
            lock_score = (
                score * 3 +
                max(0, edge) * 2 +
                (5 if sharp else 0)
            )
            candidates.append({**g, "_lock_score": lock_score})

    if not candidates:
        return None

    best = max(candidates, key=lambda x: x["_lock_score"])
    best["Is Lock Of Day"] = True
    print(f"  [Lock] 🔒 Lock of Day: {best.get('Predicted Winner')} (edge {best.get('Model Edge'):+.1f}%)")
    return best


def select_dangerous_underdog(games: list) -> dict | None:
    """
    Dangerous Underdog = team Vegas underrates by 10%+ that model likes.
    Most exciting pick of the day — higher payout, real model edge.
    """
    dogs = []
    for g in games:
        vegas_home_implied = g.get("Vegas Implied %", 50) or 50
        model_home_prob    = g.get("Home Win Probability", 50) or 50
        edge               = model_home_prob - vegas_home_implied
        pick_is_home       = g.get("Predicted Winner") == g.get("Home Team")

        # Underdog = Vegas gives them < 45% chance
        is_dog = (pick_is_home and vegas_home_implied < 45) or \
                 (not pick_is_home and vegas_home_implied > 55)
        model_edge = abs(edge)

        if is_dog and model_edge >= 8:
            dogs.append({**g, "_dog_edge": model_edge})

    if not dogs:
        return None

    best = max(dogs, key=lambda x: x["_dog_edge"])
    best["Is Dangerous Dog"] = True
    winner = best.get("Predicted Winner", "")
    print(f"  [Dog] 🐶 Dangerous Underdog: {winner} (+{best['_dog_edge']:.1f}% model edge)")
    return best


    """
    Return odds dict for one game.
    Priority 1 — Live consensus odds (vig-removed) from The Odds API.
    Priority 2 — Synthesized probability from team win% + home-field advantage
                 (much smarter than flat 50/50 when no API key is set).
    """
    if all_odds:
        if (home_name, away_name) in all_odds:
            return all_odds[(home_name, away_name)]
        for (h, a), odds in all_odds.items():
            if (home_name in h or h in home_name) and (away_name in a or a in away_name):
                return odds

    # Synthesized fallback: convert season win% + HFA to an implied probability
    total = (home_wpct + away_wpct) or 1.0
    home_prior     = home_wpct / total
    home_with_hfa  = min(max(home_prior + 0.035, 0.01), 0.99)
    if home_with_hfa >= 0.5:
        home_ml = round(-home_with_hfa / (1 - home_with_hfa) * 100)
    else:
        home_ml = round((1 - home_with_hfa) / home_with_hfa * 100)
    return {
        "home_moneyline":  home_ml,
        "away_moneyline":  -home_ml,
        "home_prob_novig": round(home_with_hfa, 4),
        "is_live":         False,
    }


def predict_run_differential(home_stats, away_stats, home_pitcher, away_pitcher, park_factor=1.0):
    """
    Phase 2 — Run differential model.
    Estimates expected runs for each team and predicts margin of victory.
    Uses starter ERA, bullpen ERA, runs per game, and park factor.
    Returns (predicted_margin, home_cover_prob, away_cover_prob).
    home_cover_prob = P(home wins by 2+, covers -1.5)
    """
    from scipy.stats import norm

    def expected_runs(batting_rpg, sp_era, bullpen_era, pf):
        """Estimate runs scored in a game."""
        # SP expected to pitch ~5.5 innings, bullpen ~3.5
        sp_runs      = (sp_era / 9.0) * 5.5
        bullpen_runs = (bullpen_era / 9.0) * 3.5
        pitcher_total = sp_runs + bullpen_runs
        # Blend batting runs per game with pitcher-allowed runs
        expected = (batting_rpg * 0.50 + pitcher_total * 0.50) * pf
        return max(1.5, min(12.0, expected))

    home_sp_era     = home_pitcher.get('blended_era', home_pitcher.get('season_era', 4.20))
    away_sp_era     = away_pitcher.get('blended_era', away_pitcher.get('season_era', 4.20))
    home_bull_era   = home_stats.get('bullpen_era', 4.00)
    away_bull_era   = away_stats.get('bullpen_era', 4.00)
    home_rpg        = home_stats.get('runs_per_game', 4.50)
    away_rpg        = away_stats.get('runs_per_game', 4.50)

    # Each team's expected runs scored vs opposing pitching
    home_scored     = expected_runs(home_rpg, away_sp_era, away_bull_era, park_factor)
    away_scored     = expected_runs(away_rpg, home_sp_era, home_bull_era, park_factor)

    # Predicted margin (positive = home wins by that margin)
    predicted_margin = round(home_scored - away_scored, 2)

    # MLB run margin is approximately normally distributed with σ≈3.1 runs
    sigma = 3.1

    # P(home covers -1.5) = P(home_margin > 1.5)
    home_cover = float(norm.sf(1.5, loc=predicted_margin, scale=sigma))
    away_cover = float(norm.cdf(1.5, loc=predicted_margin, scale=sigma))

    # Also compute P(away covers +1.5) = P(home_margin < 1.5)
    home_cover = round(max(0.05, min(0.95, home_cover)), 4)
    away_cover = round(max(0.05, min(0.95, away_cover)), 4)

    print(f"  [RunLine] Margin: {predicted_margin:+.1f} | Home cover: {home_cover:.1%} | Away cover: {away_cover:.1%}")
    return predicted_margin, home_cover, away_cover


def get_ats_edge(cover_prob, rl_odds):
    """
    Calculate ATS edge: model cover probability vs Vegas implied cover probability.
    rl_odds = American odds on that side covering (e.g. -130, +110).
    Returns edge as a percentage (positive = model likes this side more than Vegas).
    """
    if rl_odds is None:
        return None
    # Convert American odds to implied probability
    if rl_odds < 0:
        implied = abs(rl_odds) / (abs(rl_odds) + 100)
    else:
        implied = 100 / (rl_odds + 100)
    edge = (cover_prob - implied) * 100
    return round(edge, 1)


def get_rl_edge_score(ats_edge):
    """Convert ATS edge % to 1-10 score, same scale as ML edge score."""
    if ats_edge is None:
        return None
    abs_e = abs(ats_edge)
    if abs_e >= 20: return 10
    if abs_e >= 16: return 9
    if abs_e >= 13: return 8
    if abs_e >= 10: return 7
    if abs_e >= 8:  return 6
    if abs_e >= 6:  return 5
    if abs_e >= 4:  return 4
    if abs_e >= 2:  return 3
    return 2


def get_vegas_line(home_name, away_name, all_odds, home_wpct=0.500, away_wpct=0.500):
    """
    Look up Vegas odds for a specific game from the all_odds dict.
    Tries exact match first, then fuzzy team name matching.
    Falls back to win-percentage implied probability if no odds found.
    Returns dict with home_prob_novig, is_live, and run line fields.
    """
    if not all_odds:
        total = home_wpct + away_wpct
        home_implied = (home_wpct / total) if total > 0 else 0.5
        return {"home_prob_novig": round(home_implied, 4), "home_moneyline": -110, "away_moneyline": -110, "is_live": False}

    # Try exact key match
    result = all_odds.get((home_name, away_name))
    if result:
        return result

    # Try partial name matching (team names may differ slightly)
    home_key = home_name.split()[-1].lower()
    away_key = away_name.split()[-1].lower()
    for (h, a), v in all_odds.items():
        if home_key in h.lower() and away_key in a.lower():
            return v
        if home_key in a.lower() and away_key in h.lower():
            # Swapped — flip the probabilities
            flipped = dict(v)
            flipped["home_prob_novig"] = 1 - v.get("home_prob_novig", 0.5)
            return flipped

    # No match — fallback
    total = home_wpct + away_wpct
    home_implied = (home_wpct / total) if total > 0 else 0.5
    return {"home_prob_novig": round(home_implied, 4), "is_live": False}


def get_lineup_strength(team_id, team_name, fallback_strength):
    """
    Feature 3 — Starting Lineup Awareness.
    Fetches today's confirmed batting order from MLB API.
    Uses batch player lookup to get season OPS for each batter.
    Scales average OPS to a 1-10 lineup strength score.
    Falls back to team-average lineup_strength if lineup not yet posted.
    """
    global _lineup_cache
    if team_id in _lineup_cache:
        return _lineup_cache[team_id]

    today = today_et()
    data  = _get(f"{MLB_BASE}/schedule",
                 {"sportId": 1, "date": today, "hydrate": "lineups,team"})
    if not data:
        _lineup_cache[team_id] = fallback_strength
        return fallback_strength

    batting_order = []
    for date_entry in data.get("dates", []):
        for game in date_entry.get("games", []):
            for side in ("home", "away"):
                t = game.get("teams", {}).get(side, {})
                if t.get("team", {}).get("id") == team_id:
                    lineups    = game.get("lineups", {})
                    side_key   = "homePlayers" if side == "home" else "awayPlayers"
                    side_lineup = lineups.get(side_key, [])
                    if side_lineup:
                        batting_order = side_lineup
                    break

    if not batting_order:
        _lineup_cache[team_id] = fallback_strength
        return fallback_strength

    # Batch-fetch season hitting stats for the batting order
    player_ids = ",".join(str(p["id"]) for p in batting_order[:9] if p.get("id"))
    if not player_ids:
        _lineup_cache[team_id] = fallback_strength
        return fallback_strength

    try:
        pdata = _get(f"{MLB_BASE}/people",
                     {"personIds": player_ids,
                      "hydrate": f"stats(group=hitting,type=season,season={CURRENT_SEASON})"})
        ops_list = []
        for person in (pdata or {}).get("people", []):
            for stat_group in person.get("stats", []):
                for split in stat_group.get("splits", []):
                    ops_str = split.get("stat", {}).get("ops")
                    if ops_str:
                        ops_list.append(float(ops_str))
                        break

        if not ops_list:
            _lineup_cache[team_id] = fallback_strength
            return fallback_strength

        avg_ops  = sum(ops_list) / len(ops_list)
        # Scale: OPS 0.600 → 1, OPS 0.900 → 10
        strength = round(min(max((avg_ops - 0.600) / 0.300 * 10, 1), 10), 2)
        print(f"  [{team_name}] Live lineup: avg OPS={avg_ops:.3f} → strength={strength}")
        _lineup_cache[team_id] = strength
        return strength

    except Exception as e:
        print(f"  [{team_name}] Lineup lookup failed: {e}")
        _lineup_cache[team_id] = fallback_strength
        return fallback_strength


def get_bullpen_fatigue(team_id, team_name):
    """
    Feature 4 — Bullpen Fatigue Tracking.
    Sums non-starter pitch counts from the last 3 completed games.
    Returns a 0.0–1.0 fatigue score (higher = more tired bullpen).
    """
    global _bullpen_fatigue_cache
    if team_id in _bullpen_fatigue_cache:
        return _bullpen_fatigue_cache[team_id]

    today      = datetime.datetime.now(ET).date()
    start_date = (today - datetime.timedelta(days=4)).strftime("%Y-%m-%d")
    end_date   = (today - datetime.timedelta(days=1)).strftime("%Y-%m-%d")

    data = _get(f"{MLB_BASE}/schedule", {
        "sportId": 1, "startDate": start_date, "endDate": end_date,
        "teamId": team_id, "hydrate": "boxscore,team",
    })

    if not data:
        _bullpen_fatigue_cache[team_id] = 0.35
        return 0.35

    total_bullpen_pitches = 0
    games_counted         = 0

    for date_entry in data.get("dates", []):
        for game in date_entry.get("games", []):
            if game.get("status", {}).get("abstractGameState") != "Final":
                continue
            boxscore = game.get("boxscore", {})
            for side in ("home", "away"):
                t = boxscore.get("teams", {}).get(side, {})
                if t.get("team", {}).get("id") != team_id:
                    continue
                pitchers = t.get("pitchers", [])
                players  = t.get("players", {})
                for pid in pitchers[1:]:          # skip starter (index 0)
                    pstats = players.get(f"ID{pid}", {}).get("stats", {}).get("pitching", {})
                    total_bullpen_pitches += pstats.get("pitchesThrown", 0) or 0
                games_counted += 1

    if games_counted == 0:
        _bullpen_fatigue_cache[team_id] = 0.35
        return 0.35

    avg_pitches = total_bullpen_pitches / games_counted
    # 0 pitches/game → 0.0, 60+ pitches/game → 0.85
    fatigue = round(min(avg_pitches / 60, 1.0) * 0.85, 3)
    print(f"  [{team_name}] Bullpen: {avg_pitches:.0f} avg bp-pitches/game → fatigue={fatigue:.3f}")
    _bullpen_fatigue_cache[team_id] = fatigue
    return fatigue


def get_recent_pitcher_era(pitcher_id, season_era):
    """
    Feature 5 — Rolling Pitcher Form (last 5 starts).
    Computes ERA from the pitcher's most recent 5 game log entries,
    then returns a (blended_era, rolling_era) tuple.
    blended = 50/50 season + rolling for stability.
    rolling_era is returned separately so callers can compute the delta
    (rolling - season) as a trend feature for the ML model.
    """
    if not pitcher_id:
        return season_era, season_era

    try:
        data = _get(f"{MLB_BASE}/people/{pitcher_id}/stats",
                    {"stats": "gameLog", "group": "pitching",
                     "season": CURRENT_SEASON})
        if not data:
            return season_era, season_era

        splits = data.get("stats", [{}])[0].get("splits", [])
        recent = splits[-5:] if len(splits) >= 5 else splits
        if not recent:
            return season_era, season_era

        total_er, total_ip = 0, 0.0
        for game in recent:
            stat = game.get("stat", {})
            er   = stat.get("earnedRuns", 0) or 0
            ip_s = str(stat.get("inningsPitched", "0"))
            parts = ip_s.split(".")
            ip = int(parts[0]) + (int(parts[1]) / 3 if len(parts) > 1 and parts[1] else 0)
            total_er += er
            total_ip += ip

        if total_ip < 3:
            return season_era, season_era

        rolling_era = round((total_er / total_ip) * 9, 2)
        # 60% recent form, 40% season ERA — recent starts are more predictive
        blended     = round(season_era * 0.40 + rolling_era * 0.60, 2)
        return blended, rolling_era

    except Exception:
        return season_era, season_era


def _load_resolved_from_supabase():
    """Load resolved results from Supabase for model retraining."""
    try:
        import psycopg2, psycopg2.extras
        db_url = os.environ.get("DATABASE_URL", "")
        if not db_url:
            print("  [ML] DATABASE_URL not set", flush=True)
            return None

        # Try both port 5432 and 6543 (Supabase connection pooler)
        urls_to_try = [db_url]
        if ':5432/' in db_url:
            urls_to_try.append(db_url.replace(':5432/', ':6543/'))
        elif ':6543/' in db_url:
            urls_to_try.append(db_url.replace(':6543/', ':5432/'))

        conn = None
        for url in urls_to_try:
            for ssl_mode in ['require', 'prefer']:
                try:
                    conn = psycopg2.connect(
                        url,
                        sslmode=ssl_mode,
                        connect_timeout=15,
                        options='-c statement_timeout=30000'
                    )
                    port = '6543' if '6543' in url else '5432'
                    print(f"  [ML] Supabase connected (port={port}, ssl={ssl_mode})", flush=True)
                    break
                except Exception as e:
                    print(f"  [ML] Failed port={'6543' if '6543' in url else '5432'} ssl={ssl_mode}: {e}", flush=True)
            if conn:
                break

        if not conn:
            print("  [ML] All Supabase connection attempts failed", flush=True)
            return None

        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        # Phase 4.5 Step 5: team_wpct/opp_wpct/team_rpg/opp_rpg were MISSING from
        # this SELECT, which is why they were constant 0.0 in every retrain since
        # June 2 — row.get() returned None and fell through to the 0.0 default.
        # elo_diff added so Elo finally reaches the model instead of only the UI.
        cur.execute("""
            SELECT date, home_team, away_team, predicted_winner, actual_winner,
                   correct, confidence, model_edge, home_win_prob,
                   home_sp_era, away_sp_era, bullpen_era, opp_bullpen_era,
                   park_factor, temp, wind_speed, wind_dir_out, home,
                   pitcher_recent_delta, opp_pitcher_recent_delta,
                   team_wpct, opp_wpct, team_rpg, opp_rpg,
                   home_elo, away_elo, elo_diff,
                   vegas_implied, model_cohort
            FROM results
            WHERE actual_winner IS NOT NULL AND actual_winner != ''
            ORDER BY date ASC
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        if not rows:
            print("  [ML] No resolved rows found", flush=True)
            return None
        df = pd.DataFrame([dict(r) for r in rows])
        df = df.rename(columns={
            'home_team': 'Home Team', 'away_team': 'Away Team',
            'predicted_winner': 'Predicted Winner', 'actual_winner': 'Actual Winner',
            'correct': 'Correct?', 'confidence': 'Confidence',
            'model_edge': 'Model Edge', 'home_win_prob': 'Home Win %',
            'home_sp_era': 'pitcher_era', 'away_sp_era': 'opp_pitcher_era',
        })
        print(f"  [ML] Loaded {len(df)} resolved rows from Supabase", flush=True)
        return df
    except Exception as e:
        print(f"  [ML] Supabase load error: {e}", flush=True)
        return None


def _baseline_report(y_true, p_model, p_market=None, p_elo=None, label="holdout"):
    """
    Phase 4.5 Step 9 — permanent baseline comparison.

    THE most important addition in this phase. From June 2 to Aug 11 the model
    scored Brier 0.27-0.28 while the coin-flip baseline was 0.25 — i.e. it was
    WORSE THAN GUESSING for ten weeks — and nothing in the system said so.
    "Grade D" read as "early days" when it actually meant "negative skill".

    Skill score = 1 - (model_brier / baseline_brier).
    Positive = the model beats doing nothing. Negative = it does not.

    Returns a dict; also printed and persisted to pipeline_health.
    """
    import numpy as np
    y = np.asarray(y_true, dtype=float)
    p = np.clip(np.asarray(p_model, dtype=float), 0.0, 1.0)
    n = len(y)
    if n == 0:
        return {}

    base_rate = float(y.mean())
    brier_model    = float(np.mean((p - y) ** 2))
    brier_coinflip = float(np.mean((0.5 - y) ** 2))
    brier_baserate = float(np.mean((base_rate - y) ** 2))
    logloss = float(-np.mean(
        y * np.log(np.clip(p, 1e-6, 1)) + (1 - y) * np.log(np.clip(1 - p, 1e-6, 1))
    ))
    accuracy = float(np.mean((p >= 0.5) == (y == 1)))

    ref = brier_baserate if brier_baserate > 0 else 0.25
    skill = 1.0 - (brier_model / ref)

    # Bootstrap CI on the skill score — is it distinguishable from zero?
    rng = np.random.RandomState(42)
    boots = []
    for _ in range(400):
        idx = rng.randint(0, n, n)
        yb, pb = y[idx], p[idx]
        br = float(np.mean((pb - yb) ** 2))
        rb = float(np.mean((yb.mean() - yb) ** 2)) or 0.25
        boots.append(1.0 - (br / rb))
    ci_lo, ci_hi = float(np.percentile(boots, 2.5)), float(np.percentile(boots, 97.5))

    out = {
        "n": n, "brier_model": round(brier_model, 4),
        "brier_coinflip": round(brier_coinflip, 4),
        "brier_baserate": round(brier_baserate, 4),
        "logloss": round(logloss, 4), "accuracy": round(accuracy, 4),
        "base_rate": round(base_rate, 4),
        "skill_score": round(skill, 4),
        "skill_ci_low": round(ci_lo, 4), "skill_ci_high": round(ci_hi, 4),
        "min_prob": round(float(p.min()), 4), "max_prob": round(float(p.max()), 4),
        "pct_extreme": round(float(np.mean((p > 0.85) | (p < 0.15))), 4),
    }

    if p_market is not None:
        m = np.clip(np.asarray(p_market, dtype=float), 0.0, 1.0)
        if len(m) == n and np.isfinite(m).all():
            out["brier_market"] = round(float(np.mean((m - y) ** 2)), 4)
    if p_elo is not None:
        e = np.clip(np.asarray(p_elo, dtype=float), 0.0, 1.0)
        if len(e) == n and np.isfinite(e).all():
            out["brier_elo"] = round(float(np.mean((e - y) ** 2)), 4)

    verdict = "✅ REAL SKILL" if out["skill_ci_low"] > 0 else (
              "⚠️  ambiguous" if skill > 0 else "❌ WORSE THAN BASELINE")
    print(f"  [ML] ── Baselines ({label}, n={n}) ──")
    print(f"  [ML]   model    Brier {out['brier_model']:.4f} | acc {out['accuracy']:.1%} | logloss {out['logloss']:.4f}")
    print(f"  [ML]   coinflip Brier {out['brier_coinflip']:.4f}")
    print(f"  [ML]   baserate Brier {out['brier_baserate']:.4f}  (rate {out['base_rate']:.3f})")
    if "brier_market" in out:
        print(f"  [ML]   market   Brier {out['brier_market']:.4f}")
    if "brier_elo" in out:
        print(f"  [ML]   elo-only Brier {out['brier_elo']:.4f}")
    print(f"  [ML]   SKILL {out['skill_score']:+.4f}  95% CI [{out['skill_ci_low']:+.4f}, {out['skill_ci_high']:+.4f}]  {verdict}")
    print(f"  [ML]   prob range {out['min_prob']:.3f}–{out['max_prob']:.3f}, {out['pct_extreme']:.1%} extreme")
    return out


def _deployment_gate(report: dict) -> tuple:
    """
    Phase 4.5 Step 10 — hard gate. A model that fails cannot silently ship.
    Returns (passed: bool, reasons: list[str]).
    """
    reasons = []
    if not report:
        return False, ["no evaluation report"]
    if report.get("n", 0) < 100:
        reasons.append(f"holdout n={report.get('n',0)} < 100")
    if report.get("brier_model", 1) >= report.get("brier_baserate", 0.25):
        reasons.append(f"Brier {report.get('brier_model')} does not beat baserate {report.get('brier_baserate')}")
    if report.get("skill_ci_low", -1) <= 0:
        reasons.append(f"skill CI lower bound {report.get('skill_ci_low')} not > 0")
    if report.get("max_prob", 1) > 0.90:
        reasons.append(f"max_prob {report.get('max_prob')} > 0.90 (overconfident)")
    if report.get("min_prob", 0) < 0.10:
        reasons.append(f"min_prob {report.get('min_prob')} < 0.10 (overconfident)")
    return (len(reasons) == 0), reasons


def retrain_model_if_needed():
    """
    Daily Learning — retrain whenever new resolved games have been added.
    Loads training data from Supabase first, falls back to CSV.
    Minimum of 10 resolved games to start learning.
    """
    global ml_model

    # Try Supabase first (permanent storage, survives Render restarts)
    resolved = None
    df_supabase = _load_resolved_from_supabase()
    if df_supabase is not None and len(df_supabase) >= 10:
        resolved = df_supabase
        current_count = len(resolved)
        print(f"  [ML] Using Supabase data: {current_count} resolved games")
    else:
        # Fall back to CSV
        if not os.path.exists(TRACKER_CSV):
            print("  [ML] No tracker CSV found")
            return
        df       = _safe_read_tracker_csv()
        resolved = df[df["Actual Winner"].notna() & (df["Actual Winner"].str.strip() != "")]
        current_count = len(resolved)
        print(f"  [ML] Using CSV data: {current_count} resolved games")

    # Read last-retrain game count from flag file (format: "YYYY-MM-DD:count")
    last_count = 0
    if os.path.exists(RETRAIN_FLAG):
        try:
            content = open(RETRAIN_FLAG).read().strip()
            parts   = content.split(":")
            last_count = int(parts[1]) if len(parts) >= 2 else 0
        except Exception:
            last_count = 0

    if current_count < 10:
        print(f"  [ML] Daily learning: {current_count} resolved games (need ≥10 to start)")
        return
    if current_count <= last_count:
        print(f"  [ML] Daily learning: no new resolved games since last run ({last_count} games)")
        return

    print(f"  [ML] Daily learning: {current_count - last_count} new game(s) → retraining…")
    try:
        # ── Feature columns — Vegas implied (home_prob_novig) intentionally excluded ──
        # Vegas implied is used ONLY for edge computation after prediction.
        # Including it in training creates circular reasoning: model learns to
        # agree with Vegas rather than finding independent signal.
        # ── Phase 4.5: cohort-aware feature gating ───────────────────────
        # Cohort A (1,153 rows, pre-Aug-12) has NULL team quality — unfixable,
        # feature_snapshot was never populated. combined.fillna(0) turns those
        # NULLs into 0.0, so if we enable these features now the model sees:
        #     0.0    -> "this is an old row"
        #     0.556  -> "this is a new row"
        # It would learn the cohort boundary, not team strength. That is worse
        # than the original bug: a real-looking feature encoding pure artifact.
        #
        # So the team-quality features stay OUT until Cohort B alone can carry
        # a training set. They switch on automatically at the threshold.
        COHORT_B_MIN = 200
        cohort_b_rows = 0
        try:
            if "model_cohort" in resolved.columns:
                cohort_b_rows = int((resolved["model_cohort"] == "B").sum())
        except Exception:
            cohort_b_rows = 0

        TEAM_QUALITY_FEATS = ["team_wpct", "opp_wpct", "team_rpg", "opp_rpg", "elo_diff"]

        ml_feat_cols = [
            "pitcher_era", "opp_pitcher_era",
            "bullpen_era", "opp_bullpen_era",
            "park_factor", "temp", "wind_speed", "wind_dir_out",
            "home",
            "pitcher_recent_delta", "opp_pitcher_recent_delta",
        ]

        if cohort_b_rows >= COHORT_B_MIN:
            ml_feat_cols = TEAM_QUALITY_FEATS + ml_feat_cols
            print(f"  [ML] ✅ Cohort B has {cohort_b_rows} rows (≥{COHORT_B_MIN}) — "
                  f"team quality + Elo ENABLED, training on Cohort B only")
            # Train only on rows that genuinely have the features
            resolved = resolved[resolved["model_cohort"] == "B"].copy()
        else:
            need = COHORT_B_MIN - cohort_b_rows
            print(f"  [ML] Cohort B: {cohort_b_rows}/{COHORT_B_MIN} rows "
                  f"({need} more needed) — team quality + Elo held back")
            print(f"  [ML]   (enabling now would teach the model the cohort "
                  f"boundary, not team strength)")

        # ── Step 6: Feature health gate — retrain-blocking ────────────────
        # The June 2–Aug 11 failure was four features silently constant at 0.0
        # for ten weeks. RandomForest reported exactly 0.0 importance every
        # single retrain and nothing was watching. These assertions make that
        # class of bug impossible to ship again.
        def _assert_feature_health(df: "pd.DataFrame", cols: list) -> list:
            """Returns list of failures. Empty list = healthy."""
            problems = []
            RANGES = {
                "team_wpct": (0.0, 1.0), "opp_wpct": (0.0, 1.0),
                "team_rpg": (0.0, 15.0), "opp_rpg": (0.0, 15.0),
                "pitcher_era": (0.0, 30.0), "opp_pitcher_era": (0.0, 30.0),
                "bullpen_era": (0.0, 30.0), "opp_bullpen_era": (0.0, 30.0),
                "elo_diff": (-600.0, 600.0),
                "park_factor": (0.5, 2.0), "home": (0.0, 1.0),
            }
            for c in cols:
                if c not in df.columns:
                    problems.append(f"{c}: ABSENT from dataframe")
                    continue
                s = pd.to_numeric(df[c], errors="coerce")
                if s.notna().sum() == 0:
                    problems.append(f"{c}: all null")
                    continue
                if float(s.std(skipna=True) or 0.0) == 0.0:
                    problems.append(f"{c}: CONSTANT (std=0) — not reaching model")
                miss = float(s.isna().mean())
                if miss > 0.50:
                    problems.append(f"{c}: {miss:.0%} missing")
                if c in RANGES:
                    lo, hi = RANGES[c]
                    bad = int(((s < lo) | (s > hi)).sum())
                    if bad > 0:
                        problems.append(f"{c}: {bad} values outside [{lo},{hi}]")
            return problems

        # Build training rows — only use rows with real feature data
        live_rows = []
        skipped = 0
        for _, row in resolved.iterrows():
            try:
                actual = str(row.get("Actual Winner", "") or "").strip()
                home   = str(row.get("Home Team", "") or "").strip()
                if not actual or not home:
                    continue

                feat = {}
                non_zero = 0
                for f in ml_feat_cols:
                    val = row.get(f, None)
                    try:
                        fval = float(val) if val is not None and str(val).strip() not in ('', 'nan', 'None') else 0.0
                    except (ValueError, TypeError):
                        fval = 0.0
                    feat[f] = fval
                    if fval != 0.0:
                        non_zero += 1

                if non_zero < 3:
                    skipped += 1
                    continue

                feat["result"]        = 1 if actual == home else 0
                feat["_date"]         = str(row.get("date", "") or "")
                live_rows.append(feat)
            except Exception:
                continue

        print(f"  [ML] Training rows: {len(live_rows)} valid, {skipped} skipped (no features)")

        if not live_rows:
            print(f"  [ML] Retrain skipped — no rows with real feature data")
            return

        available_feats = ml_feat_cols

        live_df   = pd.DataFrame(live_rows).fillna(0)
        hist_path = os.path.join(SCRIPT_DIR, "mlb_games_history.csv")
        # ── Step 6: Feature health gate ───────────────────────────────────
        # Runs on live_df BEFORE the fillna(0) below, because fillna would
        # mask exactly the "missing column" failure we're trying to catch.
        health_problems = _assert_feature_health(live_df, ml_feat_cols)
        if health_problems:
            print(f"  [ML] ⚠️  FEATURE HEALTH: {len(health_problems)} problem(s)")
            for p in health_problems:
                print(f"       • {p}")
            fatal = [p for p in health_problems if "CONSTANT" in p or "ABSENT" in p]
            if fatal:
                print(f"  [ML] ❌ RETRAIN ABORTED — {len(fatal)} feature(s) not reaching the model.")
                print(f"  [ML]    Keeping existing model. Fix the data path before retraining.")
                return
        else:
            print(f"  [ML] ✅ Feature health: all {len(ml_feat_cols)} features live")

        if os.path.exists(hist_path):
            hist        = pd.read_csv(hist_path).fillna(0)
            common_cols = list(set(live_df.columns) & set(hist.columns))
            combined    = pd.concat([hist[common_cols], live_df[common_cols]], ignore_index=True)
        else:
            combined = live_df

        combined = combined.fillna(0)

        # ── Fix 1: Chronological split — prevents temporal leakage ────────────
        # Random split allows future games to appear in training, leaking
        # team stats that weren't known at prediction time.
        # Sort by date and split in time order.
        if "_date" in combined.columns:
            combined = combined.sort_values("_date").reset_index(drop=True)

        features = combined.drop(columns=["result", "_date"], errors="ignore")
        target   = combined["result"]

        n         = len(features)
        tr_end    = int(n * 0.70)
        cal_end   = int(n * 0.85)

        X_tr  = features.iloc[:tr_end]
        y_tr  = target.iloc[:tr_end]
        X_cal = features.iloc[tr_end:cal_end]
        y_cal = target.iloc[tr_end:cal_end]
        X_te  = features.iloc[cal_end:]
        y_te  = target.iloc[cal_end:]

        print(f"  [ML] Chronological split: {tr_end} train | {cal_end-tr_end} cal | {n-cal_end} test")

        # ── Fix 2: Regularized RF — prevents overfitting on small data ─────────
        # At ~150 rows with 15 features, default RF memorizes training examples.
        # max_depth=5 and min_samples_leaf=8 cap complexity appropriately.
        cal_rf = _train_calibrated(
            RandomForestClassifier(
                n_estimators=200,
                max_depth=5,             # was 8 — shallower for small data
                min_samples_leaf=8,      # require 8+ samples per leaf
                max_features="sqrt",
                random_state=42,
            ),
            X_tr, y_tr, X_cal, y_cal,
        )

        # ── Fix 3: Regularized XGBoost — prevents overfitting on small data ────
        # Default XGBoost (300 trees, depth 6) severely overfits at 150 rows.
        # reg_alpha (L1) + reg_lambda (L2) + min_child_weight constrain complexity.
        cal_xgb = _train_calibrated(
            XGBClassifier(
                n_estimators=100,        # was 300 — fewer trees for small data
                max_depth=3,             # was 6 — much shallower
                learning_rate=0.05,
                subsample=0.8,
                colsample_bytree=0.8,
                reg_alpha=0.1,           # L1 regularization — new
                reg_lambda=1.0,          # L2 regularization — new
                min_child_weight=5,      # require 5+ samples per leaf — new
                use_label_encoder=False,
                eval_metric="logloss",
                random_state=42,
                verbosity=0,
            ),
            X_tr, y_tr, X_cal, y_cal,
        )

        # ── Also retrain LightGBM ─────────────────────────────────────────────
        lgbm_path = os.path.join(SCRIPT_DIR, "lgbm_model.pkl")
        retrain_lightgbm(X_tr.values.tolist(), y_tr.tolist(), lgbm_path)

        # ── Phase 2: Logistic Regression — adds diversity to ensemble ─────────
        # Tree models capture nonlinear interactions.
        # LogReg captures linear relationships they miss.
        # L2 regularization (C=0.1) prevents overfitting on small data.
        try:
            from sklearn.linear_model import LogisticRegression
            from sklearn.preprocessing import StandardScaler
            scaler  = StandardScaler()
            X_tr_sc = scaler.fit_transform(X_tr)
            X_cal_sc = scaler.transform(X_cal)
            cal_lr  = _train_calibrated(
                LogisticRegression(C=0.1, max_iter=1000, random_state=42),
                X_tr_sc, y_tr, X_cal_sc, y_cal,
            )
            lr_scaler = scaler
            print(f"  [ML] Logistic regression trained (4th ensemble model)")
        except Exception as e:
            cal_lr    = None
            lr_scaler = None
            print(f"  [ML] LogReg skipped: {e}")

        # ── Phase 2: Optimized ensemble weights ───────────────────────────────
        # Learn the best blend weights on the calibration set
        # instead of using fixed 35/40/25 weights
        optimal_weights = [0.30, 0.35, 0.20, 0.15]  # default: RF/XGB/LGBM/LR
        try:
            from scipy.optimize import minimize
            import numpy as np

            lgbm_model_loaded = None
            try:
                import lightgbm as lgb
                if os.path.exists(lgbm_path):
                    with open(lgbm_path, "rb") as f:
                        lgbm_model_loaded = pickle.load(f)
            except Exception:
                pass

            # Get calibration set probabilities from each model
            rf_cal_probs  = cal_rf.predict_proba(X_cal)[:, 1]
            xgb_cal_probs = cal_xgb.predict_proba(X_cal)[:, 1]
            lgbm_cal_probs = lgbm_model_loaded.predict_proba(X_cal)[:, 1] if lgbm_model_loaded else rf_cal_probs
            lr_cal_probs  = cal_lr.predict_proba(X_cal_sc)[:, 1] if cal_lr else rf_cal_probs

            probs_list = [rf_cal_probs, xgb_cal_probs, lgbm_cal_probs, lr_cal_probs]
            y_cal_arr  = np.array(y_cal)

            def ensemble_brier(weights):
                w = np.array(weights)
                w = np.clip(w, 0.05, 0.60)  # each model min 5%, max 60%
                w = w / w.sum()
                blended = sum(wi * pi for wi, pi in zip(w, probs_list))
                return float(np.mean((blended - y_cal_arr) ** 2))

            result = minimize(
                ensemble_brier,
                [0.30, 0.35, 0.20, 0.15],
                method='Nelder-Mead',
                options={'maxiter': 500, 'xatol': 0.001}
            )
            raw_w = np.clip(result.x, 0.05, 0.60)
            optimal_weights = list(raw_w / raw_w.sum())
            print(f"  [ML] Optimized weights — RF:{optimal_weights[0]:.2f} XGB:{optimal_weights[1]:.2f} LGBM:{optimal_weights[2]:.2f} LR:{optimal_weights[3]:.2f}")
        except Exception as e:
            print(f"  [ML] Weight optimization skipped: {e}")

        # ── Phase 2: Feature importance tracking ──────────────────────────────
        try:
            import psycopg2
            db_url = os.environ.get("DATABASE_URL", "")
            if db_url:
                importances = dict(zip(features.columns, cal_rf.base_model.feature_importances_
                    if hasattr(cal_rf.base_model, 'feature_importances_') else [0]*len(features.columns)))
                # Average across all RF trees for stability
                if hasattr(cal_rf.base_model, 'estimators_'):
                    import numpy as np
                    avg_imp = np.mean([tree.feature_importances_ for tree in cal_rf.base_model.estimators_], axis=0)
                    importances = dict(zip(features.columns, avg_imp.tolist()))
                conn = psycopg2.connect(db_url, sslmode='require', connect_timeout=10)
                cur  = conn.cursor()
                cur.execute("""
                    INSERT INTO feature_importance_history (date, importances_json, model_version)
                    VALUES (%s, %s, %s)
                """, (today_et(), json.dumps(importances), f"v{current_count}"))
                conn.commit(); cur.close(); conn.close()
                top3 = sorted(importances.items(), key=lambda x: x[1], reverse=True)[:3]
                print(f"  [ML] Feature importance saved — top 3: {', '.join(f'{k}:{v:.3f}' for k,v in top3)}")
        except Exception as e:
            print(f"  [ML] Feature importance save skipped: {e}")

        # ── Steps 9 + 10: Baseline comparison and deployment gate ─────────────
        accuracy = 0.0
        gate_report = {}
        try:
            import numpy as np
            if len(y_te) >= 10:
                # Blend the ensemble the same way inference does
                w = optimal_weights
                rf_te  = cal_rf.predict_proba(X_te)[:, 1]
                xgb_te = cal_xgb.predict_proba(X_te)[:, 1]
                try:
                    lr_te = cal_lr.predict_proba(scaler.transform(X_te))[:, 1] if cal_lr else rf_te
                except Exception:
                    lr_te = rf_te
                blend_te = (rf_te * w[0] + xgb_te * w[1] + rf_te * w[2] + lr_te * w[3])

                # Elo-only baseline from elo_diff, if the feature is present
                p_elo = None
                if "elo_diff" in X_te.columns:
                    ed = pd.to_numeric(X_te["elo_diff"], errors="coerce").fillna(0.0).values
                    p_elo = 1.0 / (1.0 + np.power(10.0, (-ed) / 400.0))

                gate_report = _baseline_report(
                    np.asarray(y_te, dtype=float), blend_te,
                    p_market=None, p_elo=p_elo, label="holdout",
                )
                accuracy = gate_report.get("accuracy", 0.0)

                train_blend = cal_rf.predict_proba(X_tr)[:, 1]
                train_brier = float(np.mean((train_blend - np.asarray(y_tr, dtype=float)) ** 2))
                gap = gate_report.get("brier_model", 0) - train_brier
                print(f"  [ML]   train Brier {train_brier:.4f} | overfit gap {gap:+.4f}"
                      + ("  ⚠️ overfitting" if gap > 0.03 else ""))

                passed, reasons = _deployment_gate(gate_report)
                if passed:
                    print(f"  [ML] ✅ DEPLOYMENT GATE PASSED — model has demonstrable skill")
                else:
                    print(f"  [ML] ⚠️  DEPLOYMENT GATE FAILED:")
                    for r in reasons:
                        print(f"  [ML]      • {r}")
                    print(f"  [ML]    Model still saved (Cohort B is accumulating), but it does")
                    print(f"  [ML]    NOT yet beat the baseline. Do not build Best Bets on it.")
                gate_report["gate_passed"] = passed
                gate_report["gate_reasons"] = reasons
            else:
                print(f"  [ML] Test set too small ({len(y_te)} games) — evaluation skipped")
        except Exception as e:
            print(f"  [ML] Baseline evaluation failed: {e}")

        # Persist the report so it surfaces outside the logs
        try:
            import psycopg2, json as _json
            db_url = os.environ.get("DATABASE_URL", "")
            if db_url and gate_report:
                _c = psycopg2.connect(db_url, sslmode="require", connect_timeout=10)
                _cur = _c.cursor()
                _cur.execute("""
                    UPDATE pipeline_health
                    SET baseline_report = %s,
                        skill_score     = %s,
                        gate_passed     = %s
                    WHERE date = %s
                """, (_json.dumps(gate_report), gate_report.get("skill_score"),
                      bool(gate_report.get("gate_passed", False)), today_et()))
                _c.commit(); _cur.close(); _c.close()
        except Exception as e:
            print(f"  [ML] Baseline persist skipped: {e}")

        ensemble = {
            "rf":              cal_rf,
            "xgb":             cal_xgb,
            "lr":              cal_lr,
            "lr_scaler":       lr_scaler,
            "lgbm_path":       lgbm_path,
            "weights":         optimal_weights,
            "feature_cols":    list(features.columns),
        }
        ml_model = ensemble

        with open(MODEL_CACHE, "wb") as f:
            pickle.dump(ensemble, f)

        with open(RETRAIN_FLAG, "w") as f:
            f.write(f"{today_et()}:{current_count}")

        print(f"  [ML] Ensemble updated (RF+XGB calibrated) — {accuracy:.1%} holdout accuracy "
              f"({len(combined)} training rows from {current_count} resolved games)")

    except Exception as e:
        print(f"  [ML] Retrain failed: {e}")


def get_team_stats(team_id, team_name="Team", savant_batting=None):
    """Fetch live team stats: win pct, runs per game, bullpen ERA, batting avg."""
    if team_id in _team_stats_cache:
        return _team_stats_cache[team_id]

    defaults = {
        "wpct": 0.500, "rpg": 4.5, "bullpen_era": 4.00,
        "hard_hit": 0.38, "defense": 5.0,
        "lineup_strength": 5.0, "vs_hand_split": 0.0
    }

    standings = fetch_standings()
    if team_id in standings:
        defaults["wpct"] = standings[team_id]["wpct"]
        defaults["rpg"] = standings[team_id]["rpg"]

    # Team pitching ERA (used as overall pitching proxy incl. bullpen)
    pitch_data = _get(f"{MLB_BASE}/teams/{team_id}/stats",
                      params={"stats": "season", "group": "pitching", "season": CURRENT_SEASON})
    if pitch_data:
        splits = pitch_data.get("stats", [{}])[0].get("splits", [])
        if splits:
            era_str = splits[0].get("stat", {}).get("era", None)
            if era_str:
                defaults["bullpen_era"] = float(era_str)

    # Team hitting avg (proxy for lineup strength)
    hit_data = _get(f"{MLB_BASE}/teams/{team_id}/stats",
                    params={"stats": "season", "group": "hitting", "season": CURRENT_SEASON})
    if hit_data:
        splits = hit_data.get("stats", [{}])[0].get("splits", [])
        if splits:
            stat = splits[0].get("stat", {})
            avg_str = stat.get("avg", None)
            if avg_str:
                avg = float(avg_str)
                # Scale batting avg to lineup_strength 1-10
                defaults["lineup_strength"] = round(min(max((avg - 0.200) / (0.290 - 0.200) * 10, 1), 10), 2)

    # Replace placeholder hard_hit with real Baseball Savant data if available
    if savant_batting and team_name in savant_batting:
        sv = savant_batting[team_name]
        defaults["hard_hit"]    = sv["hard_hit"]
        defaults["barrel_rate"] = sv["barrel_rate"]
        defaults["xwoba_off"]   = sv["xwoba"]

    print(f"  [{team_name}] wpct={defaults['wpct']:.3f}, rpg={defaults['rpg']}, bullpen_era={defaults['bullpen_era']}")
    _team_stats_cache[team_id] = defaults
    return defaults


def get_pitcher_stats(pitcher_id, pitcher_name="TBD"):
    """Fetch live pitcher ERA with home/away splits. Falls back to overall ERA."""
    if not pitcher_id or pitcher_name == "TBD":
        return {"home_era": 4.00, "away_era": 4.50}

    if pitcher_id in _pitcher_stats_cache:
        return _pitcher_stats_cache[pitcher_id]

    result = {"home_era": 4.00, "away_era": 4.50}

    # Overall season ERA as base
    season_data = _get(f"{MLB_BASE}/people/{pitcher_id}/stats",
                       params={"stats": "season", "group": "pitching", "season": CURRENT_SEASON})
    if season_data:
        splits = (season_data.get("stats") or [{}])[0].get("splits", [])
        if splits:
            era_str = splits[0].get("stat", {}).get("era")
            if era_str:
                overall_era = float(era_str)
                result["home_era"] = overall_era
                result["away_era"] = overall_era

    # Home/away splits
    split_data = _get(f"{MLB_BASE}/people/{pitcher_id}/stats",
                      params={"stats": "homeAndAway", "group": "pitching", "season": CURRENT_SEASON})
    if split_data:
        for split in (split_data.get("stats") or [{}])[0].get("splits", []):
            era_str = split.get("stat", {}).get("era")
            if not era_str:
                continue
            is_home = split.get("isHome", None)
            split_code = split.get("split", {}).get("code", "")
            if is_home is True or split_code == "H":
                result["home_era"] = float(era_str)
            elif is_home is False or split_code == "A":
                result["away_era"] = float(era_str)

    # Feature 5 — blend each split ERA with rolling last-5-start ERA (50/50)
    # Also capture the raw rolling ERA to compute a trend delta for the ML model
    season_home_era = result["home_era"]
    season_away_era = result["away_era"]
    result["home_era"], home_rolling = get_recent_pitcher_era(pitcher_id, season_home_era)
    result["away_era"], away_rolling = get_recent_pitcher_era(pitcher_id, season_away_era)
    # Positive delta = pitcher struggling recently vs season avg; negative = hot streak
    result["home_recent_delta"] = round(home_rolling - season_home_era, 2)
    result["away_recent_delta"] = round(away_rolling - season_away_era, 2)

    # Phase 4.4: Pitcher momentum ratio
    # recent/season < 0.85 = trending better than average (buy signal)
    # recent/season > 1.15 = trending worse than average (sell signal)
    # Avoids division by zero when season ERA is 0 (start of season)
    if season_home_era and season_home_era > 0:
        result["momentum_ratio"] = round(home_rolling / season_home_era, 3)
    else:
        result["momentum_ratio"] = 1.0  # neutral — no data yet

    momentum = result["momentum_ratio"]
    if momentum < 0.85:
        result["momentum_signal"] = "hot"    # pitcher trending better
    elif momentum > 1.15:
        result["momentum_signal"] = "cold"   # pitcher trending worse
    else:
        result["momentum_signal"] = "neutral"

    # Blend ERA with Baseball Savant xwOBA for a more predictive quality metric
    sv = get_savant_pitcher_stats(pitcher_id)
    if sv:
        result["home_era"] = xwoba_to_era_adj(sv["xwoba"], result["home_era"])
        result["away_era"] = xwoba_to_era_adj(sv["xwoba"], result["away_era"])
        result["xwoba"]    = sv["xwoba"]
        print(f"  [{pitcher_name}] home_era={result['home_era']}, away_era={result['away_era']}  (xwOBA={sv['xwoba']:.3f})")
    else:
        print(f"  [{pitcher_name}] home_era={result['home_era']}, away_era={result['away_era']}")
    _pitcher_stats_cache[pitcher_id] = result
    return result


def fetch_todays_schedule():
    """
    Fetch today's MLB schedule with probable pitchers and team IDs.
    Returns a list of game dicts.
    """
    today = today_et()
    print(f"Fetching MLB schedule for {today}...")
    data = _get(f"{MLB_BASE}/schedule",
                params={"sportId": 1, "date": today,
                        "hydrate": "probablePitcher,team,venue"})
    if not data:
        return None

    games = []
    for date_entry in data.get("dates", []):
        for game in date_entry.get("games", []):
            home_team = game["teams"]["home"]["team"]
            away_team = game["teams"]["away"]["team"]
            home_prob = game["teams"]["home"].get("probablePitcher", {})
            away_prob = game["teams"]["away"].get("probablePitcher", {})

            games.append({
                "home_name": home_team.get("name", "Home Team"),
                "home_id": home_team.get("id"),
                "away_name": away_team.get("name", "Away Team"),
                "away_id": away_team.get("id"),
                "venue": game.get("venue", {}).get("name", "Unknown Park"),
                "home_pitcher_id": home_prob.get("id"),
                "home_pitcher_name": home_prob.get("fullName", "TBD"),
                "away_pitcher_id": away_prob.get("id"),
                "away_pitcher_name": away_prob.get("fullName", "TBD"),
                "game_date": game.get("gameDate", ""),
            })

    return games if games else None


# ============================================================
# B. WEATHER — Live via Open-Meteo (free, no API key required)
# ============================================================

# (latitude, longitude, CF bearing degrees from north, indoor/retractable)
BALLPARK_COORDS = {
    "Yankee Stadium":           (40.8296,  -73.9262,  40,  False),
    "Fenway Park":              (42.3467,  -71.0972,  65,  False),
    "Wrigley Field":            (41.9484,  -87.6553, 315,  False),
    "Dodger Stadium":           (34.0739, -118.2400,  10,  False),
    "Oracle Park":              (37.7786, -122.3893,  60,  False),
    "Chase Field":              (33.4453, -112.0667, 340,  True),
    "T-Mobile Park":            (47.5914, -122.3324, 315,  False),
    "Camden Yards":             (39.2838,  -76.6218,   5,  False),
    "Nationals Park":           (38.8730,  -77.0074,   5,  False),
    "Truist Park":              (33.8907,  -84.4678, 315,  False),
    "American Family Field":    (43.0280,  -87.9712,  20,  True),
    "Minute Maid Park":         (29.7573,  -95.3555, 280,  True),
    "Angel Stadium":            (33.8003, -117.8827, 350,  False),
    "Petco Park":               (32.7077, -117.1569, 330,  False),
    "Busch Stadium":            (38.6226,  -90.1928,  10,  False),
    "Great American Ball Park": (39.0974,  -84.5080,  40,  False),
    "Progressive Field":        (41.4962,  -81.6852, 345,  False),
    "Kauffman Stadium":         (39.0517,  -94.4803, 350,  False),
    "Target Field":             (44.9817,  -93.2781, 300,  False),
    "Comerica Park":            (42.3390,  -83.0485, 345,  False),
    "Globe Life Field":         (32.7474,  -97.0828,  20,  True),
    "Coors Field":              (39.7559, -104.9942, 340,  False),
    "loanDepot park":           (25.7781,  -80.2197, 325,  True),
    "Citizens Bank Park":       (39.9057,  -75.1665,   5,  False),
    "Citi Field":               (40.7571,  -73.8458, 330,  False),
    "PNC Park":                 (40.4469,  -80.0057, 350,  False),
    "Guaranteed Rate Field":    (41.8300,  -87.6339, 350,  False),
    "Oakland Coliseum":         (37.7516, -122.2005, 320,  False),
    "Tropicana Field":          (27.7683,  -82.6534,   0,  True),
    "Rogers Centre":            (43.6414,  -79.3894,   0,  True),
    "Sutter Health Park":       (38.5802, -121.5002, 350,  False),
}


def _angle_diff(a, b):
    """Smallest angular difference between two compass bearings."""
    return abs((a - b + 180) % 360 - 180)


def _wind_category(wind_deg, cf_bearing, indoor):
    """
    Classify wind as 'out' (helps hitters), 'in' (hurts), or 'neutral'.
    Meteorological convention: wind_deg is the direction the wind is COMING FROM.
    Wind blows OUT when it comes from behind home plate (≈ cf_bearing + 180°).
    Wind blows IN  when it comes from center field (≈ cf_bearing).
    """
    if indoor:
        return "neutral"
    out_source = (cf_bearing + 180) % 360
    if _angle_diff(wind_deg, out_source) < 45:
        return "out"
    if _angle_diff(wind_deg, cf_bearing) < 45:
        return "in"
    return "neutral"


def get_weather(ballpark, game_date_utc=""):
    """
    Fetch live ballpark weather from Open-Meteo (free, no API key required).
    Returns real temperature, wind speed/direction (relative to CF), and
    rain probability at game start time and mid-game (+3 h).
    Falls back to league-average conditions if the park is unknown.
    """
    global _weather_cache
    cache_key = f"{ballpark}:{game_date_utc[:13]}"
    if cache_key in _weather_cache:
        return _weather_cache[cache_key]

    default = {
        "temp_start": 70, "temp_mid": 68,
        "wind_start_speed": 7,  "wind_start_dir": "neutral",
        "wind_mid_speed":   6,  "wind_mid_dir":   "neutral",
        "rain_chance": 10,
    }

    coords = BALLPARK_COORDS.get(ballpark)
    if not coords:
        _weather_cache[cache_key] = default
        return default

    lat, lon, cf_bearing, indoor = coords

    # Parse game start hour (UTC); default to 23:00 (≈ 7 PM ET)
    game_hour_utc = 23
    if game_date_utc:
        try:
            dt = datetime.datetime.strptime(game_date_utc, "%Y-%m-%dT%H:%M:%SZ")
            game_hour_utc = dt.hour
        except Exception:
            pass

    try:
        today = today_et()
        r = requests.get(
            "https://api.open-meteo.com/v1/forecast",
            params={
                "latitude":         lat,
                "longitude":        lon,
                "hourly":           "temperature_2m,windspeed_10m,winddirection_10m,precipitation_probability",
                "temperature_unit": "fahrenheit",
                "windspeed_unit":   "mph",
                "timezone":         "UTC",
                "start_date":       today,
                "end_date":         today,
            },
            timeout=10,
        )
        d      = r.json()
        hours  = d.get("hourly", {})
        times  = hours.get("time", [])
        temps  = hours.get("temperature_2m", [])
        winds  = hours.get("windspeed_10m", [])
        wdirs  = hours.get("winddirection_10m", [])
        precs  = hours.get("precipitation_probability", [])

        # Map UTC hours to list indices
        start_idx = next(
            (i for i, t in enumerate(times) if f"T{game_hour_utc:02d}:00" in t),
            min(game_hour_utc, len(times) - 1)
        )
        mid_game_hour = (game_hour_utc + 3) % 24
        mid_idx = next(
            (i for i, t in enumerate(times) if f"T{mid_game_hour:02d}:00" in t),
            min(start_idx + 3, len(times) - 1)
        )

        def _v(lst, idx, fallback):
            return lst[idx] if lst and idx < len(lst) else fallback

        if indoor:
            t = _v(temps, start_idx, 72)
            result = {
                "temp_start": t, "temp_mid": t,
                "wind_start_speed": 0, "wind_start_dir": "neutral",
                "wind_mid_speed":   0, "wind_mid_dir":   "neutral",
                "rain_chance": 0,
            }
        else:
            ws  = _v(winds, start_idx, 7);   wm  = _v(winds, mid_idx, 6)
            wds = _v(wdirs, start_idx, 180);  wdm = _v(wdirs, mid_idx, 180)
            result = {
                "temp_start":       round(_v(temps, start_idx, 70), 1),
                "temp_mid":         round(_v(temps, mid_idx,   68), 1),
                "wind_start_speed": round(ws, 1),
                "wind_start_dir":   _wind_category(wds, cf_bearing, indoor),
                "wind_mid_speed":   round(wm, 1),
                "wind_mid_dir":     _wind_category(wdm, cf_bearing, indoor),
                "rain_chance":      _v(precs, start_idx, 10),
            }

        tag = "indoor/dome" if indoor else f"{result['wind_start_speed']}mph {result['wind_start_dir']}"
        print(f"  [{ballpark}] Weather: {result['temp_start']}°F, wind {tag}, rain {result['rain_chance']}%")
        _weather_cache[cache_key] = result
        return result

    except Exception as e:
        print(f"  [Weather] {ballpark} API error: {e} — using defaults")
        _weather_cache[cache_key] = default
        return default

def get_park_factor(ballpark):
    park_factors = {
        "Coors Field": 1.30,
        "Great American Ball Park": 1.15,
        "Yankee Stadium": 1.08,
        "Dodger Stadium": 0.96,
        "Oracle Park": 0.92,
        "Petco Park": 0.90,
        "T-Mobile Park": 0.93,
    }
    return park_factors.get(ballpark, 1.00)

def fetch_kalshi_odds():
    """
    Kalshi prediction market implied probabilities.
    Uses Kalshi REST API — cleaner signal than Vegas (no vig estimation needed).
    Returns dict keyed by (home_team, away_team) -> home_prob float.
    """
    global _kalshi_cache
    if _kalshi_cache:
        return _kalshi_cache

    api_key = os.environ.get("KALSHI_API_KEY", "").strip()
    if not api_key:
        print("  [Kalshi] KALSHI_API_KEY not set — skipping prediction market signal")
        return {}

    try:
        r = requests.get(
            "https://api.elections.kalshi.com/trade-api/v2/markets",
            params={"status": "open", "category": "sports", "limit": 200},
            headers={"Authorization": f"Bearer {api_key}"},
            timeout=10,
        )
        if r.status_code != 200:
            print(f"  [Kalshi] API returned {r.status_code}")
            return {}

        result = {}
        today  = today_et()
        for m in r.json().get("markets", []):
            title = m.get("title", "")
            # Match MLB game markets — "Will [Team] beat [Team]?"
            if "MLB" not in title and "baseball" not in title.lower():
                continue
            # Parse home/away from title
            for home, away in _mlb_team_pairs():
                if home in title and away in title:
                    prob = m.get("last_price", None)
                    if prob is not None:
                        result[(home, away)] = round(float(prob) / 100, 4)
                    break

        _kalshi_cache = result
        print(f"  [Kalshi] Prediction market odds loaded for {len(result)} games")
        return result

    except Exception as e:
        print(f"  [Kalshi] WARNING: {e}")
        return {}


def _mlb_team_pairs():
    """Known MLB team name pairs for Kalshi market parsing."""
    return [
        ("Yankees","Red Sox"),("Dodgers","Giants"),("Cubs","White Sox"),
        ("Mets","Yankees"),("Angels","Dodgers"),("Athletics","Giants"),
        # Add more as needed — Kalshi titles vary
    ]


def fetch_umpire_data(game_date):
    """
    Fetch home plate umpire assignments from MLB Stats API.
    Returns dict keyed by game_pk -> umpire_name.
    Umpires with wide zones hurt strikeout pitchers; tight zones hurt power arms.
    """
    try:
        r = requests.get(
            f"https://statsapi.mlb.com/api/v1/schedule",
            params={
                "sportId": 1,
                "date": game_date,
                "hydrate": "officials",
            },
            timeout=8,
        )
        if r.status_code != 200:
            return {}

        result = {}
        for date_entry in r.json().get("dates", []):
            for game in date_entry.get("games", []):
                officials = game.get("officials", [])
                for off in officials:
                    if off.get("officialType") == "Home Plate":
                        result[game["gamePk"]] = off.get("official", {}).get("fullName", "")
        return result
    except Exception as e:
        print(f"  [Umpire] Failed: {e}")
        return {}


# Umpire ERA factors — above 0 = hitter-friendly (wide zone), below = pitcher-friendly
# Source: Baseball Savant umpire scorecards (season averages)
UMPIRE_ERA_FACTOR = {
    "Laz Diaz":          0.18,   # Wide zone, hitter friendly
    "Angel Hernandez":   0.12,
    "CB Bucknor":        0.15,
    "Joe West":          0.08,
    "Hunter Wendelstedt":0.10,
    "Phil Cuzzi":        0.06,
    "Mark Carlson":     -0.05,
    "Doug Eddings":     -0.08,
    "Ted Barrett":      -0.06,
    "Nic Lentz":        -0.10,   # Tight zone, pitcher friendly
    "Jerry Layne":      -0.07,
    "Brian Gorman":      0.04,
    "Dan Iassogna":     -0.04,
    "Jim Reynolds":      0.09,
    "Bill Miller":       0.05,
    "Marty Foster":     -0.03,
    "Tim Timmons":       0.07,
    "Alfonso Marquez":   0.02,
    "Kerwin Danley":     0.03,
    "James Hoye":       -0.02,
}

def get_umpire_factor(umpire_name):
    """
    Returns ERA adjustment factor for umpire.
    Positive = hitter-friendly, negative = pitcher-friendly.
    """
    if not umpire_name:
        return 0.0
    factor = UMPIRE_ERA_FACTOR.get(umpire_name, 0.0)
    if factor != 0.0:
        print(f"  [Umpire] {umpire_name}: ERA adj {factor:+.2f}")
    return factor


def fetch_bullpen_usage(team_id, days=3):
    """
    Fetch bullpen pitcher usage from last N days via MLB Stats API.
    Returns total pitches thrown by relief pitchers — higher = more fatigued.
    """
    try:
        from datetime import datetime, timedelta
        end_date   = datetime.strptime(today_et(), "%Y-%m-%d")
        start_date = end_date - timedelta(days=days)

        r = requests.get(
            f"https://statsapi.mlb.com/api/v1/teams/{team_id}/stats",
            params={
                "stats":       "gameLog",
                "group":       "pitching",
                "startDate":   start_date.strftime("%Y-%m-%d"),
                "endDate":     end_date.strftime("%Y-%m-%d"),
                "sportId":     1,
            },
            timeout=8,
        )
        if r.status_code != 200:
            return 0

        total_pitches = 0
        for split in r.json().get("stats", [{}])[0].get("splits", []):
            stat = split.get("stat", {})
            # Only relief pitchers (not starters)
            if split.get("positionAbbrev") in ("RP", "CL"):
                total_pitches += int(stat.get("numberOfPitches", 0))
        return total_pitches

    except Exception as e:
        return 0


def get_series_context(home_team_id, away_team_id, game_date):
    """
    Determine game number within a series (1, 2, or 3+).
    Game 1 of series: managers save aces, fresh bullpens.
    Game 3+: fatigue accumulates, bullpen more taxed.
    Returns series_game_num (1-4).
    """
    try:
        from datetime import datetime, timedelta
        game_dt    = datetime.strptime(game_date, "%Y-%m-%d")
        start_date = (game_dt - timedelta(days=3)).strftime("%Y-%m-%d")

        r = requests.get(
            "https://statsapi.mlb.com/api/v1/schedule",
            params={
                "sportId":     1,
                "startDate":   start_date,
                "endDate":     game_date,
                "teamId":      home_team_id,
                "hydrate":     "team",
            },
            timeout=8,
        )
        if r.status_code != 200:
            return 1

        series_games = []
        for date_entry in r.json().get("dates", []):
            for game in date_entry.get("games", []):
                home = game.get("teams", {}).get("home", {}).get("team", {}).get("id")
                away = game.get("teams", {}).get("away", {}).get("team", {}).get("id")
                if {home, away} == {home_team_id, away_team_id}:
                    series_games.append(date_entry.get("date"))

        series_games = sorted(set(series_games))
        if game_date in series_games:
            return series_games.index(game_date) + 1
        return 1

    except Exception:
        return 1


def get_travel(team_name, prev_city=None):
    """
    Phase 4.5 — Time zone crossing fatigue.
    Replaces raw travel distance with actual TZ crossings.
    A west coast team flying east crosses 3 time zones — meaningful fatigue.
    A cross-country flight within the same TZ has minimal impact.
    Returns a fatigue score: 0 = no travel, 1-3 = TZ crossings.
    """
    # Each team's home time zone
    TEAM_TZ = {
        "New York Yankees":     "America/New_York",
        "New York Mets":        "America/New_York",
        "Boston Red Sox":       "America/New_York",
        "Tampa Bay Rays":       "America/New_York",
        "Baltimore Orioles":    "America/New_York",
        "Toronto Blue Jays":    "America/New_York",
        "Philadelphia Phillies":"America/New_York",
        "Pittsburgh Pirates":   "America/New_York",
        "Washington Nationals": "America/New_York",
        "Atlanta Braves":       "America/New_York",
        "Miami Marlins":        "America/New_York",
        "Cincinnati Reds":      "America/New_York",
        "Cleveland Guardians":  "America/New_York",
        "Detroit Tigers":       "America/New_York",
        "Chicago White Sox":    "America/Chicago",
        "Chicago Cubs":         "America/Chicago",
        "Kansas City Royals":   "America/Chicago",
        "Milwaukee Brewers":    "America/Chicago",
        "Minnesota Twins":      "America/Chicago",
        "Houston Astros":       "America/Chicago",
        "Texas Rangers":        "America/Chicago",
        "St. Louis Cardinals":  "America/Chicago",
        "Colorado Rockies":     "America/Denver",
        "Arizona Diamondbacks": "America/Phoenix",
        "Los Angeles Angels":   "America/Los_Angeles",
        "Los Angeles Dodgers":  "America/Los_Angeles",
        "Athletics":            "America/Los_Angeles",
        "Oakland Athletics":    "America/Los_Angeles",
        "Seattle Mariners":     "America/Los_Angeles",
        "San Diego Padres":     "America/Los_Angeles",
        "San Francisco Giants": "America/Los_Angeles",
    }

    # UTC offset for each zone (standard time — close enough for fatigue calc)
    TZ_OFFSET = {
        "America/New_York":    -5,
        "America/Chicago":     -6,
        "America/Denver":      -7,
        "America/Phoenix":     -7,
        "America/Los_Angeles": -8,
    }

    if not prev_city or team_name not in TEAM_TZ or prev_city not in TEAM_TZ:
        return 1  # default: minor travel assumed

    home_tz = TEAM_TZ[team_name]
    away_tz = TEAM_TZ[prev_city]

    if home_tz == away_tz:
        return 0  # same time zone — no fatigue

    home_offset = TZ_OFFSET.get(home_tz, -6)
    away_offset = TZ_OFFSET.get(away_tz, -6)
    crossings   = abs(home_offset - away_offset)
    return crossings  # 1, 2, or 3 time zone crossings


# ============================================================
# C. ADVANCED MODEL
# ============================================================

def advanced_model(inputs):
    (
        team_wpct, opp_wpct,
        team_rpg, opp_rpg,
        pitcher_home_era, pitcher_away_era,
        opp_pitcher_home_era, opp_pitcher_away_era,
        bullpen_era, opp_bullpen_era,
        park_factor,
        umpire_favor,
        travel_miles, opp_travel_miles,
        bullpen_usage, opp_bullpen_usage,
        is_home,
        temp_start, temp_mid,
        wind_start_speed, wind_start_dir,
        wind_mid_speed, wind_mid_dir,
        rain_chance,
        team_hard_hit, opp_hard_hit,
        team_defense, opp_defense,
        team_lineup_strength, opp_lineup_strength,
        team_vs_hand_split, opp_vs_hand_split
    ) = inputs

    home_adv = 0.04 if is_home else 0

    team_pitcher_era = pitcher_home_era if is_home else pitcher_away_era
    opp_pitcher_era = opp_pitcher_home_era if not is_home else opp_pitcher_away_era
    era_diff = (opp_pitcher_era - team_pitcher_era) / 10

    bullpen_diff = (opp_bullpen_era - bullpen_era) / 10
    bullpen_fatigue = (opp_bullpen_usage - bullpen_usage) * 0.5

    park_adj = (park_factor - 1.00) * 0.5
    umpire_adj = umpire_favor * 0.03
    travel_adj = (opp_travel_miles - travel_miles) / 3000

    def weather_effect(temp, wind_speed, wind_dir):
        effect = 0
        if temp >= 85:
            effect += 0.05
        elif temp <= 55:
            effect -= 0.05
        if wind_dir == "out":
            effect += wind_speed * 0.01
        elif wind_dir == "in":
            effect -= wind_speed * 0.01
        return effect

    weather_avg = (
        weather_effect(temp_start, wind_start_speed, wind_start_dir) +
        weather_effect(temp_mid, wind_mid_speed, wind_mid_dir)
    ) / 2

    rain_adj = -rain_chance * 0.001
    wpct_diff = team_wpct - opp_wpct
    rpg_diff = (team_rpg - opp_rpg) / 10
    hard_hit_diff = (team_hard_hit - opp_hard_hit) * 0.5
    defense_diff = (team_defense - opp_defense) * 0.5
    lineup_diff = (team_lineup_strength - opp_lineup_strength) * 0.7
    split_diff = (team_vs_hand_split - opp_vs_hand_split) * 0.6

    score = (
        wpct_diff * 2.0 +
        era_diff * 1.8 +
        bullpen_diff * 1.2 +
        bullpen_fatigue +
        rpg_diff * 1.2 +
        home_adv +
        park_adj +
        umpire_adj +
        travel_adj +
        weather_avg +
        rain_adj +
        hard_hit_diff +
        defense_diff +
        lineup_diff +
        split_diff
    )

    return 1 / (1 + math.exp(-score))


# ============================================================
# D. MACHINE LEARNING MODEL
# ============================================================

def load_ml_model():
    """
    Load the RF + XGBoost calibrated ensemble.
    Checks for a cached pickle first (fast, includes any daily retraining).
    Falls back to training from the historical CSV if no cache exists.
    Returns a dict {"rf": calibrated_rf, "xgb": calibrated_xgb, "feature_cols": [...]}
    or None if no data is available.
    """
    if os.path.exists(MODEL_CACHE):
        try:
            with open(MODEL_CACHE, "rb") as f:
                model = pickle.load(f)
            # Accept both old single-model cache and new ensemble dict
            if isinstance(model, dict) and "rf" in model:
                print("  [ML] Loaded cached RF+XGB ensemble from disk")
                return model
            # Old cache format — discard and retrain below
        except Exception:
            pass

    try:
        data = pd.read_csv(os.path.join(SCRIPT_DIR, "mlb_games_history.csv")).fillna(0)
    except Exception:
        return None

    # Chronological sort if date column present
    if "date" in data.columns:
        data = data.sort_values("date").reset_index(drop=True)

    features = data.drop(columns=["result", "date"], errors="ignore")
    target   = data["result"]

    # Chronological 3-way split — no temporal leakage
    n      = len(features)
    tr_end  = int(n * 0.70)
    cal_end = int(n * 0.85)
    X_tr  = features.iloc[:tr_end];  y_tr  = target.iloc[:tr_end]
    X_cal = features.iloc[tr_end:cal_end]; y_cal = target.iloc[tr_end:cal_end]
    X_te  = features.iloc[cal_end:]; y_te  = target.iloc[cal_end:]

    print("  [ML] Training RF + XGBoost ensemble on historical data (regularized)…")
    cal_rf  = _train_calibrated(
        RandomForestClassifier(
            n_estimators=200, max_depth=5,
            min_samples_leaf=8, max_features="sqrt", random_state=42,
        ),
        X_tr, y_tr, X_cal, y_cal,
    )
    cal_xgb = _train_calibrated(
        XGBClassifier(
            n_estimators=100, max_depth=3, learning_rate=0.05,
            subsample=0.8, colsample_bytree=0.8,
            reg_alpha=0.1, reg_lambda=1.0, min_child_weight=5,
            use_label_encoder=False, eval_metric="logloss",
            random_state=42, verbosity=0,
        ),
        X_tr, y_tr, X_cal, y_cal,
    )

    ensemble = {"rf": cal_rf, "xgb": cal_xgb, "feature_cols": list(features.columns)}

    if len(y_te) >= 5:
        accuracy = sum(cal_rf.predict(X_te) == y_te) / len(y_te)
        print(f"  [ML] Ensemble ready — {accuracy:.1%} holdout accuracy ({len(data)} games)")
    else:
        print(f"  [ML] Ensemble ready ({len(data)} games, test set too small to evaluate)")

    try:
        with open(MODEL_CACHE, "wb") as f:
            pickle.dump(ensemble, f)
    except Exception:
        pass

    return ensemble

ml_model = load_ml_model()


# ============================================================
# FINAL COMBINED PROBABILITY
# ============================================================

def get_platoon_advantage(home_pitcher_id, away_pitcher_id, home_team_id, away_team_id):
    """
    Platoon advantage — L/R handedness matchup adjustment.
    LHB vs RHP and RHB vs LHP both carry ~3-5% win probability shift.
    Returns a float adjustment (-0.05 to +0.05) to add to home win probability.
    Positive = home team has platoon advantage.
    """
    try:
        def get_pitcher_hand(pitcher_id):
            if not pitcher_id:
                return None
            r = requests.get(f"{MLB_BASE}/people/{pitcher_id}", timeout=5)
            if r.status_code != 200:
                return None
            p = r.json().get("people", [{}])[0]
            return p.get("pitchHand", {}).get("code")  # 'L' or 'R'

        def get_team_batter_hand_split(team_id):
            """Returns fraction of lineup that bats left."""
            r = requests.get(
                f"{MLB_BASE}/teams/{team_id}/roster?rosterType=active",
                timeout=5
            )
            if r.status_code != 200:
                return 0.5
            roster = r.json().get("roster", [])
            batters = [p for p in roster if p.get("position", {}).get("type") != "Pitcher"]
            if not batters:
                return 0.5
            left_bats = sum(1 for p in batters
                           if p.get("person", {}).get("batSide", {}).get("code") in ("L", "S"))
            return left_bats / len(batters)

        home_hand   = get_pitcher_hand(home_pitcher_id)
        away_hand   = get_pitcher_hand(away_pitcher_id)
        home_l_pct  = get_team_batter_hand_split(home_team_id)   # home batters vs away pitcher
        away_l_pct  = get_team_batter_hand_split(away_team_id)   # away batters vs home pitcher

        # Platoon edge: LHB vs LHP is a disadvantage (~3%), LHB vs RHP is neutral/slight advantage
        # RHB vs LHP is slight advantage; switch hitters are neutral
        # Net adjustment to home win prob:
        home_adv = 0.0

        # Away pitcher vs home batters
        if away_hand == 'L':
            # Left-handed pitchers suppress left-handed batters
            home_adv -= (home_l_pct - 0.5) * 0.06
        elif away_hand == 'R':
            home_adv += (home_l_pct - 0.5) * 0.04

        # Home pitcher vs away batters
        if home_hand == 'L':
            home_adv += (away_l_pct - 0.5) * 0.06
        elif home_hand == 'R':
            home_adv -= (away_l_pct - 0.5) * 0.04

        home_adv = round(max(-0.05, min(0.05, home_adv)), 4)
        if abs(home_adv) > 0.005:
            print(f"  [Platoon] Home adj: {home_adv:+.3f} "
                  f"(home_pitcher={home_hand}, away_pitcher={away_hand}, "
                  f"home_l_pct={home_l_pct:.0%}, away_l_pct={away_l_pct:.0%})")
        return home_adv

    except Exception as e:
        print(f"  [Platoon] WARNING: {e}")
        return 0.0


def final_win_probability(adv_prob, ml_prob, vegas_prob=None, vegas_is_live=False,
                          platoon_adj=0.0, kalshi_prob=None, series_game=1):
    """
    Blend four signal sources with platoon and series context adjustments.
    Live Vegas: 18% weight. Kalshi: 7% when available. Recent form prioritized.
    Series game 3+: slight regression toward 50% (variance increases in tired series).
    """
    # Dynamic weights
    w_vegas  = (0.18 if vegas_is_live else 0.08) if vegas_prob  is not None else 0.0
    w_kalshi = 0.07                                if kalshi_prob is not None else 0.0
    w_model  = 1.0 - w_vegas - w_kalshi
    w_each   = w_model / 2.0

    combined = (
        adv_prob  * w_each +
        ml_prob   * w_each +
        (vegas_prob  * w_vegas  if vegas_prob  is not None else 0) +
        (kalshi_prob * w_kalshi if kalshi_prob is not None else 0)
    )

    # Platoon adjustment (capped at ±4%)
    combined = combined + platoon_adj

    # Series context — game 3+ teams are more evenly matched (fatigue equalizes)
    if series_game >= 3:
        combined = combined * 0.97 + 0.50 * 0.03

    return round(max(0.02, min(0.98, combined)) * 100, 2)


# ============================================================
# MOBILE APP JSON EXPORT
# ============================================================

_MOBILE_KEYS = [
    "Game Time", "Home Team", "Away Team", "Venue", "Weather",
    "Home Starting Pitcher", "Away Starting Pitcher",
    "Home Win %", "Away Win %",
    # Phase 4: Elo + momentum
    "Home Elo", "Away Elo", "Elo Win Prob",
    "Home Pitcher Momentum", "Away Pitcher Momentum",
    "Home Momentum Ratio", "Away Momentum Ratio",
    "Home Runs / Game", "Away Runs / Game",
    "Home Bullpen ERA", "Away Bullpen ERA",
    "Home SP ERA (at Home)", "Away SP ERA (on Road)",
    "Park Factor", "Adv Model %", "ML Model %",
    "Vegas Implied %", "Vegas Is Live", "Kalshi Is Live", "Model Edge",
    "Home Win Probability", "Predicted Winner", "Confidence",
    "HasFullPrediction",
    # Spread / Run line
    "Spread Pick", "Spread Line", "Spread Odds",
    "Spread Cover %", "Spread Edge", "Spread Edge Score", "Run Line Live",
    "Predicted Margin", "Home Cover %", "Away Cover %",
    "Home RL Line", "Away RL Line", "Home RL Odds", "Away RL Odds",
    # Sportsbook intelligence
    "Sharp Money Signal", "Model Agrees Sharp",
    "RLM Detected", "RLM Sharp Side", "RLM Label", "RLM Strength",
    "Steam Detected", "Steam Label", "Steam Velocity",
    "Trap Detected", "Trap Label",
    "Public Home Pct", "Line Movement",
    # Phase 3: Bookmaker tiers
    "Sharp Consensus", "Square Consensus", "Book Divergence", "Sharp Square Alert",
    # Engagement
    "Is Lock Of Day", "Is Dangerous Dog",
    # Confidence intervals (Phase 2)
    "CI Low", "CI High", "CI Width",
    # Other
    "Series Game", "Lineup Confirmed",
]


def _save_predictions_json(games_list, lock_game=None, dog_game=None):
    """Export today's game predictions as JSON for the mobile app."""
    today_str = today_et()

    def clean(g):
        return {k: g.get(k) for k in _MOBILE_KEYS}

    conf_rank = {"High": 0, "Medium": 1}
    best_bets = sorted(
        [g for g in games_list if g.get("Confidence") in ("High", "Medium")],
        key=lambda g: (conf_rank.get(g.get("Confidence", ""), 99),
                       -abs(g.get("Home Win Probability", 50))),
    )[:6]

    payload = {
        "date":          today_str,
        "generated_at":  datetime.datetime.now().isoformat(timespec="seconds"),
        "best_bets":     [clean(g) for g in best_bets],
        "games":         [clean(g) for g in games_list],
        "lock_of_day":   clean(lock_game) if lock_game else None,
        "dangerous_dog": clean(dog_game)  if dog_game  else None,
    }
    with open(PREDICTIONS_JSON, "w") as fh:
        json.dump(payload, fh, indent=2, default=str)
    print(f"Predictions JSON saved → {PREDICTIONS_JSON}")

    # BUGFIX 2026-07-16: pass the UNSLIMMED games to the DB writer.
    # payload["games"] has been through clean(), which keeps only _MOBILE_KEYS —
    # that silently stripped "_feature_snapshot" and every training-only field,
    # so feature_snapshot wrote NULL on all 737 rows from Jun 1 onward.
    # The mobile JSON stays slim; the DB write gets the full dicts.
    _save_predictions_supabase(payload, raw_games=games_list)


def _save_predictions_supabase(payload, raw_games=None):
    """
    Write today's predictions to the predictions table AND the results table.

    raw_games: the UNSLIMMED game dicts. payload["games"] has passed through
    clean(), which keeps only _MOBILE_KEYS and therefore strips
    "_feature_snapshot" plus every training-only field. Passing the raw dicts
    separately is what makes feature_snapshot actually populate — it wrote
    NULL on all 1,072 pre-repair rows, which is why historical backfill of the
    dead team-quality features was impossible.
    """
    import psycopg2, psycopg2.extras

    db_url  = os.environ.get("DATABASE_URL", "")
    date_str = payload.get("date", "")

    # Use the full dicts for the DB write when available.
    db_payload = payload
    if raw_games:
        db_payload = dict(payload)
        db_payload["games"] = raw_games

    if not db_url or not date_str:
        print("  [Supabase] ❌ Missing DATABASE_URL or date — skipping predictions write")
        _save_game_features_to_results(db_payload)
        return

    # Write predictions table via psycopg2 (same connection that works for results)
    try:
        conn = psycopg2.connect(db_url, sslmode='require', connect_timeout=15)
        cur  = conn.cursor()

        games_json     = json.dumps(payload.get("games",     []), default=str)
        best_bets_json = json.dumps(payload.get("best_bets", []), default=str)
        lock_json      = json.dumps(payload.get("lock_of_day"),   default=str) if payload.get("lock_of_day")   else None
        dog_json       = json.dumps(payload.get("dangerous_dog"), default=str) if payload.get("dangerous_dog") else None
        generated_at   = payload.get("generated_at", datetime.datetime.now().isoformat())

        cur.execute("""
            INSERT INTO predictions (date, games, best_bets, generated_at, lock_of_day, dangerous_dog)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (date) DO UPDATE SET
                games         = EXCLUDED.games,
                best_bets     = EXCLUDED.best_bets,
                generated_at  = EXCLUDED.generated_at,
                lock_of_day   = EXCLUDED.lock_of_day,
                dangerous_dog = EXCLUDED.dangerous_dog
        """, (date_str, games_json, best_bets_json, generated_at, lock_json, dog_json))

        conn.commit()
        cur.close()
        conn.close()
        print(f"  [Supabase] ✅ Predictions saved for {date_str} via psycopg2")

    except Exception as e:
        print(f"  [Supabase] ❌ Predictions save failed: {e}")

    # 2. Save each game's features to results table
    _save_game_features_to_results(db_payload)


def _save_game_features_to_results(payload):
    """Write today's game features to results table so model can retrain on them."""
    try:
        import psycopg2
        db_url = os.environ.get("DATABASE_URL", "")
        if not db_url:
            return

        date  = payload.get("date", "")
        games = payload.get("games", [])
        if not date or not games:
            return

        def sf(v):
            try: return float(v) if v is not None and str(v).strip() not in ('','nan','None') else None
            except: return None

        conn = psycopg2.connect(db_url, sslmode='require', connect_timeout=15)
        cur  = conn.cursor()
        inserted = 0
        updated  = 0

        for g in games:
            home = g.get('Home Team', '')
            away = g.get('Away Team', '')
            if not home or not away:
                continue

            # Try update first
            cur.execute("""
                UPDATE results SET
                    home_sp_era      = %s,
                    away_sp_era      = %s,
                    bullpen_era      = %s,
                    opp_bullpen_era  = %s,
                    park_factor      = %s,
                    home_win_prob    = %s,
                    model_edge       = %s,
                    confidence       = %s,
                    predicted_winner = %s,
                    game_time        = %s,
                    spread_pick      = %s,
                    spread_line      = %s,
                    spread_cover_pct = %s,
                    spread_edge      = %s,
                    feature_snapshot = %s,
                    team_wpct        = %s,
                    opp_wpct         = %s,
                    team_rpg         = %s,
                    opp_rpg          = %s,
                    home_elo         = %s,
                    away_elo         = %s,
                    elo_diff         = %s,
                    vegas_implied    = %s,
                    model_cohort     = 'B'
                WHERE date=%s AND home_team=%s AND away_team=%s
            """, (
                sf(g.get('Home SP ERA (at Home)')),
                sf(g.get('Away SP ERA (on Road)')),
                sf(g.get('Home Bullpen ERA')),
                sf(g.get('Away Bullpen ERA')),
                sf(g.get('Park Factor')),
                sf(g.get('Home Win Probability')),
                sf(g.get('Model Edge')),
                g.get('Confidence') or None,
                g.get('Predicted Winner') or None,
                g.get('Game Time') or None,
                g.get('Spread Pick') or None,
                g.get('Spread Line') or None,
                sf(g.get('Spread Cover %')),
                sf(g.get('Spread Edge')),
                json.dumps(g.get('_feature_snapshot')) if g.get('_feature_snapshot') else None,
                # Phase 4.5 Step 3 — team quality + Elo. These values were
                # always present in the game dict; they were simply never
                # persisted, which is why the training SELECT found nothing
                # and four features sat constant at 0.0 for ten weeks.
                sf(g.get('Home Win %')),
                sf(g.get('Away Win %')),
                sf(g.get('Home Runs / Game')),
                sf(g.get('Away Runs / Game')),
                sf(g.get('Home Elo')),
                sf(g.get('Away Elo')),
                (sf(g.get('Home Elo')) - sf(g.get('Away Elo')))
                    if (sf(g.get('Home Elo')) is not None and sf(g.get('Away Elo')) is not None) else None,
                sf(g.get('Vegas Implied %')),
                date, home, away
            ))

            if cur.rowcount > 0:
                updated += 1
            else:
                # Insert new row
                _he, _ae = sf(g.get('Home Elo')), sf(g.get('Away Elo'))
                cur.execute("""
                    INSERT INTO results
                        (date, home_team, away_team, predicted_winner, confidence,
                         model_edge, home_win_prob, game_time,
                         home_sp_era, away_sp_era, bullpen_era, opp_bullpen_era,
                         park_factor, feature_snapshot,
                         team_wpct, opp_wpct, team_rpg, opp_rpg,
                         home_elo, away_elo, elo_diff, vegas_implied, model_cohort)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                            %s,%s,%s,%s,%s,%s,%s,%s,'B')
                    ON CONFLICT (date, home_team, away_team) DO UPDATE SET
                        predicted_winner = EXCLUDED.predicted_winner,
                        confidence       = EXCLUDED.confidence,
                        model_edge       = EXCLUDED.model_edge,
                        home_win_prob    = EXCLUDED.home_win_prob,
                        game_time        = EXCLUDED.game_time,
                        home_sp_era      = EXCLUDED.home_sp_era,
                        away_sp_era      = EXCLUDED.away_sp_era,
                        bullpen_era      = EXCLUDED.bullpen_era,
                        opp_bullpen_era  = EXCLUDED.opp_bullpen_era,
                        park_factor      = EXCLUDED.park_factor,
                        feature_snapshot = EXCLUDED.feature_snapshot,
                        team_wpct        = EXCLUDED.team_wpct,
                        opp_wpct         = EXCLUDED.opp_wpct,
                        team_rpg         = EXCLUDED.team_rpg,
                        opp_rpg          = EXCLUDED.opp_rpg,
                        home_elo         = EXCLUDED.home_elo,
                        away_elo         = EXCLUDED.away_elo,
                        elo_diff         = EXCLUDED.elo_diff,
                        vegas_implied    = EXCLUDED.vegas_implied,
                        model_cohort     = 'B'
                """, (
                    date, home, away,
                    g.get('Predicted Winner') or None,
                    g.get('Confidence') or None,
                    sf(g.get('Model Edge')),
                    sf(g.get('Home Win Probability')),
                    g.get('Game Time') or None,
                    sf(g.get('Home SP ERA (at Home)')),
                    sf(g.get('Away SP ERA (on Road)')),
                    sf(g.get('Home Bullpen ERA')),
                    sf(g.get('Away Bullpen ERA')),
                    sf(g.get('Park Factor')),
                    json.dumps(g.get('_feature_snapshot')) if g.get('_feature_snapshot') else None,
                    sf(g.get('Home Win %')),
                    sf(g.get('Away Win %')),
                    sf(g.get('Home Runs / Game')),
                    sf(g.get('Away Runs / Game')),
                    _he, _ae,
                    (_he - _ae) if (_he is not None and _ae is not None) else None,
                    sf(g.get('Vegas Implied %')),
                ))
                inserted += 1

        conn.commit()
        cur.close(); conn.close()
        print(f"  [Supabase] Results table: {inserted} inserted, {updated} updated with features for {date}")

    except Exception as e:
        print(f"  [Supabase] Results feature save failed: {e}", flush=True)


def _save_tracker_json(tracker_df):
    """Export tracker data as JSON for the mobile app (display columns only)."""
    display_cols = [
        "Date", "Game Time", "Home Team", "Away Team",
        "Predicted Winner", "Home Win %", "Model Edge",
        "Confidence", "Actual Winner", "Correct?",
    ]
    cols = [c for c in display_cols if c in tracker_df.columns]
    df   = tracker_df[cols].copy()

    records = df.to_dict(orient="records")
    records = [
        {k: (None if str(v) == "nan" else v) for k, v in r.items()}
        for r in records
    ]
    records.reverse()   # most-recent first

    def tier(conf_name):
        tier_recs = [r for r in records if r.get("Confidence") == conf_name]
        resolved  = [r for r in tier_recs
                     if r.get("Correct?") and str(r["Correct?"]) not in ("", "None")]
        correct   = sum(1 for r in resolved if "Correct" in str(r.get("Correct?", "")))
        return {
            "total":   len(resolved),
            "correct": correct,
            "pct":     round(correct / len(resolved) * 100) if resolved else 0,
        }

    resolved = [r for r in records
                if r.get("Correct?") and str(r["Correct?"]) not in ("", "None")]
    correct  = sum(1 for r in resolved if "Correct" in str(r.get("Correct?", "")))

    payload = {
        "records": records,
        "summary": {
            "total":   len(resolved),
            "correct": correct,
            "pct":     round(correct / len(resolved) * 100) if resolved else 0,
            "high":    tier("High"),
            "medium":  tier("Medium"),
            "low":     tier("Low"),
        },
    }
    with open(TRACKER_JSON, "w") as fh:
        json.dump(payload, fh, indent=2, default=str)
    print(f"Tracker JSON saved → {TRACKER_JSON}")


# ============================================================
# DAILY REPORT GENERATION
# ============================================================

def save_daily_report(games_list, lock_game=None, dog_game=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    script_dir  = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "mlb_daily_report.xlsx")
    today_str   = datetime.datetime.now(ET).strftime("%B %d, %Y")

    # ── Colour palette ──────────────────────────────────────────
    NAVY         = "1B3A6B"
    WHITE        = "FFFFFF"
    ROW_ODD      = "F4F6FA"
    ROW_EVEN     = "FFFFFF"
    MID_GRAY     = "C8CDD8"
    BEST_BET_BG  = "FFF8E1"   # warm gold tint for Best Bets rows
    EDGE_POS     = "1A6B3A"   # strong positive edge — dark green
    EDGE_NEG     = "8B1A1A"   # negative edge — dark red
    # Section header accent colours
    HDR_GAME     = "2C4F8A"
    HDR_PITCHER  = "1A6B50"
    HDR_STATS    = "5A3A8A"
    HDR_MODEL    = "7A4A00"
    HDR_PRED     = "1A5A1A"
    HDR_EDGE     = "8B4500"
    # Confidence
    GREEN_DARK   = "1A6B3A";  GREEN_LIGHT  = "D6F0E0"
    YELLOW_DARK  = "7A6000";  YELLOW_FILL  = "FFF3C4"
    RED_DARK     = "8B1A1A";  RED_FILL     = "FFD6D6"
    WINNER_FILL  = "C6EFCE"
    PROB_HIGH    = "C6EFCE"
    PROB_MID     = "DBEAFE"
    PROB_LOW     = "FFCCCC"

    thin   = Side(style="thin",   color=MID_GRAY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(c):  return PatternFill("solid", fgColor=c)
    def al(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ── Column definitions ──────────────────────────────────────
    # (key, display header, alignment, width, header-colour)
    # Widths are calibrated to the longest realistic data value in each column.
    columns = [
        ("Game Time",             "Game\nTime",             "center", 13, HDR_GAME),
        ("Home Team",             "Home Team",              "left",   22, HDR_GAME),
        ("Away Team",             "Away Team",              "left",   22, HDR_GAME),
        ("Venue",                 "Venue",                  "left",   27, HDR_GAME),
        ("Weather",               "Weather",                "left",   32, HDR_GAME),
        ("Home Starting Pitcher", "Home Starter",           "left",   22, HDR_PITCHER),
        ("Away Starting Pitcher", "Away Starter",           "left",   22, HDR_PITCHER),
        ("Home Win %",            "Home\nWin %",            "center", 11, HDR_STATS),
        ("Away Win %",            "Away\nWin %",            "center", 11, HDR_STATS),
        ("Home Runs / Game",      "Home\nRuns/G",           "center", 10, HDR_STATS),
        ("Away Runs / Game",      "Away\nRuns/G",           "center", 10, HDR_STATS),
        ("Home Bullpen ERA",      "Home\nBull ERA",         "center", 12, HDR_STATS),
        ("Away Bullpen ERA",      "Away\nBull ERA",         "center", 12, HDR_STATS),
        ("Home SP ERA (at Home)", "Home SP\nERA",           "center", 11, HDR_PITCHER),
        ("Away SP ERA (on Road)", "Away SP\nERA",           "center", 11, HDR_PITCHER),
        ("Park Factor",           "Park\nFactor",           "center", 10, HDR_GAME),
        ("Adv Model %",           "Adv\nModel %",           "center", 12, HDR_MODEL),
        ("ML Model %",            "ML\nModel %",            "center", 12, HDR_MODEL),
        ("Vegas Implied %",       "Vegas\nImplied %",       "center", 13, HDR_MODEL),
        ("Model Edge",            "Model\nEdge",            "center", 12, HDR_EDGE),
        ("Home Win Probability",  "Home Win\nProbability",  "center", 15, HDR_PRED),
        ("Predicted Winner",      "Predicted Winner",       "left",   22, HDR_PRED),
        ("Confidence",            "Confidence",             "center", 13, HDR_PRED),
    ]
    num_cols = len(columns)
    last_col = get_column_letter(num_cols)

    wb = Workbook()
    ws = wb.active
    ws.title = "MLB Predictions"

    # ── Identify best bets (High or Medium confidence) ──
    _conf_rank = {"High": 0, "Medium": 1}
    best_bets = sorted(
        [g for g in games_list if g.get("Confidence") in ("High", "Medium")],
        key=lambda g: (_conf_rank.get(g.get("Confidence", ""), 99),
                       -abs(g.get("Home Win Probability", 50))),
    )[:6]
    best_bet_keys = {(g["Home Team"], g["Away Team"]) for g in best_bets}

    current_row = 1

    # ── Title row ───────────────────────────────────────────────
    ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
    tc = ws[f"A{current_row}"]
    tc.value     = f"⚾   MLB Daily Prediction Report   —   {today_str}"
    tc.font      = Font(color=WHITE, bold=True, size=15, name="Calibri")
    tc.fill      = fill(NAVY)
    tc.alignment = al("center")
    ws.row_dimensions[current_row].height = 36
    current_row += 1

    # ── Best Bets section ────────────────────────────────────────
    GOLD         = "B8860B"
    GOLD_LIGHT   = "FFF8DC"
    GOLD_BORDER  = "DAA520"
    BEST_HDR_COL = "8B6914"

    if best_bets:
        # Section header
        ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
        bh = ws[f"A{current_row}"]
        bh.value     = "★  TODAY'S BEST BETS  —  Highest model edge with model agreement"
        bh.font      = Font(color=WHITE, bold=True, size=11, name="Calibri")
        bh.fill      = fill(GOLD)
        bh.alignment = al("left")
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Best bet sub-headers
        bb_cols = [
            ("Game Time", "Time",      9),
            ("Away Team", "Away",     20),
            ("Home Team", "Home",     20),
            ("Predicted Winner", "Pick",     20),
            ("Home Win Probability", "Win Prob", 11),
            ("Model Edge", "Edge",     10),
            ("Vegas Implied %", "Vegas %",  10),
            ("Confidence", "Confidence", 12),
            ("Weather", "Weather",   22),
        ]
        for ci, (_, hdr, w) in enumerate(bb_cols, start=1):
            cell = ws.cell(row=current_row, column=ci, value=hdr)
            cell.font      = Font(color=WHITE, bold=True, size=9, name="Calibri")
            cell.fill      = fill(BEST_HDR_COL)
            cell.alignment = al("center", wrap=True)
            cell.border    = Border(
                left=Side(style="thin", color=GOLD_BORDER),
                right=Side(style="thin", color=GOLD_BORDER),
                top=Side(style="thin", color=GOLD_BORDER),
                bottom=Side(style="thin", color=GOLD_BORDER),
            )
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for g in best_bets:
            edge      = g.get("Model Edge", 0)
            edge_str  = f"+{edge}%" if edge >= 0 else f"{edge}%"
            conf      = g.get("Confidence", "")
            conf_bg   = GREEN_LIGHT if conf == "High" else YELLOW_FILL
            conf_fg   = GREEN_DARK  if conf == "High" else YELLOW_DARK
            bb_vals   = [
                g.get("Game Time", ""),
                g.get("Away Team", ""),
                g.get("Home Team", ""),
                g.get("Predicted Winner", ""),
                f"{g.get('Home Win Probability', '')}%",
                edge_str,
                f"{g.get('Vegas Implied %', '')}%",
                conf,
                g.get("Weather", ""),
            ]
            for ci, val in enumerate(bb_vals, start=1):
                cell = ws.cell(row=current_row, column=ci, value=sanitize_cell(val))
                cell.border    = Border(
                    left=Side(style="thin", color=GOLD_BORDER),
                    right=Side(style="thin", color=GOLD_BORDER),
                    top=Side(style="thin", color=GOLD_BORDER),
                    bottom=Side(style="thin", color=GOLD_BORDER),
                )
                cell.alignment = al("center" if ci not in (2, 3, 4, 9) else "left")
                if ci == 4:
                    cell.fill = fill(WINNER_FILL)
                    cell.font = Font(name="Calibri", size=10, bold=True, color=GREEN_DARK)
                elif ci == 6:
                    cell.fill = fill(GREEN_LIGHT if edge >= 0 else RED_FILL)
                    cell.font = Font(name="Calibri", size=10, bold=True,
                                     color=EDGE_POS if edge >= 0 else EDGE_NEG)
                elif ci == 8:
                    cell.fill = fill(conf_bg)
                    cell.font = Font(name="Calibri", size=10, bold=True, color=conf_fg)
                else:
                    cell.fill = fill(GOLD_LIGHT)
                    cell.font = Font(name="Calibri", size=10)
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # Spacer
        ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
        ws.row_dimensions[current_row].height = 8
        current_row += 1

    # ── All-games section header ─────────────────────────────────
    ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
    ah = ws[f"A{current_row}"]
    ah.value     = "ALL GAMES TODAY"
    ah.font      = Font(color=WHITE, bold=True, size=10, name="Calibri")
    ah.fill      = fill(NAVY)
    ah.alignment = al("left")
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    # ── Column headers ───────────────────────────────────────────
    hdr_row = current_row
    for ci, (key, header, align, width, hdr_color) in enumerate(columns, start=1):
        cell = ws.cell(row=hdr_row, column=ci, value=header)
        cell.font      = Font(color=WHITE, bold=True, size=9, name="Calibri")
        cell.fill      = fill(hdr_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(left=thin, right=thin,
                                top=Side(style="medium", color=WHITE),
                                bottom=Side(style="medium", color=WHITE))
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[hdr_row].height = 32
    current_row += 1

    data_start_row = current_row

    # Number formats per column key
    NUM_FMT = {
        "Home Win %":            "0.0%",    # stored as 0.684 → displays 68.4%
        "Away Win %":            "0.0%",
        "Home Runs / Game":      "0.0",
        "Away Runs / Game":      "0.0",
        "Home Bullpen ERA":      "0.00",
        "Away Bullpen ERA":      "0.00",
        "Home SP ERA (at Home)": "0.00",
        "Away SP ERA (on Road)": "0.00",
        "Park Factor":           "0.00",
        "Adv Model %":           '0.0"%"',  # stored as 82.44 → displays 82.4%
        "ML Model %":            '0.0"%"',
        "Vegas Implied %":       '0.0"%"',
        "Home Win Probability":  '0.0"%"',
    }

    # ── Data rows ────────────────────────────────────────────────
    for game in games_list:
        ri      = current_row
        is_bb   = (game.get("Home Team"), game.get("Away Team")) in best_bet_keys
        row_bg  = BEST_BET_BG if is_bb else (ROW_ODD if ri % 2 == 1 else ROW_EVEN)
        conf    = game.get("Confidence", "Medium")
        home_wp = game.get("Home Win Probability", 50)
        edge    = game.get("Model Edge", 0)

        for ci, (key, _, align, __, ___) in enumerate(columns, start=1):
            val  = game.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=sanitize_cell(val))
            cell.border    = border
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = al(align)

            # Apply number format where defined
            if key in NUM_FMT:
                cell.number_format = NUM_FMT[key]

            if key == "Predicted Winner":
                cell.fill = fill(WINNER_FILL)
                cell.font = Font(name="Calibri", size=10, bold=True, color=GREEN_DARK)

            elif key == "Home Win Probability":
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.fill = fill(PROB_HIGH if home_wp >= 65 else
                                 PROB_LOW  if home_wp <= 40 else PROB_MID)

            elif key == "Confidence":
                if conf == "High":
                    cell.fill = fill(GREEN_LIGHT)
                    cell.font = Font(name="Calibri", size=10, bold=True, color=GREEN_DARK)
                elif conf == "Medium":
                    cell.fill = fill(YELLOW_FILL)
                    cell.font = Font(name="Calibri", size=10, bold=True, color=YELLOW_DARK)
                else:
                    cell.fill = fill(RED_FILL)
                    cell.font = Font(name="Calibri", size=10, bold=True, color=RED_DARK)

            elif key == "Model Edge":
                edge_disp = f"+{val}%" if isinstance(val, (int, float)) and val >= 0 else f"{val}%"
                cell.value = edge_disp
                if isinstance(val, (int, float)):
                    if val >= 5:
                        cell.fill = fill(GREEN_LIGHT)
                        cell.font = Font(name="Calibri", size=10, bold=True, color=EDGE_POS)
                    elif val <= -5:
                        cell.fill = fill(RED_FILL)
                        cell.font = Font(name="Calibri", size=10, bold=True, color=EDGE_NEG)
                    else:
                        cell.fill = fill(row_bg)
                        cell.font = Font(name="Calibri", size=10)
                else:
                    cell.fill = fill(row_bg)

            else:
                cell.fill = fill(row_bg)

            if is_bb and key not in ("Predicted Winner", "Home Win Probability",
                                     "Confidence", "Model Edge"):
                cell.font = Font(name="Calibri", size=10, bold=True)

        ws.row_dimensions[ri].height = 22
        current_row += 1

    # ── Freeze & filter ──────────────────────────────────────────
    ws.freeze_panes = f"A{data_start_row}"
    ws.auto_filter.ref = (
        f"A{hdr_row}:{last_col}{hdr_row + len(games_list)}"
    )

    # ── Legend ───────────────────────────────────────────────────
    legend_start = current_row + 1
    ws.merge_cells(f"A{legend_start}:{last_col}{legend_start}")
    lh = ws[f"A{legend_start}"]
    lh.value     = "Confidence & Colour Legend"
    lh.font      = Font(bold=True, size=10, name="Calibri", color=WHITE)
    lh.fill      = fill(NAVY)
    lh.alignment = al("left")
    ws.row_dimensions[legend_start].height = 20

    legend_items = [
        ("★ Best Bet",  GOLD_LIGHT,  GOLD,       "High or Medium confidence game with model edge ≥ 5% over Vegas implied — top picks of the day"),
        ("High",        GREEN_LIGHT, GREEN_DARK, "All three models agree within 8% — strongest conviction"),
        ("Medium",      YELLOW_FILL, YELLOW_DARK,"Models agree within 18% — reasonable edge, some caution"),
        ("Low",         RED_FILL,    RED_DARK,   "Models diverge more than 18% — conflicting signals, use caution"),
        ("Edge +5%+",   GREEN_LIGHT, EDGE_POS,   "Model predicts home team at least 5% more likely to win than Vegas implies"),
        ("Edge −5%−",   RED_FILL,    EDGE_NEG,   "Vegas implies home team 5%+ more likely than model does"),
    ]
    for offset, (label, bg, fg, desc) in enumerate(legend_items, start=1):
        row = legend_start + offset
        lc  = ws.cell(row=row, column=1, value=label)
        lc.fill      = fill(bg)
        lc.font      = Font(bold=True, size=10, name="Calibri", color=fg)
        lc.alignment = al("center")
        lc.border    = border
        ws.merge_cells(f"B{row}:{last_col}{row}")
        dc = ws.cell(row=row, column=2, value=desc)
        dc.font      = Font(size=10, name="Calibri")
        dc.alignment = al("left")
        dc.fill      = fill(ROW_ODD)
        ws.row_dimensions[row].height = 18

    # ── Column Definitions ───────────────────────────────────────
    def_start = legend_start + len(legend_items) + 2
    ws.merge_cells(f"A{def_start}:{last_col}{def_start}")
    dh = ws[f"A{def_start}"]
    dh.value     = "Column Definitions"
    dh.font      = Font(bold=True, size=10, name="Calibri", color=WHITE)
    dh.fill      = fill(NAVY)
    dh.alignment = al("left")
    ws.row_dimensions[def_start].height = 20

    definitions = [
        ("Game Time",               "The scheduled first pitch time in Eastern Time."),
        ("Weather",                 "Live ballpark conditions from Open-Meteo at game start: temperature, wind speed & direction relative to center field (OUT = helps hitters, IN = hurts), and rain %. Domes show 'dome'."),
        ("Home SP ERA (at Home)",   "Home starter's ERA in home games only this season."),
        ("Away SP ERA (on Road)",   "Away starter's ERA in road games only this season."),
        ("Park Factor",             "Ballpark scoring effect. >1.0 = hitter-friendly (Coors 1.30). <1.0 = pitcher-friendly (Petco 0.90)."),
        ("Adv Model %",             "Formula model: win %, runs/game, bullpen ERA, pitcher ERA, park factor, weather. 42% of final probability."),
        ("ML Model %",              "Random Forest trained on 7,300+ real MLB games. Learns patterns from history. 42% of final probability."),
        ("Vegas Implied %",         "Vig-removed consensus probability averaged across all sportsbooks. 16% weight when live, 10% from win% fallback."),
        ("Model Edge",              "Final blended model % minus Vegas implied %. Positive = model favours home team more than the market does. This is the key number for identifying value."),
        ("Home Win Probability",    "Final blended probability. Green ≥ 65%, blue 40–65%, red ≤ 40%."),
        ("Predicted Winner",        "Team predicted to win. Home team if probability ≥ 50%, away otherwise."),
        ("Confidence",              "Model agreement. High = within 8% (strongest). Medium = within 18%. Low = diverge >18%."),
    ]
    for offset, (col_name, desc) in enumerate(definitions, start=1):
        row = def_start + offset
        bg  = ROW_ODD if offset % 2 == 1 else ROW_EVEN
        nc  = ws.cell(row=row, column=1, value=col_name)
        nc.font      = Font(name="Calibri", size=10, bold=True, color=NAVY)
        nc.fill      = fill(bg)
        nc.alignment = al("left")
        nc.border    = border
        ws.merge_cells(f"B{row}:{last_col}{row}")
        dc = ws.cell(row=row, column=2, value=desc)
        dc.font      = Font(name="Calibri", size=10)
        dc.fill      = fill(bg)
        dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dc.border    = border
        ws.row_dimensions[row].height = 28

    # ── Results Tracker tab ──────────────────────────────────────
    tracker_df = update_results_tracker(games_list)
    _save_tracker_json(tracker_df)

    wt = wb.create_sheet("Results Tracker")

    TR_NAVY     = "1B3A6B"
    TR_WHITE    = "FFFFFF"
    TR_ODD      = "F4F6FA"
    TR_EVEN     = "FFFFFF"
    TR_GREEN_BG = "C6EFCE";  TR_GREEN_FG = "276221"
    TR_RED_BG   = "FFD6D6";  TR_RED_FG   = "8B1A1A"
    TR_PENDING  = "FFF3C4";  TR_PEND_FG  = "7A6000"

    tr_thin   = Side(style="thin", color="C8CDD8")
    tr_border = Border(left=tr_thin, right=tr_thin, top=tr_thin, bottom=tr_thin)

    # Compute running record by confidence tier
    resolved_rows = tracker_df[
        tracker_df["Actual Winner"].notna() &
        (tracker_df["Actual Winner"].str.strip() != "") &
        (tracker_df["Actual Winner"].str.lower() != "nan")
    ]

    def _record(subset):
        correct = sum(
            1 for _, r in subset.iterrows()
            if str(r.get("Predicted Winner", "")).strip() == str(r.get("Actual Winner", "")).strip()
        )
        total = len(subset)
        pct   = round(correct / total * 100) if total else 0
        return correct, total, pct

    overall_c, overall_t, overall_pct = _record(resolved_rows)
    high_c,    high_t,    high_pct    = _record(resolved_rows[resolved_rows["Confidence"] == "High"])
    med_c,     med_t,     med_pct     = _record(resolved_rows[resolved_rows["Confidence"] == "Medium"])
    low_c,     low_t,     low_pct     = _record(resolved_rows[resolved_rows["Confidence"] == "Low"])

    # Title row
    wt.merge_cells("A1:I1")
    ttl = wt["A1"]
    ttl.value     = f"⚾   MLB Results Tracker   —   Updated {today_str}"
    ttl.font      = Font(color=TR_WHITE, bold=True, size=14, name="Calibri")
    ttl.fill      = PatternFill("solid", fgColor=TR_NAVY)
    ttl.alignment = Alignment(horizontal="center", vertical="center")
    wt.row_dimensions[1].height = 30

    # Record summary row
    wt.merge_cells("A2:I2")
    rec_note = wt["A2"]
    if overall_t > 0:
        rec_note.value = (
            f"Overall: {overall_c}–{overall_t - overall_c} ({overall_pct}%)     "
            f"High confidence: {high_c}–{high_t - high_c} ({high_pct}%)     "
            f"Medium confidence: {med_c}–{med_t - med_c} ({med_pct}%)     "
            f"Low confidence: {low_c}–{low_t - low_c} ({low_pct}%)"
        )
    else:
        rec_note.value = "No resolved games yet — record will appear here once results are filled in."
    rec_note.font      = Font(size=10, bold=True, name="Calibri", color=TR_NAVY)
    rec_note.alignment = Alignment(horizontal="left", vertical="center")
    rec_note.fill      = PatternFill("solid", fgColor="D6E4F0")
    wt.row_dimensions[2].height = 22

    # Instruction row
    wt.merge_cells("A3:I3")
    note = wt["A3"]
    note.value     = 'Fill in "Actual Winner" with the winning team name — the "Correct?" column updates automatically.'
    note.font      = Font(size=9, italic=True, name="Calibri", color="444444")
    note.alignment = Alignment(horizontal="left", vertical="center")
    note.fill      = PatternFill("solid", fgColor="EEF1F8")
    wt.row_dimensions[3].height = 16

    # Column headers
    tr_cols = [
        ("Date",             11),
        ("Game Time",        13),
        ("Home Team",        22),
        ("Away Team",        22),
        ("Predicted Winner", 22),
        ("Home Win %",       12),
        ("Model Edge",       12),
        ("Confidence",       13),
        ("Actual Winner",    22),
        ("Correct?",         13),
    ]
    for col_idx, (col_name, col_width) in enumerate(tr_cols, start=1):
        cell = wt.cell(row=4, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, size=10, name="Calibri", color=TR_WHITE)
        cell.fill      = PatternFill("solid", fgColor=TR_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = tr_border
        wt.column_dimensions[get_column_letter(col_idx)].width = col_width
    wt.row_dimensions[4].height = 22

    # Data rows
    for row_idx, (_, tr_row) in enumerate(tracker_df.iterrows(), start=5):
        bg        = TR_ODD if row_idx % 2 == 0 else TR_EVEN
        predicted = str(tr_row.get("Predicted Winner", "")).strip()
        actual    = str(tr_row.get("Actual Winner", "")).strip()
        edge_val  = tr_row.get("Model Edge", "")

        if actual and actual.lower() not in ("", "nan"):
            is_correct     = actual == predicted
            correct_val    = "✓  Correct" if is_correct else "✗  Wrong"
            correct_fill   = TR_GREEN_BG if is_correct else TR_RED_BG
            correct_font_c = TR_GREEN_FG if is_correct else TR_RED_FG
        else:
            correct_val    = "Pending"
            correct_fill   = TR_PENDING
            correct_font_c = TR_PEND_FG

        try:
            edge_num = float(edge_val)
            edge_str = f"+{edge_num}%" if edge_num >= 0 else f"{edge_num}%"
        except (ValueError, TypeError):
            edge_str = str(edge_val) if edge_val else ""

        row_vals = [
            tr_row.get("Date", ""),
            tr_row.get("Game Time", ""),
            tr_row.get("Home Team", ""),
            tr_row.get("Away Team", ""),
            predicted,
            tr_row.get("Home Win %", ""),
            edge_str,
            tr_row.get("Confidence", ""),
            actual if actual and actual.lower() != "nan" else "",
            correct_val,
        ]

        LEFT_COLS = {3, 4, 5, 9}
        for col_idx, val in enumerate(row_vals, start=1):
            cell = wt.cell(row=row_idx, column=col_idx, value=sanitize_cell(val))
            cell.font      = Font(size=10, name="Calibri")
            cell.alignment = Alignment(
                horizontal="left" if col_idx in LEFT_COLS else "center",
                vertical="center"
            )
            cell.border = tr_border

            if col_idx == 10:
                cell.fill = PatternFill("solid", fgColor=correct_fill)
                cell.font = Font(size=10, name="Calibri", bold=True, color=correct_font_c)
            elif col_idx == 8:
                conf = str(val)
                if conf == "High":
                    cell.fill = PatternFill("solid", fgColor=TR_GREEN_BG)
                    cell.font = Font(size=10, name="Calibri", bold=True, color=TR_GREEN_FG)
                elif conf == "Medium":
                    cell.fill = PatternFill("solid", fgColor=TR_PENDING)
                    cell.font = Font(size=10, name="Calibri", bold=True, color=TR_PEND_FG)
                elif conf == "Low":
                    cell.fill = PatternFill("solid", fgColor=TR_RED_BG)
                    cell.font = Font(size=10, name="Calibri", bold=True, color=TR_RED_FG)
                else:
                    cell.fill = PatternFill("solid", fgColor=bg)
            elif col_idx == 7:
                try:
                    ev = float(edge_val)
                    if ev >= 5:
                        cell.fill = PatternFill("solid", fgColor=TR_GREEN_BG)
                        cell.font = Font(size=10, name="Calibri", bold=True, color=TR_GREEN_FG)
                    elif ev <= -5:
                        cell.fill = PatternFill("solid", fgColor=TR_RED_BG)
                        cell.font = Font(size=10, name="Calibri", bold=True, color=TR_RED_FG)
                    else:
                        cell.fill = PatternFill("solid", fgColor=bg)
                except (ValueError, TypeError):
                    cell.fill = PatternFill("solid", fgColor=bg)
            else:
                cell.fill = PatternFill("solid", fgColor=bg)

        wt.row_dimensions[row_idx].height = 18

    wt.freeze_panes = "A5"

    wb.save(output_path)
    print(f"Daily MLB report saved to {output_path}")
    _save_predictions_json(games_list, lock_game=lock_game, dog_game=dog_game)
    return output_path


RUN_LOCK = os.path.join(SCRIPT_DIR, ".pipeline_running")
RUN_LOCK_STALE_SECS = 20 * 60   # a real run takes ~8 min; 20 is a safe ceiling


def _safe_read_tracker_csv():
    """
    Read the tracker CSV defensively.

    On 2026-08-12 four overlapping pipeline runs raced on this file; one
    truncated it mid-write while another read it, and pandas raised
    EmptyDataError("No columns to parse from file"), failing the whole run.
    Supabase is the source of truth, so an unreadable local CSV should
    degrade to an empty frame rather than kill the pipeline.
    """
    try:
        if not os.path.exists(TRACKER_CSV) or os.path.getsize(TRACKER_CSV) == 0:
            print("  [Tracker] CSV missing or empty — using empty frame (Supabase is source of truth)")
            return pd.DataFrame()
        return pd.read_csv(TRACKER_CSV, dtype=str)
    except Exception as e:
        print(f"  [Tracker] CSV unreadable ({e}) — using empty frame")
        return pd.DataFrame()



def _acquire_run_lock() -> bool:
    """
    Prevent concurrent pipeline runs.

    run_daily_predictions() takes ~8 minutes and writes several fixed-path
    files (results_tracker.csv, mlb_daily_report.xlsx, daily_predictions.json).
    HTTP clients time out well before it finishes and retry, so on 2026-08-12
    four overlapping runs raced on those files — one truncated
    results_tracker.csv while another was reading it, producing
    "No columns to parse from file" and a failed run.

    Uses O_EXCL create, which is atomic. Stale locks (from a crashed run)
    are cleared automatically.
    """
    try:
        if os.path.exists(RUN_LOCK):
            age = time.time() - os.path.getmtime(RUN_LOCK)
            if age < RUN_LOCK_STALE_SECS:
                print(f"  [Lock] Pipeline already running ({age:.0f}s ago) — skipping this invocation")
                return False
            print(f"  [Lock] Clearing stale lock ({age:.0f}s old)")
            os.remove(RUN_LOCK)
        fd = os.open(RUN_LOCK, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        os.write(fd, str(time.time()).encode())
        os.close(fd)
        return True
    except FileExistsError:
        print("  [Lock] Lost race to another run — skipping this invocation")
        return False
    except Exception as e:
        print(f"  [Lock] Could not acquire lock ({e}) — proceeding")
        return True


def _release_run_lock():
    try:
        if os.path.exists(RUN_LOCK):
            os.remove(RUN_LOCK)
    except Exception:
        pass


def run_daily_predictions():
    if not _acquire_run_lock():
        return None
    try:
        return _run_daily_predictions_inner()
    finally:
        _release_run_lock()


def _run_daily_predictions_inner():
    games_today = []

    # Feature 1 — auto-fill yesterday's results before anything else
    print("Checking yesterday's results...")
    auto_fill_results()

    schedule = fetch_todays_schedule()
    if not schedule:
        print("No schedule data — using placeholder games.")
        schedule = [
            {"home_name": "Yankees", "home_id": 147,
             "away_name": "Red Sox", "away_id": 111,
             "venue": "Yankee Stadium",
             "home_pitcher_id": None, "home_pitcher_name": "TBD",
             "away_pitcher_id": None, "away_pitcher_name": "TBD"},
        ]

    print(f"Running predictions for {len(schedule)} game(s)...\n")

    # Fetch once up-front (all cached for the run)
    savant_batting = fetch_savant_team_batting()   # Savant team batting
    all_odds       = fetch_vegas_odds()            # Feature 2 — live Vegas odds
    umpire_map     = fetch_umpire_data(today_et()) # Umpire assignments
    print(f"  [Umpire] {len(umpire_map)} umpire assignments loaded")

    for game in schedule:
        home_name = game["home_name"]
        away_name = game["away_name"]
        home_id   = game["home_id"]
        away_id   = game["away_id"]
        venue     = game["venue"]
        print(f"--- {away_name} at {home_name} ({venue}) ---")
        try:

            home_stats   = get_team_stats(home_id, home_name, savant_batting=savant_batting)
            away_stats   = get_team_stats(away_id, away_name, savant_batting=savant_batting)
            home_pitch   = get_pitcher_stats(game["home_pitcher_id"], game["home_pitcher_name"])
            away_pitch   = get_pitcher_stats(game["away_pitcher_id"], game["away_pitcher_name"])
            weather      = get_weather(venue, game_date_utc=game.get("game_date", ""))
            park_factor  = get_park_factor(venue)
            umpire       = get_umpire_factor("Default")
            travel_home  = get_travel(home_name)
            travel_away  = get_travel(away_name)

            # Feature 3 — live lineup strength (falls back to team avg if not yet posted)
            home_lineup = get_lineup_strength(home_id, home_name, home_stats["lineup_strength"])
            away_lineup = get_lineup_strength(away_id, away_name, away_stats["lineup_strength"])

            # Lineup confirmation — check if MLB API returned real batting orders
            home_lineup_confirmed = home_lineup != home_stats["lineup_strength"]
            away_lineup_confirmed = away_lineup != away_stats["lineup_strength"]
            lineup_confirmed = home_lineup_confirmed and away_lineup_confirmed

            # Feature 4 — real bullpen fatigue (replaces placeholder 0.35)
            bullpen_home = get_bullpen_fatigue(home_id, home_name)
            bullpen_away = get_bullpen_fatigue(away_id, away_name)

            # Bullpen usage last 3 days from new fetch
            home_bullpen_usage = fetch_bullpen_usage(home_id, days=3)
            away_bullpen_usage = fetch_bullpen_usage(away_id, days=3)

            # Feature 2 — enhanced Vegas odds (vig-removed consensus or win% fallback)
            vegas = get_vegas_line(home_name, away_name, all_odds,
                                   home_wpct=home_stats["wpct"],
                                   away_wpct=away_stats["wpct"])

            # Phase 4.1: Elo win probability
            elo_prob = elo_win_probability(home_name, away_name)
            home_elo = get_elo(home_name)
            away_elo = get_elo(away_name)

            wind_dir_out = 1 if weather["wind_start_dir"] == "out" else 0

            adv_inputs = (
                home_stats["wpct"],        away_stats["wpct"],
                home_stats["rpg"],         away_stats["rpg"],
                home_pitch["home_era"],    home_pitch["away_era"],
                away_pitch["home_era"],    away_pitch["away_era"],
                home_stats["bullpen_era"], away_stats["bullpen_era"],
                park_factor,
                umpire,
                travel_home, travel_away,
                bullpen_home, bullpen_away,
                True,
                weather["temp_start"],       weather["temp_mid"],
                weather["wind_start_speed"], weather["wind_start_dir"],
                weather["wind_mid_speed"],   weather["wind_mid_dir"],
                weather["rain_chance"],
                home_stats["hard_hit"],    away_stats["hard_hit"],
                home_stats["defense"],     away_stats["defense"],
                home_lineup,               away_lineup,
                home_stats["vs_hand_split"], away_stats["vs_hand_split"]
            )

            adv_prob = advanced_model(adv_inputs)

            ml_features = {
                "team_wpct":       home_stats["wpct"],
                "opp_wpct":        away_stats["wpct"],
                "team_rpg":        home_stats["rpg"],
                "opp_rpg":         away_stats["rpg"],
                "pitcher_era":     home_pitch["home_era"],
                "opp_pitcher_era": away_pitch["away_era"],
                "bullpen_era":     home_stats["bullpen_era"],
                "opp_bullpen_era": away_stats["bullpen_era"],
                "park_factor":     park_factor,
                "temp":            weather["temp_start"],
                "wind_speed":      weather["wind_start_speed"],
                "wind_dir_out":    wind_dir_out,
                "home":            1,
                "pitcher_recent_delta":     home_pitch.get("home_recent_delta", 0.0),
                "opp_pitcher_recent_delta": away_pitch.get("away_recent_delta", 0.0),
            }
            if ml_model and isinstance(ml_model, dict):
                feat_df  = pd.DataFrame([ml_features])
                feat_cols = ml_model.get("feature_cols")
                if feat_cols:
                    for c in feat_cols:
                        if c not in feat_df.columns:
                            feat_df[c] = 0.0
                    feat_df = feat_df[feat_cols]
                rf_prob  = ml_model["rf"].predict_proba(feat_df)[0, 1]
                xgb_prob = ml_model["xgb"].predict_proba(feat_df)[0, 1]
                lgbm_prob = get_lightgbm_prediction(feat_df.values[0].tolist())

                # Phase 2: 4-model ensemble with optimized weights
                weights  = ml_model.get("weights", [0.30, 0.35, 0.20, 0.15])
                lr_prob  = None
                try:
                    if ml_model.get("lr") and ml_model.get("lr_scaler"):
                        feat_sc  = ml_model["lr_scaler"].transform(feat_df)
                        lr_prob  = ml_model["lr"].predict_proba(feat_sc)[0, 1]
                except Exception:
                    pass

                if lgbm_prob is not None and lr_prob is not None:
                    ml_prob = round(
                        rf_prob   * weights[0] +
                        xgb_prob  * weights[1] +
                        lgbm_prob * weights[2] +
                        lr_prob   * weights[3], 4)
                    print(f"  [ML] RF={rf_prob:.3f} XGB={xgb_prob:.3f} LGBM={lgbm_prob:.3f} LR={lr_prob:.3f} → {ml_prob:.3f}")
                elif lgbm_prob is not None:
                    ml_prob = round((rf_prob * 0.35 + xgb_prob * 0.40 + lgbm_prob * 0.25), 4)
                else:
                    ml_prob = round((rf_prob + xgb_prob) / 2, 4)

                # Phase 2: Confidence interval from RF bootstrap
                try:
                    import numpy as np
                    rf_base = ml_model["rf"].base_model
                    tree_preds = np.array([t.predict_proba(feat_df)[0, 1]
                                           for t in rf_base.estimators_])
                    ci_low   = round(float(np.percentile(tree_preds, 10)) * 100, 1)
                    ci_high  = round(float(np.percentile(tree_preds, 90)) * 100, 1)
                    ci_width = ci_high - ci_low
                except Exception:
                    ci_low, ci_high, ci_width = None, None, None
            else:
                ml_prob = adv_prob
            ci_low, ci_high, ci_width = None, None, None

            vegas_prob    = vegas.get("home_prob_novig",
                               (1 / (1 + 10 ** (vegas["home_moneyline"] / 100)))
                               if "home_moneyline" in vegas else 0.5)
            vegas_is_live = vegas.get("is_live", False)

            # Platoon splits — L/R handedness adjustment
            platoon_adj = get_platoon_advantage(
                game.get("home_pitcher_id"), game.get("away_pitcher_id"),
                home_id, away_id
            )

            # Kalshi prediction market signal
            kalshi_all  = fetch_kalshi_odds()
            kalshi_prob = None
            for (h, a), prob in kalshi_all.items():
                if (home_name in h or h in home_name) and (away_name in a or a in away_name):
                    kalshi_prob = prob
                    break

            # Series context
            series_game = get_series_context(home_id, away_id, today_et())

            final_prob = final_win_probability(
                adv_prob, ml_prob, vegas_prob,
                vegas_is_live=vegas_is_live,
                platoon_adj=platoon_adj,
                kalshi_prob=kalshi_prob,
                series_game=series_game,
            )

            adv_pct   = round(adv_prob    * 100, 2)
            ml_pct    = round(ml_prob     * 100, 2)
            vegas_pct = round(vegas_prob  * 100, 2)

            spread = max(adv_pct, ml_pct, vegas_pct) - min(adv_pct, ml_pct, vegas_pct)
            if spread <= 8:
                confidence = "High"
            elif spread <= 18:
                confidence = "Medium"
            else:
                confidence = "Low"

            predicted_winner = home_name if final_prob >= 50 else away_name
            model_edge = round(final_prob - vegas_pct, 1)

            # ── Market intelligence signals ───────────────────────────
            opening_prob = _opening_lines_cache.get((home_name, away_name), {}).get("home_prob", vegas_prob)
            rlm_signal   = detect_reverse_line_movement(home_name, away_name, opening_prob, vegas_prob)
            steam_signal = detect_steam_move(opening_prob, vegas_prob, hours_since_open=16)
            trap_signal  = detect_trap_line(home_name, vegas_prob)
            public_data  = get_public_betting_data(home_name, away_name)

            sharp_signal = rlm_signal.get("detected") or steam_signal.get("detected")
            model_agrees_with_sharp = (
                sharp_signal and
                rlm_signal.get("sharp_side", "") in (predicted_winner, "")
            )

            # Phase 3: Sharp/square divergence from bookmaker tiers
            sharp_consensus  = vegas.get("sharp_consensus")
            square_consensus = vegas.get("square_consensus")
            book_divergence  = vegas.get("book_divergence")
            # Strong divergence = sharps and squares disagree significantly
            sharp_square_alert = (
                book_divergence is not None and
                abs(book_divergence) > 0.03
            )

            # ── Run differential model (Phase 2) ──────────────────────
            home_pitcher_dict = {
                'blended_era': home_pitch.get('home_era', 4.20),
                'season_era':  home_pitch.get('season_era', 4.20),
            }
            away_pitcher_dict = {
                'blended_era': away_pitch.get('away_era', 4.20),
                'season_era':  away_pitch.get('season_era', 4.20),
            }
            pred_margin, home_cover_model, away_cover_model = predict_run_differential(
                home_stats, away_stats,
                home_pitcher_dict, away_pitcher_dict,
                park_factor=park_factor,
            )

            # Run line Vegas data
            rl_live          = vegas.get("rl_is_live", False)
            home_rl_line     = vegas.get("home_rl_line", -1.5)
            away_rl_line     = vegas.get("away_rl_line", 1.5)
            home_rl_odds     = vegas.get("home_rl_odds", None)
            away_rl_odds     = vegas.get("away_rl_odds", None)
            home_cover_vegas = vegas.get("home_cover_prob", None)
            away_cover_vegas = vegas.get("away_cover_prob", None)

            # ATS edge: model cover prob vs Vegas implied cover prob
            home_ats_edge = get_ats_edge(home_cover_model, home_rl_odds) if home_rl_odds else None
            away_ats_edge = get_ats_edge(away_cover_model, away_rl_odds) if away_rl_odds else None
            rl_edge_score = get_rl_edge_score(home_ats_edge) if home_ats_edge is not None else None

            # Best spread pick: whichever side has the higher ATS edge
            if home_ats_edge is not None and away_ats_edge is not None:
                if home_ats_edge >= away_ats_edge:
                    spread_pick      = home_name
                    spread_pick_line = f"{home_rl_line:+.1f}"
                    spread_pick_odds = home_rl_odds
                    spread_cover_pct = round(home_cover_model * 100, 1)
                    spread_edge      = home_ats_edge
                else:
                    spread_pick      = away_name
                    spread_pick_line = f"{away_rl_line:+.1f}"
                    spread_pick_odds = away_rl_odds
                    spread_cover_pct = round(away_cover_model * 100, 1)
                    spread_edge      = away_ats_edge
            else:
                spread_pick      = home_name if home_cover_model > 0.5 else away_name
                spread_pick_line = "-1.5" if home_cover_model > 0.5 else "+1.5"
                spread_pick_odds = None
                spread_cover_pct = round(max(home_cover_model, away_cover_model) * 100, 1)
                spread_edge      = None

            # Parse game time to Eastern for display
            game_time_et = ""
            raw_gd = game.get("game_date", "")
            if raw_gd:
                try:
                    import zoneinfo
                    dt_utc = datetime.datetime.strptime(raw_gd, "%Y-%m-%dT%H:%M:%SZ").replace(
                        tzinfo=datetime.timezone.utc)
                    dt_et  = dt_utc.astimezone(zoneinfo.ZoneInfo("America/New_York"))
                    game_time_et = dt_et.strftime("%-I:%M %p ET")
                except Exception:
                    pass

            weather_str = (
                f"{weather['temp_start']}°F · "
                + ("dome" if weather['wind_start_speed'] == 0 and weather['wind_start_dir'] == 'neutral'
                   else f"{weather['wind_start_speed']}mph {weather['wind_start_dir'].upper()}")
                + (f" · {weather['rain_chance']}% rain" if weather['rain_chance'] > 5 else "")
            )

            games_today.append({
                # Display columns
                "Game Time":               game_time_et,
                "Home Team":               home_name,
                "Away Team":               away_name,
                "Venue":                   venue,
                "Weather":                 weather_str,
                "Home Starting Pitcher":   game["home_pitcher_name"],
                "Away Starting Pitcher":   game["away_pitcher_name"],
                "Home Win %":              home_stats["wpct"],
                "Away Win %":              away_stats["wpct"],
                # Phase 4.1: Elo ratings
                "Home Elo":                home_elo,
                "Away Elo":                away_elo,
                "Elo Win Prob":            round(elo_prob * 100, 1),
                # Phase 4.4: Pitcher momentum
                "Home Pitcher Momentum":   home_pitch.get("momentum_signal", "neutral"),
                "Away Pitcher Momentum":   away_pitch.get("momentum_signal", "neutral"),
                "Home Momentum Ratio":     home_pitch.get("momentum_ratio", 1.0),
                "Away Momentum Ratio":     away_pitch.get("momentum_ratio", 1.0),
                "Home Runs / Game":        home_stats["rpg"],
                "Away Runs / Game":        away_stats["rpg"],
                "Home Bullpen ERA":        home_stats["bullpen_era"],
                "Away Bullpen ERA":        away_stats["bullpen_era"],
                "Home SP ERA (at Home)":   home_pitch["home_era"],
                "Away SP ERA (on Road)":   away_pitch["away_era"],
                "Park Factor":             park_factor,
                "Adv Model %":             adv_pct,
                "ML Model %":              ml_pct,
                "Vegas Implied %":         vegas_pct,
                "Vegas Is Live":           vegas_is_live,
                "Kalshi Is Live":          kalshi_prob is not None,
                "Series Game":             series_game,
                "Lineup Confirmed":        lineup_confirmed,
                "Model Edge":              model_edge,
                # ── Spread / Run line ─────────────────────────────────
                "Spread Pick":             spread_pick,
                "Spread Line":             spread_pick_line,
                "Spread Odds":             spread_pick_odds,
                "Spread Cover %":          spread_cover_pct,
                "Spread Edge":             spread_edge,
                "Spread Edge Score":       rl_edge_score,
                "Run Line Live":           rl_live,
                "Predicted Margin":        round(pred_margin, 1),
                "Home Cover %":            round(home_cover_model * 100, 1),
                "Away Cover %":            round(away_cover_model * 100, 1),
                "Home RL Line":            home_rl_line,
                "Away RL Line":            away_rl_line,
                "Home RL Odds":            home_rl_odds,
                "Away RL Odds":            away_rl_odds,
                # ── Sportsbook intelligence ───────────────────────────
                "Sharp Money Signal":      sharp_signal,
                "Model Agrees Sharp":      model_agrees_with_sharp,
                "RLM Detected":            rlm_signal.get("detected", False),
                "RLM Sharp Side":          rlm_signal.get("sharp_side", None),
                "RLM Label":               rlm_signal.get("label", None),
                "RLM Strength":            rlm_signal.get("signal_strength", 0),
                "Steam Detected":          steam_signal.get("detected", False),
                "Steam Label":             steam_signal.get("label", None),
                "Steam Velocity":          steam_signal.get("velocity", None),
                "Trap Detected":           trap_signal.get("detected", False),
                "Trap Label":              trap_signal.get("label", None),
                "Public Home Pct":         public_data.get("home_pct", 50),
                "Line Movement":           round((vegas_prob - opening_prob) * 100, 1),
                # Phase 3: Bookmaker tier analysis
                "Sharp Consensus":         round(sharp_consensus * 100, 1) if sharp_consensus else None,
                "Square Consensus":        round(square_consensus * 100, 1) if square_consensus else None,
                "Book Divergence":         round(book_divergence * 100, 2) if book_divergence else None,
                "Sharp Square Alert":      sharp_square_alert,
                # ── Final prediction ──────────────────────────────────
                "Home Win Probability":    final_prob,
                "Predicted Winner":        predicted_winner,
                "Confidence":              confidence,
                "CI Low":                  ci_low  if 'ci_low'  in dir() else None,
                "CI High":                 ci_high if 'ci_high' in dir() else None,
                "CI Width":                ci_width if 'ci_width' in dir() else None,
                # ML training features
                "team_wpct":        home_stats["wpct"],
                "opp_wpct":         away_stats["wpct"],
                "team_rpg":         home_stats["rpg"],
                "opp_rpg":          away_stats["rpg"],
                "pitcher_era":      home_pitch["home_era"],
                "opp_pitcher_era":  away_pitch["away_era"],
                "bullpen_era":      home_stats["bullpen_era"],
                "opp_bullpen_era":  away_stats["bullpen_era"],
                "park_factor":      park_factor,
                "temp":             weather["temp_start"],
                "wind_speed":       weather["wind_start_speed"],
                "wind_dir_out":     wind_dir_out,
                "home":             1,
                "pitcher_recent_delta":     home_pitch.get("home_recent_delta", 0.0),
                "opp_pitcher_recent_delta": away_pitch.get("away_recent_delta", 0.0),
            })
            print(f"  => Home win probability: {final_prob}%  |  Predicted Winner: {predicted_winner}  [{confidence} confidence]{' | 🔒 LOCK CANDIDATE' if model_edge >= 10 else ''}\n")

        except Exception as game_err:
            print(f"  [ERROR] Failed to process {away_name} @ {home_name}: {game_err}")
            import traceback
            traceback.print_exc()
            # Still add a basic entry so the game appears in the output
            games_today.append({
                "Game Time":         game.get("game_time_et", ""),
                "Home Team":         home_name,
                "Away Team":         away_name,
                "Venue":             venue,
                "Weather":           "",
                "Home Starting Pitcher": game.get("home_pitcher_name", "TBD"),
                "Away Starting Pitcher": game.get("away_pitcher_name", "TBD"),
                "Home Win Probability":  50,
                "Predicted Winner":      home_name,
                "Confidence":            "Low",
                "HasFullPrediction":     False,
                "Model Edge":            0,
            })
            continue

    # ── Phase 1: QA Assertions — validate before writing ─────
    pipeline_errors = []
    games_with_snapshots = []

    for g in games_today:
        home = g.get("Home Team", "")
        prob = g.get("Home Win Probability", 50) or 50
        edge = g.get("Model Edge", 0) or 0

        # QA checks
        if not g.get("Predicted Winner"):
            pipeline_errors.append(f"{home}: missing Predicted Winner")
        if not (5 <= prob <= 95):
            pipeline_errors.append(f"{home}: extreme probability {prob}% — possible data issue")
        if abs(edge) > 40:
            pipeline_errors.append(f"{home}: edge {edge}% is unrealistic — check Vegas fallback")

        # ── Feature snapshot — store exact values used at prediction time ──
        # This prevents temporal leakage when retraining uses stored values
        # instead of recalculating season stats that have changed
        snapshot = {
            "home_sp_era":        g.get("Home SP ERA (at Home)"),
            "away_sp_era":        g.get("Away SP ERA (on Road)"),
            "home_bullpen_era":   g.get("Home Bullpen ERA"),
            "away_bullpen_era":   g.get("Away Bullpen ERA"),
            "home_wpct":          g.get("Home Win %"),
            "away_wpct":          g.get("Away Win %"),
            "home_rpg":           g.get("Home Runs / Game"),
            "away_rpg":           g.get("Away Runs / Game"),
            "park_factor":        g.get("Park Factor"),
            "temp":               g.get("Weather", "").split("°")[0].replace("°F","").strip() if g.get("Weather") else None,
            "wind_speed":         g.get("Wind Speed"),
            "adv_model_pct":      g.get("Adv Model %"),
            "ml_model_pct":       g.get("ML Model %"),
            "vegas_implied_pct":  g.get("Vegas Implied %"),
            "vegas_is_live":      g.get("Vegas Is Live"),
            "model_edge":         g.get("Model Edge"),
            "sharp_money":        g.get("Sharp Money Signal"),
            "rlm_detected":       g.get("RLM Detected"),
            "captured_at":        today_et(),
        }
        g["_feature_snapshot"] = snapshot
        games_with_snapshots.append(g)

    if pipeline_errors:
        print(f"  [QA] ⚠️ {len(pipeline_errors)} validation warning(s):")
        for e in pipeline_errors[:5]:
            print(f"       {e}")
        critical = [e for e in pipeline_errors if "missing Predicted Winner" in e]
        if len(critical) > 5:
            print(f"  [QA] ❌ Too many critical errors — aborting write")
            return None
    else:
        print(f"  [QA] ✅ All {len(games_today)} games passed validation")

    games_today = games_with_snapshots

    # Feature 6 — monthly auto-recalibration of the ML model
    retrain_model_if_needed()

    # ── Lock of Day + Dangerous Underdog ──────────────────────
    lock_game = select_lock_of_the_day(games_today)
    dog_game  = select_dangerous_underdog(games_today)

    # ── Calibration metrics ────────────────────────────────────
    cal_metrics = {}
    try:
        import psycopg2
        db_url = os.environ.get("DATABASE_URL", "")
        if db_url:
            cal_conn = psycopg2.connect(db_url, sslmode='require', connect_timeout=10)
            cal_metrics = compute_calibration_metrics(cal_conn)
            if "error" not in cal_metrics:
                save_calibration_metrics(cal_conn, cal_metrics)
            cal_conn.close()
    except Exception as e:
        print(f"  [Calibration] Skipped: {e}")

    # ── Phase 1: Pipeline health monitoring ───────────────────
    # Writes a row to pipeline_health so we know immediately if something breaks
    try:
        import psycopg2
        db_url = os.environ.get("DATABASE_URL", "")
        if db_url:
            conn = psycopg2.connect(db_url, sslmode='require', connect_timeout=10)
            cur  = conn.cursor()
            cur.execute("""
                INSERT INTO pipeline_health
                    (date, games_generated, error_count, errors_json,
                     odds_api_live, lineup_confirmed, brier_score, run_completed_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT (date) DO UPDATE SET
                    games_generated  = EXCLUDED.games_generated,
                    error_count      = EXCLUDED.error_count,
                    errors_json      = EXCLUDED.errors_json,
                    odds_api_live    = EXCLUDED.odds_api_live,
                    lineup_confirmed = EXCLUDED.lineup_confirmed,
                    brier_score      = EXCLUDED.brier_score,
                    run_completed_at = NOW()
            """, (
                today_et(),
                len(games_today),
                len(pipeline_errors),
                json.dumps(pipeline_errors[:10]),
                any(g.get("Vegas Is Live") for g in games_today),
                any(g.get("Lineup Confirmed") for g in games_today),
                cal_metrics.get("brier_score"),
            ))
            conn.commit()
            cur.close()
            conn.close()
            print(f"  [Health] ✅ Pipeline health recorded — {len(games_today)} games, {len(pipeline_errors)} warnings, Vegas={'live' if any(g.get('Vegas Is Live') for g in games_today) else 'fallback'}")
    except Exception as e:
        print(f"  [Health] ❌ Health write failed: {e}")

    report = save_daily_report(games_today, lock_game=lock_game, dog_game=dog_game)

    # Opening lines MUST be stored after save_daily_report, because
    # store_opening_lines UPDATEs existing results rows and save_daily_report
    # is what creates them. Running it earlier silently updated 0 rows on any
    # day where the 9 AM run inserted the slate fresh (2026-08-16, 08-19),
    # which left those games with no opening price and therefore no CLV.
    store_opening_lines(all_odds)

    return report


# ============================================================
# EMAIL REPORT
# ============================================================

def email_report(report_path):
    to_email   = "radkins0206199@gmail.com"
    from_email = "radkins0206199@gmail.com"
    app_password = os.environ.get("GMAIL_APP_PASSWORD", "").replace(" ", "")

    if not app_password:
        print("ERROR: GMAIL_APP_PASSWORD environment variable not set.")
        return

    today_str = datetime.datetime.now(ET).strftime("%B %d, %Y")
    msg = EmailMessage()
    msg["Subject"] = f"MLB Prediction Report — {today_str}"
    msg["From"]    = from_email
    msg["To"]      = to_email
    msg.set_content(
        f"Your MLB prediction report for {today_str} is attached.\n\n"
        "Probabilities reflect live team stats, probable pitcher ERA, and Vegas lines."
    )

    with open(report_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"mlb_report_{today_et()}.xlsx"
        )

    with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
        smtp.starttls()
        smtp.login(from_email, app_password)
        smtp.send_message(msg)

    print("Email sent successfully.")


# ============================================================
# RUN EVERYTHING
# ============================================================

if __name__ == "__main__":
    report_path = run_daily_predictions()
    email_report(report_path)
